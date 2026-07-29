"""SeaBird HR Analytics - Production FastAPI Backend v3.2.10 (merged)
Python 3.12 compatible - uses openpyxl for Excel processing

CORRECTED ARCHITECTURE:
- Master: Employee database (Bio, PR, WC/BC, Vendor, Name, Store, Location, GENDER, DOJ, Contact No, Designation, Status)
          NOTE: Master has NO Emp Code column and NO Shift/Department columns.
                'Bio' is a sparse legacy biometric id (~69/1168 rows only) - do not use it as a join key.
- ESSL: Raw biometric punches from machine. ESSL's own "EMP Code" column IS the PayCode/PR
        for SeaBird staff (verified) - so ESSL joins to Employee via pr_number, not emp_code.
        ESSL_All is plant-wide (thousands of employees) - only rows matching a known pr_number
        belong to SeaBird.
- Tata: Official daily attendance from Tata. Tata_-_June.xlsx has ONE SHEET PER DAY (30 sheets) -
        every sheet must be read, not just the first. Each row also carries its own Shift and
        Department/Division - these are per-day/per-employee source of truth and must be used
        instead of a static Employee.shift.
- Tata All: Monthly historical data across all sub-contractors (trends only). Also multi-row,
            single sheet - read with the same "read all sheets" helper for consistency.
- Dump: RECONCILED OUTPUT - the system generates this; HR's own Dump workbook (also 30 sheets,
        ESSL vs Tata side-by-side) is treated as an external audit reference, not an ingestion
        source, and is not re-uploaded.

Correct Shift Rules: A=6:30-15:00(8.5h), B=15:00-23:30(8.5h), G=8:30-17:00(8.5h), C=23:30-06:30(7h)
Grace: 5 min | OT Headcount = Extra Hours / 8 | WC=no OT, BC/FLD=OT eligible

v3.2 fixes (see inline "# FIX v3.2:" markers):
  1. ESSL cross-tab upload now preserves status markers (WO, A, L, P, SP, HD) from IN/OUT rows
     instead of discarding them via parse_time_br(). These are stored in raw_punches.essl_status
     and the new ESSLAttendance.status column.
  2. Reconciliation logic now checks essl_status before declaring "Missing ESSL" - WO days
     are correctly shown as "Week Off" instead of "Missing ESSL" + "Mismatch".
  3. run_reconciliation_month now pre-fetches ALL data in 5 bulk queries and processes entirely
     in memory, then bulk-inserts results. Eliminates the ~240K individual DB queries that
     caused the 30-day reconciliation to timeout.
  4. _reconcile_single_date also uses the same batching pattern for consistency.
  5. Includes the v3.1 has_code fix: IN/OUT rows that repeat the employee code in column A
     are no longer treated as new employee blocks, so punch times are actually captured.
  6. v3.2.2: ESSL "WO" no longer blindly overrides Tata punches - only treated as Week Off
     when Tata also has no punches for that day. If Tata has punches, Tata is trusted.

MERGED v3.2.8 changes (file1 + file2):
  - Fixed endpoint placement: Tata-only and Debug endpoints moved BEFORE if __name__ block
  - Preserved force_man_hrs parameter for Tata-only OT calculation
  - Preserved get_tata_only_employees() helper and Tata-only endpoints
  - Preserved enhanced download_dump_report with monthly multi-sheet + Summary
  - Preserved /api/v1/tata-only/check endpoint for per-employee verification
     when Tata also has no punches for that day. If Tata has punches, Tata is trusted.
"""

from fastapi import FastAPI, UploadFile, File, HTTPException, Depends, Query, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, date, timedelta, time
import json
import csv
import io as bio
import os
import re
import hashlib
import secrets
import uuid
import threading
from pathlib import Path

from sqlalchemy import create_engine, event, Column, Integer, String, Float, DateTime, Date, Boolean, Text, ForeignKey, func, desc, text
from sqlalchemy.orm import declarative_base
from sqlalchemy.orm import sessionmaker, Session, relationship, joinedload

from openpyxl import load_workbook

# ============================================================================
# DATABASE SETUP
# ============================================================================

DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)


# Neon / PostgreSQL connection
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./seabird_hr.db")

# SQLAlchemy needs "postgresql://" not "postgres://"
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

# SQLite-only args won't work with Postgres
if DATABASE_URL.startswith("sqlite"):
    engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
    
    @event.listens_for(engine, "connect")
    def set_sqlite_pragma(dbapi_conn, connection_record):
        cursor = dbapi_conn.cursor()
        cursor.execute("PRAGMA journal_mode=WAL")
        cursor.execute("PRAGMA synchronous=NORMAL")
        cursor.execute("PRAGMA cache_size=10000")
        # FIX: bounded lock wait instead of relying on default busy_timeout=0 behavior.
        cursor.execute("PRAGMA busy_timeout=15000")
        cursor.close()
else:
    # Neon / PostgreSQL: aggressive pool settings to prevent SSL drops
    # FIX: the reconciliation job was getting stuck forever at "running" with no
    # error - root cause is that a Postgres connection with a dropped/stale socket
    # (Neon idles/drops connections) can hang a query or commit() indefinitely with
    # no default timeout, and threading + in-memory job dict meant nothing ever
    # surfaced that hang as a failure. TCP keepalives let the OS notice a dead
    # connection instead of waiting forever on a socket that will never respond.
    #
    # NOTE: statement_timeout is NOT set via connect_args "options" here - Neon's
    # pooled (pgbouncer) endpoint rejects unsupported startup parameters and refuses
    # the connection outright (see neon.tech/docs/connect/connection-errors).
    # Instead it's set per-session below via a plain SET command after connect,
    # which pgbouncer passes through fine.
    _pg_connect_args = {
        "sslmode": "require",
        "keepalives": 1,
        "keepalives_idle": 30,
        "keepalives_interval": 10,
        "keepalives_count": 3,
        "connect_timeout": 10,
    }
    if "sslmode" in DATABASE_URL:
        _pg_connect_args.pop("sslmode")

    engine = create_engine(
        DATABASE_URL,
        pool_pre_ping=True,      # verify connection before use
        pool_recycle=300,        # recycle every 5 min (Neon drops idle >5min)
        pool_size=6,             # Neon free tier = 10 max concurrent; Dashboard fires up to 7 parallel requests per load
        max_overflow=2,          # allow 2 extra temporarily
        pool_timeout=30,         # wait up to 30s for a connection
        connect_args=_pg_connect_args
    )

    @event.listens_for(engine, "connect")
    def set_pg_statement_timeout(dbapi_conn, connection_record):
        cursor = dbapi_conn.cursor()
        cursor.execute("SET statement_timeout = 120000")  # 120s hard cap per statement
        cursor.close()

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

Base = declarative_base()
# ============================================================================
# DATABASE MODELS
# ============================================================================

class Vendor(Base):
    __tablename__ = "vendors"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(100), unique=True, index=True, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)

class Store(Base):
    __tablename__ = "stores"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(100), unique=True, index=True, nullable=False)
    location = Column(String(100))
    created_at = Column(DateTime, default=datetime.utcnow)

class Department(Base):
    __tablename__ = "departments"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(100), unique=True, index=True, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)

class Employee(Base):
    __tablename__ = "employees"
    id = Column(Integer, primary_key=True, index=True)
    pr_number = Column(String(20), unique=True, index=True, nullable=False)
    bio_id = Column(String(20), index=True)
    emp_code = Column(String(20), index=True)
    name = Column(String(100), nullable=False)
    vendor = Column(String(50), index=True)
    store = Column(String(50), index=True)
    department = Column(String(50), index=True)
    designation = Column(String(50))
    shift = Column(String(10), index=True)
    wc = Column(String(20))
    bc = Column(String(20))
    status = Column(String(20), default="ACTIVE")
    join_date = Column(Date)
    phone = Column(String(20))
    email = Column(String(100))
    upload_log_id = Column(Integer, index=True, nullable=True)  # which master upload created/last-updated this employee
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    essl_records = relationship("ESSLAttendance", back_populates="employee", cascade="all, delete-orphan")
    tata_records = relationship("TataAttendance", back_populates="employee", cascade="all, delete-orphan")
    ot_records = relationship("Overtime", back_populates="employee", cascade="all, delete-orphan")

class ESSLAttendance(Base):
    __tablename__ = "essl_attendance"
    id = Column(Integer, primary_key=True, index=True)
    employee_id = Column(Integer, ForeignKey("employees.id"), index=True)
    pr_number = Column(String(20), index=True)
    emp_code = Column(String(20))
    emp_name = Column(String(100))
    date = Column(Date, index=True)
    in_time = Column(String(10))
    out_time = Column(String(10))
    status = Column(String(10))  # FIX v3.2: explicit status column for WO/A/L/P/SP/HD
    raw_punches = Column(Text)
    vendor = Column(String(50))
    store = Column(String(50))
    shift = Column(String(10))
    upload_log_id = Column(Integer, index=True, nullable=True)  # which upload created/last-touched this row
    created_at = Column(DateTime, default=datetime.utcnow)
    employee = relationship("Employee", back_populates="essl_records")

class TataAttendance(Base):
    __tablename__ = "tata_attendance"
    id = Column(Integer, primary_key=True, index=True)
    employee_id = Column(Integer, ForeignKey("employees.id"), index=True)
    pr_number = Column(String(20), index=True)
    emp_code = Column(String(20))
    emp_name = Column(String(100))
    date = Column(Date, index=True)
    in_time = Column(String(10))
    out_time = Column(String(10))
    man_hrs = Column(Float, default=0)
    status = Column(String(20))
    ot_hours = Column(Float, default=0)
    approved_ot = Column(Float, default=0)
    early_going = Column(String(5), default="No")
    shift_late = Column(String(5), default="No")
    vendor = Column(String(50))
    store = Column(String(50))
    department = Column(String(50))
    shift = Column(String(10))
    upload_log_id = Column(Integer, index=True, nullable=True)  # which upload created/last-touched this row
    created_at = Column(DateTime, default=datetime.utcnow)
    employee = relationship("Employee", back_populates="tata_records")

class Attendance(Base):
    __tablename__ = "attendance"
    id = Column(Integer, primary_key=True, index=True)
    employee_id = Column(Integer, ForeignKey("employees.id"), index=True)
    pr_number = Column(String(20), index=True)
    emp_code = Column(String(20))
    date = Column(Date, index=True)
    essl_in = Column(String(10))
    essl_out = Column(String(10))
    tata_in = Column(String(10))
    tata_out = Column(String(10))
    final_in = Column(String(10))
    final_out = Column(String(10))
    worked_hours = Column(Float, default=0)
    man_hrs = Column(Float, default=0)
    ot_hours = Column(Float, default=0)
    ot_headcount = Column(Float, default=0)
    attendance_status = Column(String(30), index=True)   # INDEX: daily register filter
    late_minutes = Column(Integer, default=0)
    early_minutes = Column(Integer, default=0)
    single_punch = Column(String(5), default="No")
    is_match = Column(String(5), default="No")
    match_status = Column(String(30))
    vendor = Column(String(50), index=True)              # INDEX: vendor breakdown
    store = Column(String(50), index=True)               # INDEX: store breakdown
    department = Column(String(50), index=True)          # INDEX: dept breakdown
    shift = Column(String(10), index=True)               # INDEX: shift breakdown
    category = Column(String(10))
    remark = Column(String(100))
    issue = Column(String(50), default="-", index=True)  # INDEX: issue filter
    source = Column(String(20), default="reconciliation")
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    employee = relationship("Employee")

class MonthlyTataAttendance(Base):
    __tablename__ = "monthly_tata_attendance"
    id = Column(Integer, primary_key=True, index=True)
    pr_number = Column(String(20), index=True)
    emp_name = Column(String(100))
    year_month = Column(String(7), index=True)
    total_days = Column(Integer, default=0)
    present_days = Column(Integer, default=0)
    absent_days = Column(Integer, default=0)
    leave_days = Column(Integer, default=0)
    weekoff_days = Column(Integer, default=0)
    half_days = Column(Integer, default=0)
    attendance_percentage = Column(Float, default=0)
    total_ot_hours = Column(Float, default=0)
    created_at = Column(DateTime, default=datetime.utcnow)

class AttendanceReconciliation(Base):
    __tablename__ = "attendance_reconciliation"
    id = Column(Integer, primary_key=True, index=True)
    pr_number = Column(String(20), index=True)
    emp_name = Column(String(100))
    date = Column(Date, index=True)
    essl_in = Column(String(10))
    essl_out = Column(String(10))
    tata_in = Column(String(10))
    tata_out = Column(String(10))
    in_delta_minutes = Column(Integer, default=0)
    out_delta_minutes = Column(Integer, default=0)
    match_status = Column(String(30), index=True)
    severity = Column(String(20), index=True)
    remarks = Column(Text)
    created_at = Column(DateTime, default=datetime.utcnow)

class HRAction(Base):
    __tablename__ = "hr_actions"
    id = Column(Integer, primary_key=True, index=True)
    pr_number = Column(String(20), index=True)
    emp_name = Column(String(100))
    date = Column(Date, index=True)
    action_type = Column(String(50), index=True)
    description = Column(Text)
    priority = Column(String(20), index=True)
    status = Column(String(20), default="Open")
    assigned_to = Column(String(100))
    resolution = Column(Text)
    resolved_at = Column(DateTime)
    created_at = Column(DateTime, default=datetime.utcnow)

class UploadLog(Base):
    __tablename__ = "upload_logs"
    id = Column(Integer, primary_key=True, index=True)
    file_type = Column(String(20), index=True)
    filename = Column(String(255))
    file_hash = Column(String(64))
    rows_processed = Column(Integer, default=0)
    status = Column(String(20), default="Success")
    error_message = Column(Text)
    uploaded_at = Column(DateTime, default=datetime.utcnow)

def safe_log_upload(db: Session, file_type: str, filename: str, file_hash: str,
                     rows_processed: int, status: str = "Success", error_message: Optional[str] = None) -> Dict[str, Any]:
    """Write an UploadLog row without ever letting a logging failure turn into a
    silent, unlogged 500. If this insert/commit itself fails, we catch it, roll
    back just this log write (not the caller's already-committed data), and
    return the real error so the endpoint can put it in the HTTP response body
    instead of it vanishing into server logs no one is watching."""
    try:
        db.add(UploadLog(file_type=file_type, filename=filename, file_hash=file_hash,
                          rows_processed=rows_processed, status=status, error_message=error_message))
        db.commit()
        return {"logged": True}
    except Exception as log_err:
        db.rollback()
        import traceback
        tb = traceback.format_exc()
        print(f"[UPLOAD LOG FAILED] file_type={file_type} filename={filename} error={log_err}")
        print(tb)
        return {"logged": False, "log_error": str(log_err), "log_traceback": tb}

class Overtime(Base):
    __tablename__ = "overtime"
    id = Column(Integer, primary_key=True, index=True)
    employee_id = Column(Integer, ForeignKey("employees.id"))
    pr_number = Column(String(20), index=True)
    name = Column(String(100))
    store = Column(String(50))
    date = Column(Date, index=True)
    worked_hours = Column(Float, default=0)
    ot_hours = Column(Float, default=0)
    calc_headcount = Column(Float, default=0)
    status = Column(String(20), default="Pending")
    approved_by = Column(String(100))
    approved_at = Column(DateTime)
    created_at = Column(DateTime, default=datetime.utcnow)
    employee = relationship("Employee", back_populates="ot_records")

class LatePunchPenalty(Base):
    __tablename__ = "late_punch_penalties"
    id = Column(Integer, primary_key=True, index=True)
    pr_number = Column(String(20), index=True)
    month = Column(String(7), index=True)
    late_count = Column(Integer, default=0)
    penalty_days = Column(Float, default=0)
    calculated_at = Column(DateTime, default=datetime.utcnow)

class BehavioralAlert(Base):
    """Non-penalty behavioral alerts: continuous (consecutive working-day) streaks of
    Early Departure or Half Day. No penalty attached - just a flag for HR to review."""
    __tablename__ = "behavioral_alerts"
    id = Column(Integer, primary_key=True, index=True)
    pr_number = Column(String(20), index=True)
    month = Column(String(7), index=True)
    alert_type = Column(String(30), index=True)  # "Early Departure" or "Half Day"
    streak_length = Column(Integer, default=0)
    streak_start = Column(Date)
    streak_end = Column(Date)
    threshold_used = Column(Integer, default=3)
    calculated_at = Column(DateTime, default=datetime.utcnow)


class User(Base):
    """Minimal auth users table. In production, switch to bcrypt + JWT."""
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String(50), unique=True, index=True, nullable=False)
    name = Column(String(100), nullable=False)
    hashed_password = Column(String(255), nullable=False)
    role = Column(String(20), default="user")
    is_active = Column(Boolean, default=True)
    api_token = Column(String(255), unique=True, nullable=True, index=True)
    created_at = Column(DateTime, default=datetime.utcnow)

Base.metadata.create_all(bind=engine)

def _run_lightweight_migrations():
    is_sqlite = DATABASE_URL.startswith("sqlite")
    with engine.begin() as conn:
        for table in Base.metadata.sorted_tables:
            if is_sqlite:
                existing_cols = {row[1] for row in conn.exec_driver_sql(f'PRAGMA table_info("{table.name}")')}
            else:
                existing_cols = {
                    row[0] for row in conn.exec_driver_sql(
                        "SELECT column_name FROM information_schema.columns WHERE table_name = %(t)s",
                        {"t": table.name}
                    )
                }
            if not existing_cols:
                continue
            for column in table.columns:
                if column.name not in existing_cols:
                    col_type = column.type.compile(engine.dialect)
                    conn.exec_driver_sql(f'ALTER TABLE "{table.name}" ADD COLUMN "{column.name}" {col_type}')
                    print(f"[MIGRATE] Added column {table.name}.{column.name}")

_run_lightweight_migrations()


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ============================================================================
# AUTH HELPERS (zero-dependency: uses built-in hashlib + secrets)
# ============================================================================

def _hash_password(password: str) -> str:
    """PBKDF2 hash — replace with passlib/bcrypt for production."""
    salt = "seabird-hr-static-salt-v1"
    return hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 100000).hex()


def _verify_password(plain: str, hashed: str) -> bool:
    return secrets.compare_digest(_hash_password(plain), hashed)


def _generate_token() -> str:
    return secrets.token_urlsafe(32)


def _seed_default_admin(db: Session):
    """Create a default admin user if the users table is empty."""
    if db.query(User).first() is None:
        admin = User(
            username="admin",
            name="HR Manager",
            hashed_password=_hash_password("admin123"),
            role="admin",
            api_token=_generate_token(),
        )
        db.add(admin)
        db.commit()
        print("[AUTH] Default admin created: username=admin, password=admin123")



# Seed default admin user on first boot
with SessionLocal() as _seed_db:
    _seed_default_admin(_seed_db)
from fastapi import Header
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials

security = HTTPBearer(auto_error=False)


def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db),
) -> User:
    """Dependency to protect routes. Pass token in Authorization: Bearer <token> header."""
    if not credentials:
        raise HTTPException(status_code=401, detail="Missing Authorization header")
    token = credentials.credentials
    user = db.query(User).filter(User.api_token == token, User.is_active == True).first()
    if not user:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    return user


app = FastAPI(
    title="SeaBird HR Analytics API",
    description="HR Analytics backend for SeaBird x Tata Motors workforce",
    version="3.2.8-MERGED"
)

# CORS - DYNAMIC: reflect any origin (internal tool - no hardcoded URLs needed)
# This eliminates all CORS issues regardless of which Vercel/Railway URL is used
from starlette.responses import Response

@app.middleware("http")
async def dynamic_cors_middleware(request, call_next):
    """Handle CORS for ALL origins dynamically. Internal tool - no URL whitelist needed."""
    origin = request.headers.get("origin", "*")

    # Handle preflight (OPTIONS) immediately
    if request.method == "OPTIONS":
        return Response(
            status_code=200,
            headers={
                "Access-Control-Allow-Origin": origin,
                "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, PATCH, OPTIONS",
                "Access-Control-Allow-Headers": "Authorization, Content-Type, Accept, X-Requested-With, Origin",
                "Access-Control-Allow-Credentials": "true",
                "Access-Control-Max-Age": "86400",
            },
        )

    # Handle actual request - add CORS headers even if the endpoint crashes
    try:
        response = await call_next(request)
    except Exception as exc:
        # FIX: this used to swallow the real exception and return a blank
        # "Internal Server Error", which is why every crash just showed up in
        # the browser as "HTTP 500" with no explanation - neither the user nor
        # anyone debugging it could tell what actually broke. Now we print the
        # full traceback to server logs (visible in Railway logs) AND include
        # the real error message + type in the JSON body, so the frontend's
        # error alert can show something actionable.
        import traceback
        tb = traceback.format_exc()
        print(f"[UNHANDLED EXCEPTION] {request.method} {request.url.path}")
        print(tb)
        import json as _json
        response = Response(
            status_code=500,
            content=_json.dumps({
                "detail": f"{type(exc).__name__}: {str(exc)}",
                "path": str(request.url.path),
            }),
            media_type="application/json"
        )

    response.headers["Access-Control-Allow-Origin"] = origin
    response.headers["Access-Control-Allow-Credentials"] = "true"
    response.headers["Access-Control-Expose-Headers"] = "*"
    return response

# ============================================================================
# HEALTH / CORS TEST
# ============================================================================

@app.get("/api/v1/health")
def health_check():
    """Returns ok. Use this to verify backend is alive and CORS headers are present."""
    return {"status": "ok", "version": "3.2.10-OT", "timestamp": datetime.utcnow().isoformat()}



# ============================================================================
# BUSINESS RULES ENGINE
# ============================================================================

SHIFT_RULES = {
    "A": {"name": "A Shift", "start": time(6, 30), "end": time(15, 0), "standard_hours": 8.5, "early_departure": 7.5, "half_day_threshold": 4.5, "grace_minutes": 5},
    "B": {"name": "B Shift", "start": time(15, 0), "end": time(23, 30), "standard_hours": 8.5, "early_departure": 7.5, "half_day_threshold": 4.5, "grace_minutes": 5},
    "G": {"name": "General Shift", "start": time(8, 30), "end": time(17, 0), "standard_hours": 8.5, "early_departure": 7.5, "half_day_threshold": 4.5, "grace_minutes": 5},
    "GENERAL": {"name": "General Shift", "start": time(8, 30), "end": time(17, 0), "standard_hours": 8.5, "early_departure": 7.5, "half_day_threshold": 4.5, "grace_minutes": 5},
    "C": {"name": "Night Shift", "start": time(23, 30), "end": time(6, 30), "standard_hours": 7.0, "early_departure": 6.5, "half_day_threshold": 3.5, "grace_minutes": 5}
}

CATEGORY_RULES = {
    "WC": {"name": "White Collar", "eligible_for_ot": False, "description": "Office Staff"},
    "BC": {"name": "Blue Collar", "eligible_for_ot": True, "description": "Factory Worker"},
    "FLD": {"name": "Fork Lift Driver", "eligible_for_ot": True, "description": "Fork Lift Operator"}
}

LATE_PUNCH_RULES = {"grace_minutes": 5, "half_day_after": 3, "one_day_after": 6}

# No penalty attached to early-departure / half-day streaks, but HR wants a heads-up
# when the same behaviour repeats on consecutive working days. Adjustable without a
# code change via the `consecutive_days` query param on the calculate endpoint below;
# this is just the default.
BEHAVIORAL_ALERT_CONFIG = {"default_consecutive_days": 3}

# OT Threshold Alerts
OT_WEEKLY_THRESHOLD = 12.0   # hours per rolling 7-day window
OT_MONTHLY_THRESHOLD = 48.0  # hours per calendar month

ACTIVE_STATUS = {"ACTIVE", "OK", "ON LEAVE", "TRANSFER"}
INACTIVE_STATUS = {"LEFT", "NOT OK", "RESIGNED", "TERMINATED"}

# FIX v3.2: Valid ESSL status markers that are NOT time values
ESSL_STATUS_MARKERS = {"WO", "A", "L", "P", "SP", "HD"}

# FIX v3.2.2: Valid shift codes - used to avoid comparing shift codes against raw
# ESSL time-range strings (e.g. "15:00To23:30") when computing "Alternate Shift"
VALID_SHIFT_CODES = {"A", "B", "G", "C"}

def normalize_shift(shift: Optional[str]) -> str:
    if shift is None:
        return "G"
    shift = str(shift).strip().upper()
    shift = re.sub(r'^SHIFT\s+', '', shift).strip()
    if shift == "GENERAL" or shift == "":
        return "G"
    return shift

def safe_shift(shift: Optional[str]) -> str:
    """Normalize shift and guarantee it fits in VARCHAR(10)."""
    s = normalize_shift(shift)
    return s[:10]

def safe_time(time_val: Optional[str]) -> Optional[str]:
    """Truncate time string to fit VARCHAR(10)."""
    if not time_val:
        return None
    return time_val[:10]

def parse_time_br(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, float):
        hours = int(value * 24)
        minutes = int((value * 24 - hours) * 60)
        return f"{hours:02d}:{minutes:02d}"
    time_str = str(value).strip()
    if time_str == "":
        return None
    # FIX v3.2: Also treat ESSL status markers as non-time values
    if time_str.upper() in ["#N/A", "N/A", "NA", "-", "", " "] or time_str.upper() in ESSL_STATUS_MARKERS:
        return None
    for fmt in ["%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p"]:
        try:
            return datetime.strptime(time_str, fmt).strftime("%H:%M")
        except ValueError:
            continue
    # FIX: Don't return raw strings that look like time ranges (e.g. "15:00To23:30")
    if ":" in time_str and len(time_str) <= 10:
        return time_str[:10]
    return None

def parse_time_obj(value) -> Optional[time]:
    if value is None:
        return None
    if isinstance(value, time):
        return value
    value = str(value).strip()
    if value == "" or value.upper() in ["#N/A", "N/A", "NA", "-"]:
        return None
    for fmt in ["%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p"]:
        try:
            return datetime.strptime(value, fmt).time()
        except ValueError:
            continue
    return None

def parse_date_br(date_val) -> Optional[date]:
    if date_val is None:
        return None
    if isinstance(date_val, datetime):
        return date_val.date()
    if isinstance(date_val, date):
        return date_val
    date_str = str(date_val).strip()
    dt_match = re.match(r'(\d{1,2}[/-]\d{1,2}[/-]\d{4})\s+\d{1,2}:\d{2}\s*(?:AM|PM)?', date_str, re.IGNORECASE)
    if dt_match:
        date_str = dt_match.group(1)
    iso_dt_match = re.match(r'(\d{4}-\d{2}-\d{2})[\sT]\d{2}:\d{2}', date_str)
    if iso_dt_match:
        date_str = iso_dt_match.group(1)
    for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"]:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None

def parse_year_month(month_str: Optional[str]) -> tuple[int, int]:
    """Safely parse a 'YYYY-MM' string. Tolerates extra date parts and raises a clean 400 instead of crashing."""
    if not month_str:
        raise HTTPException(status_code=400, detail="Month parameter is required. Expected YYYY-MM.")
    try:
        parts = month_str.strip().split("-")
        year, mon = int(parts[0]), int(parts[1])
        if not (1 <= mon <= 12):
            raise ValueError("month out of range")
        return year, mon
    except (ValueError, IndexError):
        raise HTTPException(status_code=400, detail=f"Invalid month format '{month_str}'. Expected YYYY-MM.")

def combine_today(t: time) -> datetime:
    return datetime.combine(datetime.today(), t)

def time_to_minutes(time_str: str) -> int:
    if not time_str or ":" not in time_str:
        return 0
    try:
        parts = time_str.split(":")
        return int(parts[0]) * 60 + int(parts[1])
    except:
        return 0

def calculate_work_duration(in_time, out_time) -> float:
    t_in = parse_time_obj(in_time)
    t_out = parse_time_obj(out_time)
    if t_in is None or t_out is None:
        return 0.0
    dt_in = combine_today(t_in)
    dt_out = combine_today(t_out)
    if dt_out < dt_in:
        dt_out += timedelta(days=1)
    duration = dt_out - dt_in
    return round(duration.total_seconds() / 3600, 2)

def get_shift_rule(shift: Optional[str]) -> Dict[str, Any]:
    shift = normalize_shift(shift)
    return SHIFT_RULES.get(shift, SHIFT_RULES["G"])

def get_standard_hours(shift: Optional[str]) -> float:
    return get_shift_rule(shift)["standard_hours"]

def get_early_departure_limit(shift: Optional[str]) -> float:
    return get_shift_rule(shift)["early_departure"]

def get_half_day_threshold(shift: Optional[str]) -> float:
    """Return the worked-hours threshold below which a day is counted as Half Day.
    Default: 4.5 hrs for A/B/G shifts, 3.5 hrs for C shift."""
    return get_shift_rule(shift).get("half_day_threshold", 4.5)

def get_grace_minutes(shift: Optional[str]) -> int:
    return get_shift_rule(shift)["grace_minutes"]

def get_shift_start_time(shift: Optional[str]) -> time:
    return get_shift_rule(shift)["start"]

def get_shift_end_time(shift: Optional[str]) -> time:
    return get_shift_rule(shift)["end"]

def normalize_category(category: Optional[str]) -> str:
    if category is None:
        return "UNKNOWN"
    category = str(category).strip().upper()
    mapping = {
        "WHITE COLLAR": "WC", "WHITECOLLAR": "WC", "WC": "WC",
        "BLUE COLLAR": "BC", "BLUECOLLAR": "BC", "BC": "BC",
        "FORK LIFT DRIVER": "FLD", "FORKLIFT DRIVER": "FLD",
        "FORK LIFT": "FLD", "FORKLIFT": "FLD", "FLD": "FLD"
    }
    return mapping.get(category, category)

def is_ot_eligible_category(category: Optional[str]) -> bool:
    category = normalize_category(category)
    if category not in CATEGORY_RULES:
        return True
    return CATEGORY_RULES[category]["eligible_for_ot"]

def calculate_ot_headcount(extra_hours: float) -> float:
    if extra_hours <= 0:
        return 0.0
    return round(extra_hours / 8.0, 2)

def calculate_ot(shift: Optional[str], category: Optional[str], in_time, out_time, man_hrs: Optional[float] = None, force_man_hrs: bool = False) -> Dict[str, Any]:
    shift = normalize_shift(shift)
    standard_hours = get_standard_hours(shift)
    category = normalize_category(category)
    ot_allowed = is_ot_eligible_category(category)

    # TATA-ONLY: If force_man_hrs is True, always use man_hrs for worked_hours (Tata priority)
    if force_man_hrs and man_hrs is not None:
        try:
            worked_hours = float(man_hrs)
        except:
            worked_hours = 0.0
    else:
        worked_hours = calculate_work_duration(in_time, out_time)
        if worked_hours <= 0 and man_hrs is not None:
            try:
                worked_hours = float(man_hrs)
            except:
                worked_hours = 0.0

    extra_hours = max(0, worked_hours - standard_hours)
    extra_hours = round(extra_hours, 2)

    if not ot_allowed:
        return {
            "worked_hours": worked_hours, "standard_hours": standard_hours, "extra_hours": extra_hours,
            "calculated_ot_hours": 0.0, "ot_headcount": 0.0, "is_ot_eligible": False, "blocked_reason": "White Collar"
        }

    ot_headcount = calculate_ot_headcount(extra_hours)
    return {
        "worked_hours": worked_hours, "standard_hours": standard_hours, "extra_hours": extra_hours,
        "calculated_ot_hours": extra_hours, "ot_headcount": ot_headcount,
        "is_ot_eligible": extra_hours > 0, "blocked_reason": None
    }

def is_late_punch(shift: Optional[str], in_time) -> bool:
    t_in = parse_time_obj(in_time)
    if t_in is None:
        return False
    rule = get_shift_rule(shift)
    cutoff = combine_today(rule["start"]) + timedelta(minutes=rule["grace_minutes"])
    actual = combine_today(t_in)
    return actual > cutoff

def get_late_minutes(shift: Optional[str], in_time) -> int:
    t_in = parse_time_obj(in_time)
    if t_in is None:
        return 0
    rule = get_shift_rule(shift)
    cutoff = combine_today(rule["start"]) + timedelta(minutes=rule["grace_minutes"])
    actual = combine_today(t_in)
    if actual <= cutoff:
        return 0
    diff = actual - cutoff
    return int(diff.total_seconds() // 60)

def is_early_departure(shift: Optional[str], worked_hours: float) -> bool:
    try:
        worked_hours = float(worked_hours)
    except:
        worked_hours = 0.0
    return worked_hours < get_early_departure_limit(shift)

def determine_attendance_status(status: Optional[str] = None, in_time=None, out_time=None, man_hrs: Optional[float] = None) -> Dict[str, Any]:
    result = {"status": "A", "is_present": False, "is_absent": True, "is_single_punch": False}
    if status is not None:
        status = str(status).strip().upper()

    # FIX v3.2.1: Handle both short codes (P, SP, A) and full words (Present, Absent, etc.)
    present_keywords = {"P", "PRESENT", "PR", "PRESENT DAY"}
    single_punch_keywords = {"SP", "SINGLE PUNCH", "SINGLEPUNCH"}
    absent_keywords = {"A", "ABSENT", "AB"}
    leave_keywords = {"L", "LEAVE", "LV"}
    weekoff_keywords = {"WO", "WEEK OFF", "WEEKOFF", "W/O"}
    half_day_keywords = {"HD", "HALF DAY", "HALFDAY", "H/D"}

    if status in present_keywords:
        result["status"] = "P"; result["is_present"] = True; result["is_absent"] = False; return result
    if status in single_punch_keywords:
        result["status"] = "SP"; result["is_present"] = True; result["is_absent"] = False; result["is_single_punch"] = True; return result
    if status in absent_keywords:
        return result
    if status in leave_keywords:
        result["status"] = "L"; result["is_present"] = False; result["is_absent"] = False; return result
    # FIX v3.2.7: WO returns "WO" as distinct status (still counts as absent for dashboards)
    if status in weekoff_keywords:
        result["status"] = "Week Off"; result["is_present"] = False; result["is_absent"] = True; return result
    if status in half_day_keywords:
        result["status"] = "HD"; result["is_present"] = True; result["is_absent"] = False; return result

    # Fallback: check punches
    has_in = parse_time_obj(in_time) is not None
    has_out = parse_time_obj(out_time) is not None
    if has_in and has_out:
        result["status"] = "P"; result["is_present"] = True; result["is_absent"] = False; return result
    if has_in or has_out:
        result["status"] = "SP"; result["is_present"] = True; result["is_absent"] = False; result["is_single_punch"] = True; return result
    return result

def normalize_employee_status(status: Optional[str]) -> str:
    if status is None:
        return "ACTIVE"  # FIX v3.2.2: Default to ACTIVE, not UNKNOWN
    status = str(status).strip().upper()
    if status == "":
        return "ACTIVE"  # FIX v3.2.2: Empty string = ACTIVE
    mapping = {
        "ACTIVE": "ACTIVE", "OK": "ACTIVE", "ON LEAVE": "ACTIVE", "TRANSFER": "ACTIVE",
        "LEFT": "INACTIVE", "NOT OK": "INACTIVE", "RESIGNED": "INACTIVE", "TERMINATED": "INACTIVE"
    }
    return mapping.get(status, "ACTIVE")  # FIX v3.2.2: Default unknown to ACTIVE

def is_employee_active(status: Optional[str]) -> bool:
    return normalize_employee_status(status) == "ACTIVE"

def calculate_late_punch_penalty(late_punch_count: int) -> Dict[str, Any]:
    """Escalating penalty: every 3 late punches in a month adds another 0.5 day.
    3 -> 0.5, 6 -> 1.0, 9 -> 1.5, 12 -> 2.0, ... (uncapped)."""
    penalty_units = late_punch_count // LATE_PUNCH_RULES["half_day_after"]
    penalty_days = penalty_units * 0.5
    next_threshold = (penalty_units + 1) * LATE_PUNCH_RULES["half_day_after"]
    if penalty_days <= 0:
        return {"penalty_days": 0.0, "penalty": "None", "action_required": False,
                "next_penalty_at": next_threshold}
    label = "Half Day" if penalty_days == 0.5 else ("One Day" if penalty_days == 1.0 else f"{penalty_days} Days")
    return {"penalty_days": penalty_days, "penalty": label, "action_required": True,
            "next_penalty_at": next_threshold}


def get_ot_totals(pr_number: str, reference_date: date, db: Session) -> tuple[float, float]:
    """Return (weekly_ot, monthly_ot) for an employee relative to reference_date.
    Weekly = rolling 7 days ending on reference_date.
    Monthly = current calendar month."""
    week_start = reference_date - timedelta(days=6)
    month_start = date(reference_date.year, reference_date.month, 1)
    month_end = date(reference_date.year, reference_date.month + 1, 1) if reference_date.month < 12 else date(reference_date.year + 1, 1, 1)

    weekly_ot = db.query(func.sum(Overtime.ot_hours)).filter(
        Overtime.pr_number == pr_number,
        Overtime.date >= week_start,
        Overtime.date <= reference_date
    ).scalar() or 0.0

    monthly_ot = db.query(func.sum(Overtime.ot_hours)).filter(
        Overtime.pr_number == pr_number,
        Overtime.date >= month_start,
        Overtime.date < month_end
    ).scalar() or 0.0

    return round(float(weekly_ot), 2), round(float(monthly_ot), 2)


def check_ot_thresholds(pr_number: str, emp_name: str, reference_date: date, db: Session) -> List[Dict[str, Any]]:
    """Check if employee OT exceeds weekly (12h) or monthly (48h) thresholds.
    Creates HRAction alerts if breached and not already flagged."""
    weekly_ot, monthly_ot = get_ot_totals(pr_number, reference_date, db)
    alerts: List[Dict[str, Any]] = []

    # Weekly threshold check
    if weekly_ot > OT_WEEKLY_THRESHOLD:
        week_start = reference_date - timedelta(days=6)
        existing = db.query(HRAction).filter(
            HRAction.pr_number == pr_number,
            HRAction.date >= week_start,
            HRAction.date <= reference_date,
            HRAction.action_type == "OT Weekly Threshold"
        ).first()
        if not existing:
            db.add(HRAction(
                pr_number=pr_number,
                emp_name=emp_name,
                date=reference_date,
                action_type="OT Weekly Threshold",
                description=(
                    f"{emp_name} ({pr_number}) has exceeded the weekly OT limit: "
                    f"{weekly_ot}h / {OT_WEEKLY_THRESHOLD}h "
                    f"(week ending {reference_date.isoformat()})"
                ),
                priority="High",
                assigned_to="HR Manager"
            ))
            alerts.append({"type": "weekly", "hours": weekly_ot, "threshold": OT_WEEKLY_THRESHOLD})

    # Monthly threshold check
    if monthly_ot > OT_MONTHLY_THRESHOLD:
        month_start = date(reference_date.year, reference_date.month, 1)
        month_end = date(reference_date.year, reference_date.month + 1, 1) if reference_date.month < 12 else date(reference_date.year + 1, 1, 1)
        existing = db.query(HRAction).filter(
            HRAction.pr_number == pr_number,
            HRAction.date >= month_start,
            HRAction.date < month_end,
            HRAction.action_type == "OT Monthly Threshold"
        ).first()
        if not existing:
            db.add(HRAction(
                pr_number=pr_number,
                emp_name=emp_name,
                date=reference_date,
                action_type="OT Monthly Threshold",
                description=(
                    f"{emp_name} ({pr_number}) has exceeded the monthly OT limit: "
                    f"{monthly_ot}h / {OT_MONTHLY_THRESHOLD}h "
                    f"for {reference_date.strftime('%Y-%m')}"
                ),
                priority="Critical",
                assigned_to="HR Manager"
            ))
            alerts.append({"type": "monthly", "hours": monthly_ot, "threshold": OT_MONTHLY_THRESHOLD})

    return alerts

def _find_max_consecutive_streak(rows: List["Attendance"], flag_fn) -> Dict[str, Any]:
    """Given Attendance rows sorted by date ascending, find the longest run of
    calendar-consecutive dates for which flag_fn(row) is True. A gap of even one
    day (e.g. a Week Off) breaks the streak."""
    best_len, best_start, best_end = 0, None, None
    cur_len, cur_start, prev_date = 0, None, None
    for r in rows:
        if flag_fn(r):
            if prev_date is not None and (r.date - prev_date).days == 1:
                cur_len += 1
            else:
                cur_len = 1
                cur_start = r.date
            prev_date = r.date
            if cur_len > best_len:
                best_len, best_start, best_end = cur_len, cur_start, r.date
        else:
            cur_len, cur_start, prev_date = 0, None, None
    return {"length": best_len, "start": best_start, "end": best_end}

def normalize_vendor(vendor: Optional[str]) -> str:
    if vendor is None:
        return "UNKNOWN"
    vendor = str(vendor).strip().upper()
    aliases = {
        "S.L.L": "SLL", "S L L": "SLL", "SEABIRD LOGISTICS LTD": "SLL", "SEABIRD LOGISOLUTIONS LIMITED": "SLL",
        "S.G": "SG", "S G": "SG", "S G ENTERPRISES": "SG", "S G ENTERPRISES_SEABIRD LOG.LTD.": "SG",
        "S.S.E": "SSE", "S S E": "SSE", "SHREE SADGURU ENT._SEABIRD LOG LTD": "SSE"
    }
    return aliases.get(vendor, vendor)

def normalize_store(store: Optional[str]) -> str:
    if store is None:
        return "UNKNOWN"
    return str(store).strip().upper()

def calculate_daily_summary(shift, category, status, in_time, out_time, man_hrs, force_man_hrs: bool = False) -> Dict[str, Any]:
    attendance = determine_attendance_status(status, in_time, out_time, man_hrs)
    ot = calculate_ot(shift, category, in_time, out_time, man_hrs, force_man_hrs=force_man_hrs)
    worked = ot["worked_hours"]

    # FIX v3.2.10: Auto-convert Present -> Half Day if worked hours <= threshold
    # Only apply if status was not explicitly set to HD/Leave/Absent/WO by uploaded data
    if attendance["status"] not in ["HD", "L", "A", "Week Off"]:
        half_day_limit = get_half_day_threshold(shift)
        if worked > 0 and worked <= half_day_limit:
            attendance["status"] = "HD"
            attendance["is_present"] = True
            attendance["is_absent"] = False

    return {
        **attendance, **ot,
        "is_late_punch": is_late_punch(shift, in_time),
        "is_early_departure": is_early_departure(shift, worked),
        "is_less_working_hours": is_early_departure(shift, worked),
        "late_minutes": get_late_minutes(shift, in_time) if is_late_punch(shift, in_time) else 0
    }



def get_tata_only_employees(pr_numbers: List[str], start: date, end: date, db: Session) -> set:
    """
    Detect employees who have Tata data but ZERO ESSL records in the given date range.
    These are 'Tata-only' employees for this period.
    """
    if not pr_numbers:
        return set()

    # Get all ESSL records for these employees in the date range
    essl_prs = db.query(ESSLAttendance.pr_number).filter(
        ESSLAttendance.pr_number.in_(pr_numbers),
        ESSLAttendance.date >= start,
        ESSLAttendance.date < end
    ).distinct().all()
    essl_prs_set = {r[0] for r in essl_prs if r[0]}

    # Tata-only = employees who have Tata data but NO ESSL data at all in this range
    tata_prs = db.query(TataAttendance.pr_number).filter(
        TataAttendance.pr_number.in_(pr_numbers),
        TataAttendance.date >= start,
        TataAttendance.date < end
    ).distinct().all()
    tata_prs_set = {r[0] for r in tata_prs if r[0]}

    # Employees with Tata but no ESSL
    tata_only = tata_prs_set - essl_prs_set
    return tata_only


def attendance_percentage(present: int, headcount: int) -> float:
    if headcount == 0:
        return 0.0
    return round((present / headcount) * 100, 2)

def absentee_percentage(absent: int, headcount: int) -> float:
    if headcount == 0:
        return 0.0
    return round((absent / headcount) * 100, 2)

# ============================================================================
# EXCEL HELPERS & SETTINGS
# ============================================================================

def read_excel_bytes_to_dicts(file_bytes: bytes, sheet_index: int = 0) -> List[Dict[str, Any]]:
    wb = load_workbook(filename=bio.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.worksheets[sheet_index] if sheet_index < len(wb.worksheets) else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            row_dict[header] = row[i] if i < len(row) else None
        result.append(row_dict)
    wb.close()
    return result

def read_excel_all_sheets_to_dicts(file_bytes: bytes, required_any_header: Optional[List[str]] = None) -> Dict[str, Any]:
    wb = load_workbook(filename=bio.BytesIO(file_bytes), read_only=True, data_only=True)
    all_rows: List[Dict[str, Any]] = []
    sheets_read = 0
    sheets_skipped = 0
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sheets_skipped += 1
            continue
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        if required_any_header and not any(h in headers for h in required_any_header):
            sheets_skipped += 1
            continue
        sheets_read += 1
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            row_dict = {}
            for i, header in enumerate(headers):
                row_dict[header] = row[i] if i < len(row) else None
            all_rows.append(row_dict)
    wb.close()
    return {"rows": all_rows, "sheets_read": sheets_read, "sheets_skipped": sheets_skipped}

def read_excel_raw_rows(file_bytes: bytes, sheet_index: int = 0) -> List[List[Any]]:
    wb = load_workbook(filename=bio.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.worksheets[sheet_index] if sheet_index < len(wb.worksheets) else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows

SETTINGS = {
    "company": {"name": "SeaBird Logistics", "client": "Tata Motors", "location": "Gurugram"},
    "shift_rules": {k: {kk: (vv.strftime("%H:%M") if isinstance(vv, time) else vv) for kk, vv in v.items()} for k, v in SHIFT_RULES.items()},
    "late_punch_rules": LATE_PUNCH_RULES,
    "category_rules": {k: v for k, v in CATEGORY_RULES.items()},
}

print("[INFO] Starting SeaBird HR Analytics API v3.2.10...")
print("[INFO] Server starting on http://0.0.0.0:8000")
print("[INFO] API docs: http://localhost:8000/docs")


# ============================================================================
# DASHBOARD ENDPOINTS (FIX v3.2.1: Restored missing endpoints)
# ============================================================================

@app.get("/api/v1/kpis")
def get_kpis(target_date: str = Query(...), db: Session = Depends(get_db)):
    today = parse_date_br(target_date)
    if not today:
        raise HTTPException(status_code=400, detail="Invalid target_date format. Use YYYY-MM-DD.")

    total_employees = db.query(Employee).count()  # FIX v3.2.8: Count ALL master employees

    # PERF FIX: joinedload eager-loads employee in the same query. Previously,
    # a.employee below triggered a separate SELECT per attendance row (N+1) -
    # with hundreds of present employees that was hundreds of extra round trips
    # on every KPI card load.
    attendance_records = db.query(Attendance).options(joinedload(Attendance.employee)).filter(Attendance.date == today).all()
    present = sum(1 for a in attendance_records if a.attendance_status not in ["Absent", "Leave", "Week Off"])
    # FIX v3.2.7: Count both Absent and Week Off in absent
    absent = sum(1 for a in attendance_records if a.attendance_status in ["Absent", "Week Off"])
    late = sum(1 for a in attendance_records if a.late_minutes > 0)
    single_punch = sum(1 for a in attendance_records if a.single_punch == "Yes")
    early_departure = sum(1 for a in attendance_records if a.early_minutes > 0)
    less_working = sum(1 for a in attendance_records if a.issue == "Less Working Hours")

    # FIX v3.2.1: Count DISTINCT employees with OT using Python set (avoids SQL NULL issues)
    ot_records = db.query(Attendance).filter(
        Attendance.date == today,
        Attendance.ot_hours > 0
    ).all()
    ot_employees = len(set(r.pr_number for r in ot_records if r.pr_number))

    total_ot = db.query(func.sum(Attendance.ot_hours)).filter(Attendance.date == today).scalar() or 0

    # FIX: Alternate shift = employees whose attendance shift differs from their master shift
    # FIX v3.2.1: Normalize both shifts before comparing to avoid false positives
    # FIX v3.2.2: Only count as alternate shift if both are valid shift codes
    VALID_SHIFTS = {"A", "B", "G", "C"}
    alternate_shift = 0
    for a in attendance_records:
        if a.employee and a.shift and a.employee.shift:
            # Normalize both: "A Shift" → "A", "General" → "G", etc.
            daily_shift = normalize_shift(a.shift)
            master_shift = normalize_shift(a.employee.shift)
            # Only count if both are valid shift codes (not time ranges like "15:00To23:30")
            if daily_shift in VALID_SHIFTS and master_shift in VALID_SHIFTS and daily_shift != master_shift:
                alternate_shift += 1

    # FIX: New joiners = employees who joined in current month
    current_month_start = date(today.year, today.month, 1)
    new_joiners = db.query(Employee).filter(
        Employee.status == "ACTIVE",
        Employee.join_date >= current_month_start,
        Employee.join_date <= today
    ).count()

    reconciliation_issues = db.query(AttendanceReconciliation).filter(AttendanceReconciliation.date == today).count()
    critical_issues = db.query(AttendanceReconciliation).filter(
        AttendanceReconciliation.date == today,
        AttendanceReconciliation.severity.in_(["Critical", "High"])
    ).count()

    # FIX: Proper attendance rate calculation
    attendance_rate = round((present / total_employees * 100), 2) if total_employees > 0 else 0

    # FIX v3.2.1: Match frontend KPIData interface exactly
    # Count on-leave employees
    on_leave = sum(1 for a in attendance_records if a.attendance_status == "Leave")

    # Count distinct vendors/stores/departments/shifts with attendance today
    vendor_count = len(set(a.vendor for a in attendance_records if a.vendor))
    store_count = len(set(a.store for a in attendance_records if a.store))
    dept_count = len(set(a.department for a in attendance_records if a.department))
    shift_count = len(set(a.shift for a in attendance_records if a.shift))

    # Pending actions and OT
    pending_actions = db.query(HRAction).filter(HRAction.status == "Open").count()
    pending_ot = db.query(Overtime).filter(Overtime.status == "Pending").count()

        # FIX v3.2.8: Headcount = Present + OT adjusted (+0.25 for 2hrs OT, +0.5 for 4hrs)
        # FIX v3.2.10: Auto-convert Present -> Half Day if worked_hours <= half_day_threshold (4.5h for A/B/G, 3.5h for C)
    headcount = present
    for a in attendance_records:
        ot = a.ot_hours or 0
        if ot >= 3.5 and ot <= 4.5:
            headcount += 0.5
        elif ot >= 1.5 and ot < 3.5:
            headcount += 0.25
    headcount = round(headcount, 2)

    # FIX v3.2.8: No Data now counts employees whose attendance record has issue="No Data"
    # (these are employees with no ESSL/Tata data who got an "Absent" record created)
    no_data_count = sum(1 for a in attendance_records if a.issue == "No Data")

    return {
        "total_employees": total_employees,
        "present": present,
        "absent": absent,
        "on_leave": on_leave,
        "late_punches": late,
        "ot_hours": round(total_ot, 2),
        "attendance_rate": attendance_rate,
        "vendor_count": vendor_count,
        "store_count": store_count,
        "department_count": dept_count,
        "shift_count": shift_count,
        "pending_actions": pending_actions,
        "pending_ot": pending_ot,
        "early_departure": early_departure,
        "single_punch": single_punch,
        "less_working_hours": less_working,
        "alternate_shift": alternate_shift,
        "ot_eligible": ot_employees,
        "headcount": headcount,
        "new_joiners": new_joiners,
        "no_data": no_data_count,
        "selected_date": today.isoformat()
    }

@app.get("/api/v1/trends")
def get_trends(target_date: str = Query(...), days: int = Query(30, ge=1, le=90), db: Session = Depends(get_db)):
    end_date = parse_date_br(target_date)
    if not end_date:
        raise HTTPException(status_code=400, detail="Invalid target_date format. Use YYYY-MM-DD.")
    start_date = end_date - timedelta(days=days)

    # PERF FIX: this used to run a fresh query PER DAY in the loop (2 queries x up
    # to 90 days = up to 180 sequential round trips for one chart). Attendance.date
    # is indexed, so pulling the whole range in one query and grouping in Python
    # is both correct and dramatically faster.
    from collections import defaultdict as _defaultdict
    rows = db.query(
        Attendance.date, Attendance.attendance_status, Attendance.late_minutes, Attendance.ot_hours
    ).filter(Attendance.date >= start_date, Attendance.date <= end_date).all()

    by_date = _defaultdict(lambda: {"present": 0, "absent": 0, "late_punches": 0, "ot_hours": 0.0})
    for rec_date, status, late_minutes, ot_hours in rows:
        bucket = by_date[rec_date]
        if status not in ("Absent", "Leave", "Week Off"):
            bucket["present"] += 1
        if status in ("Absent", "Week Off"):
            bucket["absent"] += 1
        if late_minutes and late_minutes > 0:
            bucket["late_punches"] += 1
        bucket["ot_hours"] += (ot_hours or 0)

    result = []
    d = start_date
    while d <= end_date:
        b = by_date.get(d, {"present": 0, "absent": 0, "late_punches": 0, "ot_hours": 0.0})
        result.append({
            "date": d.isoformat(),
            "present": b["present"],
            "absent": b["absent"],
            "late_punches": b["late_punches"],
            "ot_hours": round(b["ot_hours"], 2)
        })
        d += timedelta(days=1)

    return {"trends": result, "days": days, "target_date": end_date.isoformat()}

# PERF FIX: the four breakdown endpoints below used to loop over every distinct
# vendor/store/department/shift and fire 3 separate queries per group (count,
# attendance list, OT sum) - e.g. 10 stores = 30 queries for one chart, x4
# breakdown types = up to ~80-100 queries on a single dashboard load. Attendance
# already carries vendor/store/department/shift columns directly (no join to
# Employee needed for the day's stats), so this collapses to 2 queries total
# per endpoint regardless of how many distinct groups exist.
def _compute_group_breakdown(db: Session, field: str, today: date) -> Dict[str, Dict[str, Any]]:
    from collections import defaultdict as _defaultdict
    emp_col = getattr(Employee, field)
    att_col = getattr(Attendance, field)

    totals = dict(
        db.query(emp_col, func.count(Employee.id))
        .filter(emp_col.isnot(None))
        .group_by(emp_col)
        .all()
    )

    rows = db.query(att_col, Attendance.attendance_status, Attendance.late_minutes, Attendance.ot_hours) \
        .filter(Attendance.date == today, att_col.isnot(None)).all()

    agg = _defaultdict(lambda: {"present": 0, "absent": 0, "late": 0, "ot": 0.0})
    for key, status, late_minutes, ot_hours in rows:
        b = agg[key]
        if status not in ("Absent", "Leave", "Week Off"):
            b["present"] += 1
        if status in ("Absent", "Week Off"):
            b["absent"] += 1
        if late_minutes and late_minutes > 0:
            b["late"] += 1
        b["ot"] += (ot_hours or 0)

    result = {}
    for name, total in totals.items():
        b = agg.get(name, {"present": 0, "absent": 0, "late": 0, "ot": 0.0})
        result[name] = {
            "name": name,
            "present": b["present"],
            "absent": b["absent"],
            "total": total,
            "late": b["late"],
            "ot": round(b["ot"], 1),
            "percentage": round((b["present"] / total * 100), 1) if total > 0 else 0
        }
    return result


@app.get("/api/v1/breakdown/stores")
def get_store_breakdown(target_date: str = Query(...), db: Session = Depends(get_db)):
    today = parse_date_br(target_date)
    if not today:
        raise HTTPException(status_code=400, detail="Invalid target_date format. Use YYYY-MM-DD.")
    breakdown = _compute_group_breakdown(db, "store", today)
    return sorted(breakdown.values(), key=lambda x: x["present"], reverse=True)

@app.get("/api/v1/breakdown/vendors")
def get_vendor_breakdown(target_date: str = Query(...), db: Session = Depends(get_db)):
    today = parse_date_br(target_date)
    if not today:
        raise HTTPException(status_code=400, detail="Invalid target_date format. Use YYYY-MM-DD.")
    breakdown = _compute_group_breakdown(db, "vendor", today)
    return sorted(breakdown.values(), key=lambda x: x["present"], reverse=True)

@app.get("/api/v1/breakdown/departments")
def get_department_breakdown(target_date: str = Query(...), db: Session = Depends(get_db)):
    today = parse_date_br(target_date)
    if not today:
        raise HTTPException(status_code=400, detail="Invalid target_date format. Use YYYY-MM-DD.")
    breakdown = _compute_group_breakdown(db, "department", today)
    return [{"name": b["name"], "present": b["present"], "total": b["total"], "ot": b["ot"]} for b in breakdown.values()]

@app.get("/api/v1/breakdown/shifts")
def get_shift_breakdown(target_date: str = Query(...), db: Session = Depends(get_db)):
    today = parse_date_br(target_date)
    if not today:
        raise HTTPException(status_code=400, detail="Invalid target_date format. Use YYYY-MM-DD.")
    breakdown = _compute_group_breakdown(db, "shift", today)
    return [{"name": b["name"], "present": b["present"], "total": b["total"], "ot": b["ot"]} for b in breakdown.values()]

@app.get("/api/v1/actions/summary")
def get_action_summary(db: Session = Depends(get_db)):
    # PERF FIX: was 6 separate COUNT queries. One GROUP BY covers all action
    # types, plus the total is just the sum (no extra query needed).
    type_counts = dict(
        db.query(HRAction.action_type, func.count(HRAction.id))
        .filter(HRAction.status == "Open")
        .group_by(HRAction.action_type)
        .all()
    )
    return {
        "reconciliation": type_counts.get("Reconciliation", 0),
        "single_punch": type_counts.get("Single Punch", 0),
        "late_punches": type_counts.get("Late Punch", 0),
        "early_departure": type_counts.get("Early Departure", 0),
        "less_working_hours": type_counts.get("Less Working Hours", 0),
        "total_open": sum(type_counts.values())
    }

@app.get("/api/v1/actions/queue")
def get_action_queue(page: int = Query(1, ge=1), page_size: int = Query(50, ge=1, le=200), db: Session = Depends(get_db)):
    total = db.query(HRAction).filter(HRAction.status == "Open").count()
    items = db.query(HRAction).filter(HRAction.status == "Open").order_by(desc(HRAction.created_at)).offset((page - 1) * page_size).limit(page_size).all()
    data = [{"id": a.id, "type": a.action_type, "description": a.description, "priority": a.priority,
             "status": a.status, "assigned_to": a.assigned_to or "Unassigned"} for a in items]
    return {"data": data, "total": total, "page": page, "total_pages": max(1, (total + page_size - 1) // page_size)}

@app.post("/api/v1/actions/{action_id}/resolve")
def resolve_action(action_id: int, resolution: str = Form(...), db: Session = Depends(get_db)):
    action = db.query(HRAction).filter(HRAction.id == action_id).first()
    if not action:
        raise HTTPException(status_code=404, detail="Action not found")
    action.status = "Resolved"
    action.resolution = resolution
    action.resolved_at = datetime.utcnow()
    db.commit()
    return {"status": "resolved", "id": action_id}

@app.get("/api/v1/ai-insights")
def get_ai_insights(db: Session = Depends(get_db)):
    today = date.today()
    week_ago = today - timedelta(days=7)
    late_count = db.query(Attendance).filter(Attendance.date >= week_ago, Attendance.late_minutes > 0).count()
    absent_count = db.query(Attendance).filter(Attendance.date >= week_ago, Attendance.attendance_status == "Absent").count()
    ot_total = db.query(func.sum(Attendance.ot_hours)).filter(Attendance.date >= week_ago).scalar() or 0
    insights = []
    if late_count > 50:
        insights.append({"id": 1, "title": "High Late Punch Trend", "description": f"{late_count} late punches in last 7 days.",
                         "severity": "High", "recommendation": "Review shift schedules."})
    if absent_count > 30:
        insights.append({"id": 2, "title": "Absenteeism Spike", "description": f"{absent_count} absences in last 7 days.",
                         "severity": "Medium", "recommendation": "Conduct wellness checks."})
    if ot_total > 200:
        insights.append({"id": 3, "title": "Overtime Surge", "description": f"{ot_total:.1f} OT hours recently.",
                         "severity": "Medium", "recommendation": "Evaluate staffing levels."})
    if not insights:
        insights.append({"id": 1, "title": "All Systems Normal", "description": "No critical issues.",
                         "severity": "Low", "recommendation": "Continue monitoring."})
    return insights

# ============================================================================
# API ENDPOINTS - ATTENDANCE & EMPLOYEES
# ============================================================================

def determine_worked_shift(essl_in, essl_out, tata_in, tata_out, assigned_shift, existing_remark=None):
    """
    Determine which shift(s) the employee actually worked.
    Priority: 1) existing remark if it has multi-shift info, 2) compute from punches, 3) assigned shift
    """
    # If existing remark already has multi-shift info, use it
    if existing_remark and ("&" in existing_remark or "/2" in existing_remark):
        return existing_remark

    final_in = tata_in or essl_in
    final_out = tata_out or essl_out

    if not final_in or not final_out:
        return assigned_shift or "G"

    in_time = parse_time_obj(final_in)
    out_time = parse_time_obj(final_out)
    if not in_time or not out_time:
        return assigned_shift or "G"

    in_mins = in_time.hour * 60 + in_time.minute
    out_mins = out_time.hour * 60 + out_time.minute
    if out_mins < in_mins:
        out_mins += 24 * 60

    worked_hours = (out_mins - in_mins) / 60

    # Determine start shift from IN time
    if 5.5*60 <= in_mins < 8.5*60:
        start_shift = "A"
    elif 8.5*60 <= in_mins < 15*60:
        start_shift = "G"
    elif 15*60 <= in_mins < 23.5*60:
        start_shift = "B"
    else:
        start_shift = "C"

    # Determine end shift from OUT time (with tolerance)
    out_mins_norm = out_mins % (24*60)
    if 5.5*60 <= out_mins_norm < 8.5*60:
        end_shift = "A"
    elif 8.5*60 <= out_mins_norm < 15*60:
        end_shift = "G"
    elif 15*60 <= out_mins_norm < 23.5*60:
        end_shift = "B"
    else:
        end_shift = "C"

    # If worked hours is normal (<=9.5h), likely same shift
    if worked_hours <= 9.5:
        return start_shift

    # Multi-shift: format based on patterns
    if start_shift == end_shift:
        return start_shift

    if start_shift == "A" and end_shift == "G":
        return "A&G"
    if start_shift == "G" and end_shift == "B":
        return "G&B"
    if start_shift == "A" and end_shift == "B":
        return "A&B/2"
    if start_shift == "G" and end_shift == "A":
        return "G&A"
    if start_shift == "B" and end_shift == "G":
        return "B&G"

    return f"{start_shift}&{end_shift}"




@app.get("/api/v1/attendance/daily")
def get_daily_register(
    date_filter: Optional[str] = Query(None, alias="date"),
    vendor: Optional[str] = Query(None), store: Optional[str] = Query(None),
    department: Optional[str] = Query(None), status: Optional[str] = Query(None),
    issue: Optional[str] = Query(None),
    search: Optional[str] = Query(None), page: int = Query(1, ge=1), page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db)
):
    """Optimized daily register: avoids JOIN unless employee filters are used.
    Uses selectinload to prevent N+1. Orders by Attendance.pr_number (indexed)
    instead of Employee.pr_number to eliminate slow sorts."""
    from sqlalchemy.orm import selectinload

    target_date = parse_date_br(date_filter) if date_filter else date.today()
    if not target_date:
        target_date = date.today()

    # If employee-side filters are active, resolve matching PRs first in a
    # lightweight query, then filter Attendance by PR list. This avoids the
    # heavy JOIN + COUNT + OFFSET pattern that kills performance.
    pr_filter = None
    needs_emp_lookup = bool(vendor or store or department or search)

    if needs_emp_lookup:
        emp_q = db.query(Employee.pr_number)
        if vendor:   emp_q = emp_q.filter(Employee.vendor == vendor)
        if store:    emp_q = emp_q.filter(Employee.store == store)
        if department: emp_q = emp_q.filter(Employee.department == department)
        if search:
            emp_q = emp_q.filter(
                (Employee.pr_number.ilike(f"%{search}%")) |
                (Employee.name.ilike(f"%{search}%"))
            )
        pr_rows = emp_q.all()
        pr_filter = {r[0] for r in pr_rows if r[0]}
        if not pr_filter:
            return {"data": [], "total": 0, "page": page, "total_pages": 1}

    # Base Attendance query — NO JOIN, uses indexes on date + pr_number
    q = db.query(Attendance).filter(Attendance.date == target_date)
    if status: q = q.filter(Attendance.attendance_status == status)
    if issue:  q = q.filter(Attendance.issue == issue)
    if pr_filter: q = q.filter(Attendance.pr_number.in_(pr_filter))

    # Fast count (index-only scan on date)
    total = q.count()

    # Paginate and bulk-load employees in one extra query (selectinload)
    items = (
        q.options(selectinload(Attendance.employee))
          .order_by(Attendance.pr_number)
          .offset((page - 1) * page_size)
          .limit(page_size)
          .all()
    )

    data = []
    for att in items:
        emp = att.employee
        data.append({
            "id": att.id,
            "pr_number": att.pr_number,
            "name": emp.name if emp else att.pr_number,
            "vendor": att.vendor or (emp.vendor if emp else ""),
            "store": att.store or (emp.store if emp else ""),
            "department": att.department or (emp.department if emp else ""),
            "assigned_shift": att.shift or (emp.shift if emp else "G"),
            "worked_shift": determine_worked_shift(
                att.essl_in, att.essl_out, att.tata_in, att.tata_out,
                att.shift or (emp.shift if emp else "G"), att.remark
            ),
            "category": emp.wc if emp else (att.category or "BC"),
            "essl_in": att.essl_in,
            "essl_out": att.essl_out,
            "tata_in": att.tata_in,
            "tata_out": att.tata_out,
            "final_in": att.final_in,
            "final_out": att.final_out,
            "worked_hours": att.worked_hours,
            "man_hrs": att.man_hrs,
            "attendance_status": att.attendance_status,
            "ot_hours": att.ot_hours,
            "ot_headcount": att.ot_headcount,
            "late_minutes": att.late_minutes,
            "early_minutes": att.early_minutes,
            "single_punch": att.single_punch,
            "is_match": att.is_match,
            "issue": att.issue,
            "remark": att.remark
        })

    return {
        "data": data,
        "total": total,
        "page": page,
        "total_pages": max(1, (total + page_size - 1) // page_size)
    }

@app.get("/api/v1/attendance/monthly/{pr_number}")
def get_monthly_attendance(pr_number: str, month: Optional[str] = Query(None), db: Session = Depends(get_db)):
    emp = db.query(Employee).filter(Employee.pr_number == pr_number).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    target_month = month or datetime.now().strftime("%Y-%m")
    year, mon = parse_year_month(target_month)
    start = date(year, mon, 1)
    end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)
    records = db.query(Attendance).filter(Attendance.pr_number == pr_number, Attendance.date >= start, Attendance.date < end).all()
    days = []
    for r in records:
        days.append({
            "date": r.date.isoformat(), "attendance_status": r.attendance_status,
            "essl_in": r.essl_in, "essl_out": r.essl_out, "tata_in": r.tata_in, "tata_out": r.tata_out,
            "final_in": r.final_in, "final_out": r.final_out, "worked_hours": r.worked_hours, "man_hrs": r.man_hrs,
            "ot_hours": r.ot_hours, "ot_headcount": r.ot_headcount, "late_minutes": r.late_minutes, "early_minutes": r.early_minutes,
            "single_punch": r.single_punch, "is_match": r.is_match, "remark": r.remark
        })
    summary = {
        "present": sum(1 for d in days if d["attendance_status"] not in ["Absent", "Leave", "Week Off"]),
        "absent": sum(1 for d in days if d["attendance_status"] == "Absent"),
        "half_day": sum(1 for d in days if d["attendance_status"] == "Half Day"),
        "late": sum(1 for d in days if d["late_minutes"] > 0),
        "total_ot": round(sum(d["ot_hours"] for d in days), 1),
        "single_punch_days": sum(1 for d in days if d["single_punch"] == "Yes")
    }
    return {"pr_number": pr_number, "name": emp.name, "days": days, "summary": summary}

@app.get("/api/v1/employees")
def get_employees(
    search: Optional[str] = Query(None), vendor: Optional[str] = Query(None), store: Optional[str] = Query(None),
    department: Optional[str] = Query(None), status: Optional[str] = Query(None),
    page: int = Query(1, ge=1), page_size: int = Query(50, ge=1, le=200), db: Session = Depends(get_db)
):
    q = db.query(Employee)
    if search: q = q.filter((Employee.pr_number.ilike(f"%{search}%")) | (Employee.name.ilike(f"%{search}%")) | (Employee.emp_code.ilike(f"%{search}%")))
    if vendor: q = q.filter(Employee.vendor == vendor)
    if store: q = q.filter(Employee.store == store)
    if department: q = q.filter(Employee.department == department)
    if status: q = q.filter(Employee.status == status)
    total = q.count()
    items = q.order_by(Employee.pr_number).offset((page - 1) * page_size).limit(page_size).all()
    data = [{"id": e.id, "pr_number": e.pr_number, "bio_id": e.bio_id, "emp_code": e.emp_code, "name": e.name, "vendor": e.vendor,
             "wc": e.wc, "bc": e.bc, "store": e.store, "designation": e.designation, "status": e.status,
             "shift": e.shift, "department": e.department} for e in items]
    return {"data": data, "total": total, "page": page, "total_pages": max(1, (total + page_size - 1) // page_size)}

@app.get("/api/v1/vendors")
def get_vendors(db: Session = Depends(get_db)):
    today = date.today()
    vendors = db.query(Employee.vendor).distinct().all()
    result = []
    for (v,) in vendors:
        if not v: continue
        total = db.query(Employee).filter(Employee.vendor == v).count()  # FIX v3.2.8: All employees
        vendor_attendance = db.query(Attendance).join(Employee).filter(
            Employee.vendor == v, Attendance.date == today
        ).all()
        present = sum(1 for a in vendor_attendance if a.attendance_status not in ["Absent", "Leave", "Week Off"])
        ot = db.query(func.sum(Attendance.ot_hours)).join(Employee).filter(Employee.vendor == v, Attendance.date == today).scalar() or 0
        result.append({"name": v, "present": present, "total": total, "ot": round(ot, 1)})
    return result

@app.get("/api/v1/stores")
def get_stores(db: Session = Depends(get_db)):
    today = date.today()
    stores = db.query(Employee.store).distinct().all()
    result = []
    for (s,) in stores:
        if not s: continue
        total = db.query(Employee).filter(Employee.store == s).count()  # FIX v3.2.8: All employees
        store_attendance = db.query(Attendance).join(Employee).filter(
            Employee.store == s, Attendance.date == today
        ).all()
        present = sum(1 for a in store_attendance if a.attendance_status not in ["Absent", "Leave", "Week Off"])
        ot = db.query(func.sum(Attendance.ot_hours)).join(Employee).filter(Employee.store == s, Attendance.date == today).scalar() or 0
        result.append({"name": s, "present": present, "total": total, "ot": round(ot, 1)})
    return result

@app.get("/api/v1/departments")
def get_departments(db: Session = Depends(get_db)):
    today = date.today()
    depts = db.query(Employee.department).distinct().all()
    result = []
    for (d,) in depts:
        if not d: continue
        total = db.query(Employee).filter(Employee.department == d).count()  # FIX v3.2.8: All employees
        dept_attendance = db.query(Attendance).join(Employee).filter(
            Employee.department == d, Attendance.date == today
        ).all()
        present = sum(1 for a in dept_attendance if a.attendance_status not in ["Absent", "Leave", "Week Off"])
        ot = db.query(func.sum(Attendance.ot_hours)).join(Employee).filter(Employee.department == d, Attendance.date == today).scalar() or 0
        result.append({"name": d, "present": present, "total": total, "ot": round(ot, 1)})
    return result

@app.get("/api/v1/overtime/summary")
def get_ot_summary(db: Session = Depends(get_db)):
    today = date.today()
    eligible = db.query(Attendance).filter(Attendance.date == today, Attendance.ot_hours > 0).count()
    pending = db.query(Overtime).filter(Overtime.status == "Pending").count()
    approved = db.query(Overtime).filter(Overtime.status == "Approved").count()
    total = db.query(func.sum(Overtime.ot_hours)).scalar() or 0
    return {"ot_eligible": eligible, "pending_approval": pending, "approved": approved, "total_hours": round(total, 1)}

@app.get("/api/v1/overtime/requests")
def get_ot_requests(page: int = Query(1, ge=1), page_size: int = Query(50, ge=1, le=200), db: Session = Depends(get_db)):
    total = db.query(Overtime).count()
    items = db.query(Overtime).order_by(desc(Overtime.created_at)).offset((page - 1) * page_size).limit(page_size).all()
    data = [{"id": o.id, "pr_number": o.pr_number, "name": o.name, "store": o.store,
             "worked_hours": o.worked_hours, "ot_hours": o.ot_hours, "calc_headcount": o.calc_headcount, "status": o.status} for o in items]
    return {"data": data, "total": total, "page": page, "total_pages": max(1, (total + page_size - 1) // page_size)}

@app.post("/api/v1/overtime/{ot_id}/approve")
def approve_ot(ot_id: int, action: str = Form(...), db: Session = Depends(get_db)):
    ot = db.query(Overtime).filter(Overtime.id == ot_id).first()
    if not ot:
        raise HTTPException(status_code=404, detail="OT request not found")
    ot.status = "Approved" if action == "approve" else "Rejected"
    ot.approved_by = "HR Manager"
    ot.approved_at = datetime.utcnow()
    db.commit()
    return {"status": ot.status, "id": ot_id}

@app.get("/api/v1/reports")
def get_reports():
    return [
        {"id": 1, "name": "Dump Report", "description": "Export reconciled attendance (ESSL + Tata side by side)", "type": "dump"},
        {"id": 2, "name": "Monthly Summary", "description": "Monthly summary by vendor/store", "type": "monthly"},
        {"id": 3, "name": "OT Report", "description": "OT hours and approvals", "type": "ot"},
        {"id": 4, "name": "Late Punch Report", "description": "Late punch analysis", "type": "late"},
    ]

@app.get("/api/v1/settings")
def get_settings():
    return SETTINGS

# ============================================================================
# UPLOAD ENDPOINTS
# ============================================================================

@app.get("/api/v1/upload/history")
def get_upload_history(db: Session = Depends(get_db)):
    logs = db.query(UploadLog).order_by(desc(UploadLog.uploaded_at)).limit(50).all()
    return [{"id": l.id, "type": l.file_type, "filename": l.filename, "rows_processed": l.rows_processed,
             "status": l.status, "uploaded_at": l.uploaded_at.isoformat()} for l in logs]

@app.delete("/api/v1/upload/{upload_id}")
def delete_upload(upload_id: int, db: Session = Depends(get_db)):
    """FULL CASCADE DELETE:
    - Master: marks all employees from this upload as INACTIVE + deletes the log
    - ESSL/Tata/Daywise: deletes raw rows + ALL derived data (Attendance, Reconciliation,
      HRAction, Overtime) for the date range covered by this upload
    - Tata-All: deletes the log only (monthly summaries are not tagged by upload_log_id)
    """
    log = db.query(UploadLog).filter(UploadLog.id == upload_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Upload not found")

    file_type = log.file_type
    essl_deleted = 0
    tata_deleted = 0
    derived_deleted = {"attendance": 0, "reconciliation": 0, "hr_actions": 0, "overtime": 0}
    employees_deleted = 0

    if file_type == "master":
        # FIX: hard-DELETEing these employees crashes with a Postgres
        # IntegrityError (foreign key violation) the moment any of them has
        # attendance/reconciliation/overtime history, since employee_id on
        # those tables references employees.id and nothing cleans those rows
        # up first. That's what was producing the unexplained HTTP 500 here.
        # Soft-delete instead: mark them INACTIVE. This is also just correct
        # behavior - undoing a master upload shouldn't destroy attendance
        # history that's already been reconciled against those employee IDs.
        employees_deleted = db.query(Employee).filter(
            Employee.upload_log_id == upload_id
        ).update({"status": "INACTIVE"}, synchronize_session=False)
        db.commit()

    elif file_type in ("essl", "tata", "daywise"):
        # Find the date range this upload covered
        date_range = None
        if file_type == "essl":
            date_range = db.query(
                func.min(ESSLAttendance.date), func.max(ESSLAttendance.date)
            ).filter(ESSLAttendance.upload_log_id == upload_id).first()
        elif file_type == "tata":
            date_range = db.query(
                func.min(TataAttendance.date), func.max(TataAttendance.date)
            ).filter(TataAttendance.upload_log_id == upload_id).first()
        elif file_type == "daywise":
            essl_range = db.query(
                func.min(ESSLAttendance.date), func.max(ESSLAttendance.date)
            ).filter(ESSLAttendance.upload_log_id == upload_id).first()
            tata_range = db.query(
                func.min(TataAttendance.date), func.max(TataAttendance.date)
            ).filter(TataAttendance.upload_log_id == upload_id).first()
            dates = [d for d in [essl_range[0] if essl_range else None, essl_range[1] if essl_range else None,
                                 tata_range[0] if tata_range else None, tata_range[1] if tata_range else None] if d]
            if dates:
                date_range = (min(dates), max(dates))

        # Delete raw rows
        essl_deleted = db.query(ESSLAttendance).filter(
            ESSLAttendance.upload_log_id == upload_id
        ).delete(synchronize_session=False)
        tata_deleted = db.query(TataAttendance).filter(
            TataAttendance.upload_log_id == upload_id
        ).delete(synchronize_session=False)

        # Delete ALL derived data for the affected date range
        if date_range and date_range[0] and date_range[1]:
            start, end = date_range[0], date_range[1] + timedelta(days=1)
            derived_deleted["attendance"] = db.query(Attendance).filter(
                Attendance.date >= start, Attendance.date < end
            ).delete(synchronize_session=False)
            derived_deleted["reconciliation"] = db.query(AttendanceReconciliation).filter(
                AttendanceReconciliation.date >= start, AttendanceReconciliation.date < end
            ).delete(synchronize_session=False)
            derived_deleted["hr_actions"] = db.query(HRAction).filter(
                HRAction.date >= start, HRAction.date < end
            ).delete(synchronize_session=False)
            derived_deleted["overtime"] = db.query(Overtime).filter(
                Overtime.date >= start, Overtime.date < end
            ).delete(synchronize_session=False)
            month_str = start.strftime("%Y-%m")
            db.query(LatePunchPenalty).filter(LatePunchPenalty.month == month_str).delete(synchronize_session=False)
            db.query(BehavioralAlert).filter(BehavioralAlert.month == month_str).delete(synchronize_session=False)

        db.commit()

    # FIX: uploads made BEFORE upload_log_id existed have NULL on every row, so the
    # cascade filters above match nothing for them - the delete "succeeds" but only
    # removes the UploadLog row itself, leaving all the actual data behind. That's
    # confusing ("it just deletes the log"), so detect it and say so honestly instead
    # of claiming a full cascade delete happened.
    total_rows_touched = essl_deleted + tata_deleted + employees_deleted + sum(derived_deleted.values())
    untracked = (file_type in ("master", "essl", "tata", "daywise")) and total_rows_touched == 0

    # Delete the log itself
    db.delete(log)
    db.commit()

    if untracked:
        return {
            "status": "log_only_deleted",
            "id": upload_id,
            "file_type": file_type,
            "essl_rows_deleted": 0, "tata_rows_deleted": 0,
            "employees_deleted": 0, "derived_data_deleted": derived_deleted,
            "message": (
                f"This upload predates change-tracking (upload_log_id), so I could not "
                f"identify which rows it created - only the log entry was removed. "
                f"The underlying data is still in the database. "
                f"To actually remove it, use 'Clear Month' (DELETE /api/v1/attendance/clear-month"
                f"?month=YYYY-MM) for the affected month, or re-upload the corrected file with "
                f"force=true to overwrite it in place."
            )
        }

    return {
        "status": "deleted",
        "id": upload_id,
        "file_type": file_type,
        "essl_rows_deleted": essl_deleted,
        "tata_rows_deleted": tata_deleted,
        "employees_deleted": employees_deleted,
        "derived_data_deleted": derived_deleted,
        "message": (
            f"Upload {upload_id} ({file_type}) fully deleted. "
            f"Raw: {essl_deleted + tata_deleted} rows. "
            f"Derived: {sum(derived_deleted.values())} rows. "
            f"Employees deleted: {employees_deleted}."
        )
    }


# FIX: Frontend might call /uploads/ (plural). This alias prevents 404s.
@app.delete("/api/v1/uploads/{upload_id}")
def delete_upload_alias(upload_id: int, db: Session = Depends(get_db)):
    """Alias for /api/v1/upload/{upload_id} — fixes frontend 404s."""
    return delete_upload(upload_id, db)


def _clear_derived_data(start: date, end: date, db: Session):
    """Nuclear option: wipe Attendance, Reconciliation, HRAction, Overtime for a date range.
    Use this before re-running reconciliation on a month that has stale data."""
    db.query(Attendance).filter(Attendance.date >= start, Attendance.date < end).delete(synchronize_session=False)
    db.query(AttendanceReconciliation).filter(AttendanceReconciliation.date >= start, AttendanceReconciliation.date < end).delete(synchronize_session=False)
    db.query(HRAction).filter(HRAction.date >= start, HRAction.date < end).delete(synchronize_session=False)
    db.query(Overtime).filter(Overtime.date >= start, Overtime.date < end).delete(synchronize_session=False)
    db.query(LatePunchPenalty).filter(LatePunchPenalty.month == start.strftime("%Y-%m")).delete(synchronize_session=False)
    db.query(BehavioralAlert).filter(BehavioralAlert.month == start.strftime("%Y-%m")).delete(synchronize_session=False)
    db.commit()


@app.delete("/api/v1/system/reset-all")
def reset_all_data(confirm: str = Query(...), db: Session = Depends(get_db)):
    """NUCLEAR OPTION: wipes EVERYTHING - all attendance data, all reconciliation
    results, all HR actions/OT, all upload logs, AND all Employee/Vendor/Store/
    Department master data. This is what "delete all with master" means - a full
    reset back to an empty database.

    Does NOT touch the User table (login credentials) - you'd otherwise lock
    yourself out of the admin panel.

    Requires ?confirm=DELETE_EVERYTHING to run, since this is irreversible and
    there is no undo. Deletes child tables before Employee to satisfy the
    employee_id foreign keys on Attendance/ESSLAttendance/TataAttendance/Overtime.
    """
    if confirm != "DELETE_EVERYTHING":
        raise HTTPException(
            status_code=400,
            detail="This permanently deletes ALL data including master/Employee records. "
                   "Call again with ?confirm=DELETE_EVERYTHING to proceed."
        )

    counts = {}
    # Children first (FK: employee_id -> employees.id), then parents, then
    # standalone tables. Order matters for Postgres FK constraints.
    for model, key in [
        (HRAction, "hr_actions"),
        (LatePunchPenalty, "late_punch_penalties"),
        (BehavioralAlert, "behavioral_alerts"),
        (AttendanceReconciliation, "reconciliation"),
        (Overtime, "overtime"),
        (Attendance, "attendance"),
        (ESSLAttendance, "essl"),
        (TataAttendance, "tata"),
        (MonthlyTataAttendance, "monthly_tata"),
        (UploadLog, "upload_logs"),
        (Employee, "employees"),
        (Vendor, "vendors"),
        (Store, "stores"),
        (Department, "departments"),
    ]:
        counts[key] = db.query(model).delete(synchronize_session=False)
    db.commit()

    return {
        "status": "reset_complete",
        "rows_deleted": counts,
        "total_rows_deleted": sum(counts.values()),
        "message": (
            "Everything wiped: all attendance/reconciliation/HR/overtime data, all "
            "upload logs, and all master data (employees, vendors, stores, departments). "
            "Your login (User table) was left intact. Start fresh with a Master upload."
        )
    }


@app.delete("/api/v1/attendance/clear-month")
def clear_month_data(month: str = Query(...), db: Session = Depends(get_db)):
    """Wipe ALL derived reconciliation data for a month. Use before re-running reconciliation
    if your data looks corrupted or has ghost employees."""
    year, mon = parse_year_month(month)
    start = date(year, mon, 1)
    end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)
    _clear_derived_data(start, end, db)
    return {"status": "cleared", "month": month, "message": f"All Attendance/Reconciliation/Actions/OT for {month} deleted. Upload fresh ESSL/Tata and re-run reconciliation."}


@app.delete("/api/v1/attendance/clear-month-raw")
def clear_month_raw_data(month: str = Query(...), db: Session = Depends(get_db)):
    """FULL WIPE for a month, by DATE RANGE rather than upload_log_id - this is the
    fix for data uploaded before upload_log_id tracking existed (e.g. the original
    June dataset), where per-upload delete can't identify which rows to remove
    because they were never tagged. This deletes:
    - Raw ESSLAttendance + TataAttendance rows in the month
    - All derived Attendance/Reconciliation/HRAction/Overtime/LatePunchPenalty/BehavioralAlert
    It does NOT touch Employee/master data - use a fresh master upload with force=true
    to fix employee records instead.
    """
    year, mon = parse_year_month(month)
    start = date(year, mon, 1)
    end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)

    essl_deleted = db.query(ESSLAttendance).filter(
        ESSLAttendance.date >= start, ESSLAttendance.date < end
    ).delete(synchronize_session=False)
    tata_deleted = db.query(TataAttendance).filter(
        TataAttendance.date >= start, TataAttendance.date < end
    ).delete(synchronize_session=False)
    db.commit()

    _clear_derived_data(start, end, db)

    return {
        "status": "cleared",
        "month": month,
        "essl_rows_deleted": essl_deleted,
        "tata_rows_deleted": tata_deleted,
        "message": (
            f"Fully wiped raw + derived attendance data for {month} ({essl_deleted} ESSL rows, "
            f"{tata_deleted} Tata rows). Employees/master data untouched. "
            f"Old Upload History entries for this month will still be listed (they're just log "
            f"records now) - deleting them individually is safe and will correctly report nothing "
            f"left to remove. Re-upload ESSL/Tata for {month} and re-run reconciliation."
        )
    }


@app.post("/api/v1/upload/master")
def upload_master(file: UploadFile = File(...), force: bool = Form(False), deactivate_missing: bool = Form(False), db: Session = Depends(get_db)):
    # NOTE: sync `def` (not async) so Starlette runs this in a worker thread instead of
    # blocking the single event loop for the whole upload - fixes ERR_HTTP2_PING_FAILED /
    # "CORS" errors on large files, since the server can still answer keepalive pings.
    contents = file.file.read()
    file_hash = hashlib.sha256(contents).hexdigest()
    if not force:
        existing = db.query(UploadLog).filter(UploadLog.file_hash == file_hash).first()
        if existing:
            return {"status": "duplicate", "message": "Already uploaded", "previous_id": existing.id}
    try:
        rows = read_excel_bytes_to_dicts(contents)
    except Exception as e:
        db.add(UploadLog(file_type="master", filename=file.filename, file_hash=file_hash, rows_processed=0, status="Failed", error_message=str(e)))
        db.commit()
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {str(e)}")

    processed = 0
    vendors_created = 0
    stores_created = 0

    # Prefetch once instead of hitting the DB 3x per row (this was the main slowdown)
    emp_by_pr = {e.pr_number: e for e in db.query(Employee).all() if e.pr_number}
    existing_vendor_names = {v.name for v in db.query(Vendor.name).all()}
    existing_store_names = {s.name for s in db.query(Store.name).all()}

    # Create upload log early so we can tag employees with it
    upload_log = UploadLog(file_type="master", filename=file.filename, file_hash=file_hash, rows_processed=0, status="Processing")
    db.add(upload_log)
    db.flush()
    upload_log_id = upload_log.id

    seen_prs = set()

    for row in rows:
        pr = str(row.get("PR", row.get("pr_number", row.get("PR Number", row.get("PR_Number", ""))))).strip()
        if not pr:
            continue
        seen_prs.add(pr)
        bio_val = row.get("Bio", row.get("BioID", None))
        bio_id = str(bio_val).strip() if bio_val not in (None, "") else ""
        name = str(row.get("Name", row.get("EMP Name", row.get("Employee Name", "")))).strip()
        vendor_name = normalize_vendor(str(row.get("Vendor", row.get("vendor", row.get("Contractor", "")))).strip())
        store_name = normalize_store(str(row.get("Store", row.get("store", row.get("Location", "")))).strip())
        category = normalize_category(str(row.get("WC/BC", row.get("Category", row.get("category", "")))).strip())
        status_val = str(row.get("Status", row.get("status", "ACTIVE"))).strip().upper() or "ACTIVE"

        if vendor_name and vendor_name not in existing_vendor_names:
            db.add(Vendor(name=vendor_name))
            existing_vendor_names.add(vendor_name)
            vendors_created += 1
        if store_name and store_name not in existing_store_names:
            db.add(Store(name=store_name, location=store_name))
            existing_store_names.add(store_name)
            stores_created += 1

        emp = emp_by_pr.get(pr)
        if emp:
            emp.bio_id = bio_id or emp.bio_id
            emp.name = name or emp.name
            emp.vendor = vendor_name or emp.vendor
            emp.store = store_name or emp.store
            emp.designation = str(row.get("Designation", row.get("designation", emp.designation))).strip() or emp.designation
            emp.wc = category or emp.wc
            emp.status = normalize_employee_status(status_val) if status_val else emp.status
            emp.join_date = parse_date_br(row.get("DOJ", row.get("doj", ""))) or emp.join_date
            emp.upload_log_id = upload_log_id  # track which master last touched this employee
        else:
            new_emp = Employee(
                pr_number=pr, bio_id=bio_id, emp_code="", name=name or pr,
                vendor=vendor_name, store=store_name, department="",
                designation=str(row.get("Designation", row.get("designation", ""))).strip(),
                shift="G",
                wc=category, bc="",
                status=normalize_employee_status(status_val),
                join_date=parse_date_br(row.get("DOJ", row.get("doj", ""))),
                upload_log_id=upload_log_id
            )
            db.add(new_emp)
            emp_by_pr[pr] = new_emp  # so a duplicate PR later in the same file updates, not re-inserts
        processed += 1

    # CRITICAL FIX: Employees who are ACTIVE in DB but NOT in this master file are
    # marked INACTIVE. This stops reconciliation from creating attendance rows for
    # employees who left the company.
    deactivated_count = 0
    if deactivate_missing:
        for pr, emp in emp_by_pr.items():
            if pr not in seen_prs and emp.status == "ACTIVE":
                emp.status = "INACTIVE"
                deactivated_count += 1

    upload_log.status = "Success"
    upload_log.rows_processed = processed
    db.commit()
    return {
        "status": "success", "type": "master", "rows_processed": processed,
        "vendors_created": vendors_created, "stores_created": stores_created,
        "deactivated_employees": deactivated_count,
        "filename": file.filename,
        "upload_log_id": upload_log_id
    }

@app.post("/api/v1/upload/essl")
def upload_essl(file: UploadFile = File(...), force: bool = Form(False), db: Session = Depends(get_db)):
    import time as time_module
    start_time = time_module.time()

    contents = file.file.read()
    file_hash = hashlib.sha256(contents).hexdigest()
    if not force:
        existing = db.query(UploadLog).filter(UploadLog.file_hash == file_hash).first()
        if existing:
            return {"status": "duplicate", "message": "Already uploaded", "previous_id": existing.id}

    try:
        raw_rows = read_excel_raw_rows(contents)
    except Exception as e:
        db.add(UploadLog(file_type="essl", filename=file.filename, file_hash=file_hash, rows_processed=0, status="Failed", error_message=str(e)))
        db.commit()
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {str(e)}")

    if not raw_rows or len(raw_rows) < 2:
        raise HTTPException(status_code=400, detail="Empty or invalid Excel file")

    all_employees = db.query(Employee).all()
    emp_by_pr = {e.pr_number: e for e in all_employees if e.pr_number}
    emp_by_name = {}
    for e in all_employees:
        if e.name:
            emp_by_name[e.name.strip().upper()] = e
    emp_by_code = {e.emp_code: e for e in all_employees if e.emp_code}

    def resolve_employee_fast(emp_code_val: str, emp_name_val: str):
        e = emp_by_pr.get(emp_code_val)
        if e:
            return e
        if emp_name_val:
            e = emp_by_name.get(emp_name_val.strip().upper())
            if e:
                return e
        return emp_by_code.get(emp_code_val)

    header_row_idx = None
    date_columns = []
    for i, row in enumerate(raw_rows[:15]):
        if len(row) < 3:
            continue
        dates_in_row = 0
        for cell in row:
            if cell is not None:
                cell_str = str(cell).strip()
                if re.match(r'\d{4}-\d{2}-\d{2}', cell_str) or re.match(r'\d{2}/\d{2}/\d{4}', cell_str):
                    dates_in_row += 1
        if dates_in_row >= 3:
            header_row_idx = i
            for j, cell in enumerate(row):
                if cell is not None:
                    parsed = parse_date_br(str(cell).strip())
                    if parsed:
                        date_columns.append((j, parsed))
            break

    # FLAT FORMAT FALLBACK
    if header_row_idx is None:
        try:
            rows = read_excel_bytes_to_dicts(contents)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not parse Excel: {str(e)}")

        parsed_rows = []
        all_dates = set()
        for row in rows:
            emp_code = str(row.get("EMP Code", row.get("Emp Code", row.get("emp_code", "")))).strip()
            emp_name = str(row.get("EMP Name", row.get("Emp Name", row.get("name", "")))).strip()
            att_date = parse_date_br(row.get("Date", row.get("date", "")))
            in_time = safe_time(parse_time_br(row.get("IN", row.get("In", row.get("in_time", "")))))
            out_time = safe_time(parse_time_br(row.get("OUT", row.get("Out", row.get("out_time", "")))))
            raw_in = str(row.get("IN", row.get("In", row.get("in_time", "")))).strip()
            raw_out = str(row.get("OUT", row.get("Out", row.get("out_time", "")))).strip()
            essl_status = None
            if raw_in.upper() in ESSL_STATUS_MARKERS:
                essl_status = raw_in.upper()
            elif raw_out.upper() in ESSL_STATUS_MARKERS:
                essl_status = raw_out.upper()
            if emp_code and att_date:
                parsed_rows.append((emp_code, emp_name, att_date, in_time, out_time, essl_status))
                all_dates.add(att_date)

        existing_keys = set()
        if parsed_rows and all_dates:
            min_date = min(all_dates)
            max_date = max(all_dates)
            existing_records = db.query(ESSLAttendance.pr_number, ESSLAttendance.date).filter(
                ESSLAttendance.date >= min_date,
                ESSLAttendance.date <= max_date
            ).all()
            for pr, dt in existing_records:
                existing_keys.add((pr, dt.isoformat()))

        processed = 0
        to_insert = []
        to_update = []

        upload_log = UploadLog(file_type="essl", filename=file.filename, file_hash=file_hash, rows_processed=0, status="Processing")
        db.add(upload_log)
        db.flush()  # assigns upload_log.id within this same transaction, without committing yet
        upload_log_id = upload_log.id

        for emp_code, emp_name, att_date, in_time, out_time, essl_status in parsed_rows:
            emp = resolve_employee_fast(emp_code, emp_name)
            pr_number = emp.pr_number if emp else emp_code
            if emp and not emp.emp_code:
                emp.emp_code = emp_code
                emp_by_code[emp_code] = emp

            key = (pr_number, att_date.isoformat())
            if key in existing_keys:
                to_update.append({
                    "pr_number": pr_number,
                    "date": att_date,
                    "in_time": in_time,
                    "out_time": out_time,
                    "status": essl_status,
                })
            else:
                to_insert.append({
                    "employee_id": emp.id if emp else None,
                    "pr_number": pr_number,
                    "emp_code": emp_code,
                    "emp_name": emp_name,
                    "date": att_date,
                    "in_time": in_time,
                    "out_time": out_time,
                    "status": essl_status,
                    "vendor": emp.vendor if emp else "",
                    "store": emp.store if emp else "",
                    "shift": safe_shift(emp.shift) if emp else "G",
                    "upload_log_id": upload_log_id,
                })
            processed += 1

        if to_insert:
            db.execute(ESSLAttendance.__table__.insert(), to_insert)
        for upd in to_update:
            db.query(ESSLAttendance).filter(
                ESSLAttendance.pr_number == upd["pr_number"],
                ESSLAttendance.date == upd["date"]
            ).update({
                ESSLAttendance.in_time: upd["in_time"] or ESSLAttendance.in_time,
                ESSLAttendance.out_time: upd["out_time"] or ESSLAttendance.out_time,
                ESSLAttendance.status: upd["status"] or ESSLAttendance.status,
                ESSLAttendance.upload_log_id: upload_log_id,
            }, synchronize_session=False)

        upload_log.status = "Success"
        upload_log.rows_processed = processed
        db.commit()
        elapsed = round(time_module.time() - start_time, 2)
        return {"status": "success", "type": "essl", "format": "flat", "rows_processed": processed, "elapsed_seconds": elapsed, "filename": file.filename, "upload_log_id": upload_log_id}

    # CROSS-TAB FORMAT
    header_row = raw_rows[header_row_idx]
    emp_code_col = emp_name_col = shift_col = None
    for j, cell in enumerate(header_row):
        if cell is None:
            continue
        cell_str = str(cell).strip().upper()
        if any(x in cell_str for x in ["EMP CODE", "EMP_CODE", "CODE", "PAYCODE", "PAY CODE", "EMP.CODE"]):
            emp_code_col = j
        elif any(x in cell_str for x in ["EMP NAME", "EMP_NAME", "NAME", "EMPLOYEE NAME", "EMPLOYEE_NAME"]):
            emp_name_col = j
        elif cell_str in ["SHIFT", "SFT"]:
            shift_col = j

    if emp_code_col is None:
        emp_code_col = 0
    if emp_name_col is None:
        emp_name_col = 1

    current_emp_code = None
    current_emp_name = ""
    current_shift = "G"
    current_in_row = {}
    current_out_row = {}
    current_other_rows = {}
    all_records = []
    all_dates = set()

    for i in range(header_row_idx + 1, len(raw_rows)):
        row = raw_rows[i]
        if len(row) < 3:
            continue
        first_val = row[emp_code_col] if emp_code_col < len(row) else None
        # FIX v3.1 (preserved): IN/OUT rows repeat the emp code in col A - check if it is a real new employee block
        has_code = False
        if first_val is not None and str(first_val).strip() not in ["", "-", "#N/A", "N/A"]:
            first_str = str(first_val).strip()
            # Only treat as new employee if the value looks like an employee code (alphanumeric, not just a time/status)
            if re.match(r'^[A-Z0-9]+$', first_str) and first_str not in ["IN", "OUT", "TOTAL", "OT", "STATUS", "REMARKS", "REMARK", "SHIFT", "SFT"]:
                has_code = True

        if has_code:
            emp_code = str(first_val).strip()
            if current_emp_code and current_emp_code != emp_code and date_columns:
                # Flush previous employee before starting new one
                for col_idx, att_date in date_columns:
                    date_key = att_date.isoformat()
                    raw_in = current_in_row.get(date_key)
                    raw_out = current_out_row.get(date_key)
                    in_time = parse_time_br(raw_in)
                    out_time = parse_time_br(raw_out)

                    # FIX v3.2: Detect ESSL status from raw IN/OUT values
                    essl_status = None
                    if raw_in and raw_in.upper() in ESSL_STATUS_MARKERS:
                        essl_status = raw_in.upper()
                    elif raw_out and raw_out.upper() in ESSL_STATUS_MARKERS:
                        essl_status = raw_out.upper()

                    raw_punches = {}
                    for rt, vals in current_other_rows.items():
                        if date_key in vals:
                            key = {"TOTAL": "total_hours", "OT": "ot_hours", "STATUS": "status",
                                   "REMARKS": "remark", "REMARK": "remark", "SHIFT": "shift", "SFT": "shift"}.get(rt, rt.lower())
                            raw_punches[key] = vals[date_key]
                    if essl_status:
                        raw_punches["essl_status"] = essl_status

                    shift = current_other_rows.get("SHIFT", {}).get(date_key) or current_other_rows.get("SFT", {}).get(date_key) or current_shift
                    all_records.append((current_emp_code, current_emp_name, shift, date_key, in_time, out_time, essl_status, raw_punches))
                    all_dates.add(att_date)
                # Reset dicts only when flushing previous employee
                current_in_row = {}
                current_out_row = {}
                current_other_rows = {}
            # Set up new employee (or first employee)
            current_emp_code = emp_code
            current_emp_name = str(row[emp_name_col]).strip() if emp_name_col < len(row) and row[emp_name_col] else ""
            current_shift = "G"
            if shift_col is not None and shift_col < len(row) and row[shift_col]:
                current_shift = normalize_shift(str(row[shift_col]).strip())
            # FIX v3.2.2: Don't continue - let code fall through to process IN row data
        if not current_emp_code:
            continue
        row_type = None
        for j in range(min(10, len(row))):
            cell = row[j]
            if cell is not None:
                cell_str = str(cell).strip().upper()
                if cell_str in ["IN", "OUT", "TOTAL", "OT", "STATUS", "REMARKS", "REMARK", "SHIFT", "SFT"]:
                    row_type = cell_str
                    break
        if not row_type:
            continue
        target_dict = current_in_row if row_type == "IN" else current_out_row if row_type == "OUT" else current_other_rows.setdefault(row_type, {})
        for col_idx, att_date in date_columns:
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None or str(val).strip() in ["", "-", "#N/A", "N/A", "NA"]:
                continue
            target_dict[att_date.isoformat()] = str(val).strip()

    # Flush last employee
    if current_emp_code and date_columns:
        for col_idx, att_date in date_columns:
            date_key = att_date.isoformat()
            raw_in = current_in_row.get(date_key)
            raw_out = current_out_row.get(date_key)
            in_time = parse_time_br(raw_in)
            out_time = parse_time_br(raw_out)

            # FIX v3.2: Detect ESSL status from raw IN/OUT values
            essl_status = None
            if raw_in and raw_in.upper() in ESSL_STATUS_MARKERS:
                essl_status = raw_in.upper()
            elif raw_out and raw_out.upper() in ESSL_STATUS_MARKERS:
                essl_status = raw_out.upper()

            raw_punches = {}
            for rt, vals in current_other_rows.items():
                if date_key in vals:
                    key = {"TOTAL": "total_hours", "OT": "ot_hours", "STATUS": "status",
                           "REMARKS": "remark", "REMARK": "remark", "SHIFT": "shift", "SFT": "shift"}.get(rt, rt.lower())
                    raw_punches[key] = vals[date_key]
            if essl_status:
                raw_punches["essl_status"] = essl_status

            shift = current_other_rows.get("SHIFT", {}).get(date_key) or current_other_rows.get("SFT", {}).get(date_key) or current_shift
            all_records.append((current_emp_code, current_emp_name, shift, date_key, in_time, out_time, essl_status, raw_punches))
            all_dates.add(att_date)

    existing_keys = set()
    if all_records and all_dates:
        min_date = min(all_dates)
        max_date = max(all_dates)
        existing_records = db.query(ESSLAttendance.pr_number, ESSLAttendance.date).filter(
            ESSLAttendance.date >= min_date,
            ESSLAttendance.date <= max_date
        ).all()
        for pr, dt in existing_records:
            existing_keys.add((pr, dt.isoformat()))

    processed = 0
    to_insert = []
    to_update = []

    upload_log = UploadLog(file_type="essl", filename=file.filename, file_hash=file_hash, rows_processed=0, status="Processing")
    db.add(upload_log)
    db.flush()
    upload_log_id = upload_log.id

    for emp_code, emp_name, shift, date_key, in_time, out_time, essl_status, raw_punches in all_records:
        emp = resolve_employee_fast(emp_code, emp_name)
        pr_number = emp.pr_number if emp else emp_code
        if emp and not emp.emp_code:
            emp.emp_code = emp_code
            emp_by_code[emp_code] = emp
        att_date = datetime.fromisoformat(date_key).date()
        key = (pr_number, date_key)
        if key in existing_keys:
            to_update.append({
                "pr_number": pr_number,
                "date": att_date,
                "in_time": in_time,
                "out_time": out_time,
                "status": essl_status,
                "raw_punches": json.dumps(raw_punches) if raw_punches else None,
            })
        else:
            to_insert.append({
                "employee_id": emp.id if emp else None,
                "pr_number": pr_number,
                "emp_code": emp_code,
                "emp_name": emp_name,
                "date": att_date,
                "in_time": in_time,
                "out_time": out_time,
                "status": essl_status,
                "raw_punches": json.dumps(raw_punches) if raw_punches else None,
                "vendor": emp.vendor if emp else "",
                "store": emp.store if emp else "",
                "shift": safe_shift(shift),
                "upload_log_id": upload_log_id,
            })
        processed += 1

    if to_insert:
        db.execute(ESSLAttendance.__table__.insert(), to_insert)
    for upd in to_update:
        db.query(ESSLAttendance).filter(
            ESSLAttendance.pr_number == upd["pr_number"],
            ESSLAttendance.date == upd["date"]
        ).update({
            ESSLAttendance.in_time: upd["in_time"] or ESSLAttendance.in_time,
            ESSLAttendance.out_time: upd["out_time"] or ESSLAttendance.out_time,
            ESSLAttendance.status: upd["status"] or ESSLAttendance.status,
            ESSLAttendance.raw_punches: upd["raw_punches"] or ESSLAttendance.raw_punches,
            ESSLAttendance.upload_log_id: upload_log_id,
        }, synchronize_session=False)

    upload_log.status = "Success"
    upload_log.rows_processed = processed
    db.commit()
    elapsed = round(time_module.time() - start_time, 2)

    return {
        "status": "success", "type": "essl", "format": "cross-tab", "rows_processed": processed,
        "employees_found": len(set(r[0] for r in all_records)), "date_columns": len(date_columns),
        "elapsed_seconds": elapsed, "filename": file.filename, "upload_log_id": upload_log_id
    }

@app.post("/api/v1/upload/tata")
def upload_tata(file: UploadFile = File(...), force: bool = Form(False), db: Session = Depends(get_db)):
    # NOTE: sync `def` (not async) so this runs in a worker thread, not the event loop -
    # fixes ERR_HTTP2_PING_FAILED / misleading "CORS" errors on large 30-sheet files.
    import time as time_module
    start_time = time_module.time()

    contents = file.file.read()
    file_hash = hashlib.sha256(contents).hexdigest()
    if not force:
        existing = db.query(UploadLog).filter(UploadLog.file_hash == file_hash).first()
        if existing:
            return {"status": "duplicate", "message": "Already uploaded", "previous_id": existing.id}
    try:
        read_result = read_excel_all_sheets_to_dicts(contents, required_any_header=["PayCode", "Pay Code"])
        rows = read_result["rows"]
        sheets_read = read_result["sheets_read"]
        sheets_skipped = read_result["sheets_skipped"]
    except Exception as e:
        db.add(UploadLog(file_type="tata", filename=file.filename, file_hash=file_hash, rows_processed=0, status="Failed", error_message=str(e)))
        db.commit()
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {str(e)}")

    # Prefetch every employee ONCE instead of up to 3 SELECTs per row (this was the
    # main cause of the timeout - a 30-sheet file could mean 30,000+ rows x 3 queries).
    all_employees = db.query(Employee).all()
    emp_by_pr = {e.pr_number: e for e in all_employees if e.pr_number}
    emp_by_code = {e.emp_code: e for e in all_employees if e.emp_code}
    emp_by_name = {e.name.strip().upper(): e for e in all_employees if e.name}

    def resolve_employee_fast(paycode: str, emp_name_val: str):
        e = emp_by_pr.get(paycode)
        if e:
            return e
        e = emp_by_code.get(paycode)
        if e:
            return e
        if emp_name_val:
            return emp_by_name.get(emp_name_val.strip().upper())
        return None

    # Pass 1: parse all rows in memory (no DB calls yet), track which dates are touched.
    parsed = []
    all_dates = set()
    for row in rows:
        paycode = str(row.get("PayCode", row.get("Pay Code", row.get("paycode", "")))).strip()
        att_date = parse_date_br(row.get("Date", row.get("date", "")))
        if not paycode or not att_date:
            continue
        parsed.append({
            "paycode": paycode,
            "emp_name": str(row.get("Employee Name", row.get("EMP Name", row.get("EmployeeName", "")))).strip(),
            "att_date": att_date,
            "in_time": safe_time(parse_time_br(row.get("In Time", row.get("In_Time", row.get("IN", ""))))),
            "out_time": safe_time(parse_time_br(row.get("Out Time", row.get("Out_Time", row.get("OUT", ""))))),
            "man_hrs": float(row.get("Man Hrs", row.get("Man_Hrs", 0)) or 0),
            "status_val": str(row.get("Status", row.get("status", "P"))).strip().upper(),
            "dept_from_row": str(row.get("Department", row.get("department", ""))).strip(),
            "division_from_row": str(row.get("Division", "")).strip(),
            "row_shift_raw": row.get("Shift", row.get("Shift In Time", None)),
            "category_raw": str(row.get("WC/BC", "")).strip(),
            "contractor_raw": str(row.get("Contractor", "")).strip(),
            "store_raw": str(row.get("Store", "")).strip(),
            "early_going": str(row.get("Early Going", "")).strip(),
            "shift_late": str(row.get("Shift Late", "")).strip(),
        })
        all_dates.add(att_date)

    # Pass 2: bulk-fetch existing TataAttendance rows for just the date range touched
    # (ONE query instead of one SELECT per row).
    existing_map = {}
    if parsed:
        min_date, max_date = min(all_dates), max(all_dates)
        for rec in db.query(TataAttendance).filter(TataAttendance.date >= min_date, TataAttendance.date <= max_date).all():
            existing_map[(rec.pr_number, rec.date)] = rec

    processed = 0
    pending_inserts = {}  # key -> row dict, for new rows not yet in the DB or existing_map

    upload_log = UploadLog(file_type="tata", filename=file.filename, file_hash=file_hash, rows_processed=0, status="Processing")
    db.add(upload_log)
    db.flush()
    upload_log_id = upload_log.id

    for item in parsed:
        paycode = item["paycode"]; att_date = item["att_date"]
        in_time = item["in_time"]; out_time = item["out_time"]; man_hrs = item["man_hrs"]
        status_val = item["status_val"]; dept_from_row = item["dept_from_row"]; division_from_row = item["division_from_row"]

        emp = resolve_employee_fast(paycode, item["emp_name"])
        pr_number = emp.pr_number if emp else paycode

        if item["row_shift_raw"]:
            shift = safe_shift(str(item["row_shift_raw"]))
        else:
            shift = emp.shift if emp else "G"

        category = normalize_category(item["category_raw"]).strip() or (emp.wc if emp else "BC")
        if category == "":
            category = emp.wc if emp else "BC"

        if emp:
            dept_val = dept_from_row or division_from_row
            if dept_val and emp.department != dept_val:
                emp.department = dept_val
            safe_emp_shift = safe_shift(shift)
            if safe_emp_shift and emp.shift != safe_emp_shift:
                emp.shift = safe_emp_shift

        if status_val not in ["P", "A", "L", "HD", "WO", "SP"]:
            status_val = "P" if (in_time or out_time or man_hrs > 0) else "A"

        ot_result = calculate_ot(shift, category, in_time, out_time, man_hrs)
        key = (pr_number, att_date)
        existing_tata = existing_map.get(key)

        if existing_tata:
            existing_tata.in_time = safe_time(in_time) or existing_tata.in_time
            existing_tata.out_time = safe_time(out_time) or existing_tata.out_time
            existing_tata.man_hrs = man_hrs if man_hrs > 0 else existing_tata.man_hrs
            existing_tata.status = status_val
            existing_tata.ot_hours = ot_result["calculated_ot_hours"]
            existing_tata.shift = safe_shift(shift)
            existing_tata.department = dept_from_row or division_from_row or existing_tata.department
            existing_tata.upload_log_id = upload_log_id
        elif key in pending_inserts:
            # duplicate paycode+date appearing again later in the same file - update in place
            pend = pending_inserts[key]
            pend["in_time"] = safe_time(in_time) or pend["in_time"]
            pend["out_time"] = safe_time(out_time) or pend["out_time"]
            pend["man_hrs"] = man_hrs if man_hrs > 0 else pend["man_hrs"]
            pend["status"] = status_val
            pend["ot_hours"] = ot_result["calculated_ot_hours"]
            pend["shift"] = safe_shift(shift)
            pend["department"] = dept_from_row or division_from_row or pend["department"]
        else:
            pending_inserts[key] = dict(
                employee_id=emp.id if emp else None, pr_number=pr_number, emp_code=paycode, emp_name=item["emp_name"],
                date=att_date, in_time=safe_time(in_time), out_time=safe_time(out_time), man_hrs=man_hrs, status=status_val,
                ot_hours=ot_result["calculated_ot_hours"],
                early_going=item["early_going"], shift_late=item["shift_late"],
                vendor=normalize_vendor(item["contractor_raw"] or (emp.vendor if emp else "")),
                store=item["store_raw"] or (emp.store if emp else ""),
                department=dept_from_row or division_from_row or (emp.department if emp else ""),
                shift=safe_shift(shift),
                upload_log_id=upload_log_id
            )
        processed += 1

    if pending_inserts:
        db.execute(TataAttendance.__table__.insert(), list(pending_inserts.values()))

    upload_log.status = "Success"
    upload_log.rows_processed = processed
    db.commit()
    elapsed = round(time_module.time() - start_time, 2)
    return {"status": "success", "type": "tata", "rows_processed": processed,
            "sheets_read": sheets_read, "sheets_skipped": sheets_skipped, "elapsed_seconds": elapsed,
            "filename": file.filename, "message": "Official attendance stored. Ready for reconciliation.",
            "upload_log_id": upload_log_id}

@app.post("/api/v1/upload/tata-all")
def upload_tata_all(file: UploadFile = File(...), force: bool = Form(False), db: Session = Depends(get_db)):
    contents = file.file.read()
    file_hash = hashlib.sha256(contents).hexdigest()
    if not force:
        existing = db.query(UploadLog).filter(UploadLog.file_hash == file_hash).first()
        if existing:
            return {"status": "duplicate", "message": "Already uploaded", "previous_id": existing.id}
    try:
        read_result = read_excel_all_sheets_to_dicts(contents, required_any_header=["PayCode", "Pay Code"])
        rows = read_result["rows"]
    except Exception as e:
        db.add(UploadLog(file_type="tata_all", filename=file.filename, file_hash=file_hash, rows_processed=0, status="Failed", error_message=str(e)))
        db.commit()
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {str(e)}")

    processed = 0
    monthly_data = {}
    for row in rows:
        paycode = str(row.get("PayCode", row.get("Pay Code", ""))).strip()
        emp_name = str(row.get("Employee Name", row.get("EMP Name", ""))).strip()
        att_date = parse_date_br(row.get("Date", row.get("date", "")))
        status_val = str(row.get("Status", row.get("status", ""))).strip().upper()

        if not paycode or not att_date:
            continue

        year_month = att_date.strftime("%Y-%m")
        key = (paycode, year_month)
        if key not in monthly_data:
            monthly_data[key] = {"pr": paycode, "emp_name": emp_name, "year_month": year_month,
                                  "present": 0, "absent": 0, "leave": 0, "weekoff": 0, "half_day": 0, "total_ot": 0.0}

        if status_val == "P": monthly_data[key]["present"] += 1
        elif status_val == "A": monthly_data[key]["absent"] += 1
        elif status_val == "L": monthly_data[key]["leave"] += 1
        elif status_val == "WO": monthly_data[key]["weekoff"] += 1
        elif status_val == "HD": monthly_data[key]["half_day"] += 1; monthly_data[key]["present"] += 1

        ot = float(row.get("OT", 0) or 0)
        monthly_data[key]["total_ot"] += ot
        processed += 1

    # Prefetch existing monthly rows for just the year-months touched (one query instead
    # of one SELECT per unique employee/month key).
    year_months_touched = {ym for (_, ym) in monthly_data.keys()}
    existing_monthly_map = {}
    if year_months_touched:
        for rec in db.query(MonthlyTataAttendance).filter(MonthlyTataAttendance.year_month.in_(year_months_touched)).all():
            existing_monthly_map[(rec.pr_number, rec.year_month)] = rec

    for key, data in monthly_data.items():
        total_days = data["present"] + data["absent"] + data["leave"] + data["weekoff"] + data["half_day"]
        att_pct = (data["present"] / total_days * 100) if total_days > 0 else 0
        existing_monthly = existing_monthly_map.get(key)
        if existing_monthly:
            existing_monthly.present_days = data["present"]; existing_monthly.absent_days = data["absent"]
            existing_monthly.leave_days = data["leave"]; existing_monthly.weekoff_days = data["weekoff"]
            existing_monthly.half_days = data["half_day"]; existing_monthly.total_days = total_days
            existing_monthly.attendance_percentage = att_pct; existing_monthly.total_ot_hours = data["total_ot"]
        else:
            db.add(MonthlyTataAttendance(
                pr_number=data["pr"], emp_name=data["emp_name"], year_month=data["year_month"],
                total_days=total_days, present_days=data["present"], absent_days=data["absent"],
                leave_days=data["leave"], weekoff_days=data["weekoff"], half_days=data["half_day"],
                attendance_percentage=att_pct, total_ot_hours=data["total_ot"]
            ))

    db.commit()
    log_result = safe_log_upload(db, "tata_all", file.filename, file_hash, len(monthly_data), "Success")
    return {"status": "success", "type": "tata_all", "monthly_records_created": len(monthly_data),
            "filename": file.filename, "message": "Monthly attendance summary stored for trends and reports.",
            "upload_log": log_result}

# ============================================================================
# DAY-WISE COMBINED UPLOAD (ESSL In / ESSL Out / Tata in one workbook)
# ============================================================================

def _find_sheet_by_keywords(wb, must_contain: List[str], must_not_contain: Optional[List[str]] = None):
    """Case/space-insensitive sheet name match. e.g. must_contain=['essl','in'] matches
    'Essl In', 'ESSL_IN', 'essl-in', etc."""
    must_not_contain = must_not_contain or []
    for ws in wb.worksheets:
        name_norm = re.sub(r'[^a-z]', '', ws.title.lower())
        if all(kw in name_norm for kw in must_contain) and not any(kw in name_norm for kw in must_not_contain):
            return ws
    return None

def _parse_essl_side_sheet(ws) -> Dict[str, Dict[str, Any]]:
    """Parses an 'Essl In' or 'Essl Out' sheet shaped like:
    Bio | PRN | Name | Days | <single date column of times/status markers>
    Returns {pr_number: {"name": ..., "value": <time str or status marker or None>}}"""
    rows = list(ws.iter_rows(values_only=True))
    out: Dict[str, Dict[str, Any]] = {}
    if not rows:
        return out
    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue
        prn = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
        name = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        if not prn:
            continue
        raw_val = row[-1] if len(row) > 3 else None
        raw_str = str(raw_val).strip() if raw_val is not None else ""
        status_marker = raw_str.upper() if raw_str.upper() in ESSL_STATUS_MARKERS else None
        time_val = parse_time_br(raw_val) if not status_marker else None
        out[prn] = {"name": name, "time": time_val, "status": status_marker}
    return out

@app.post("/api/v1/upload/daywise")
def upload_daywise(file: UploadFile = File(...), target_date: Optional[str] = Form(None), force: bool = Form(False), db: Session = Depends(get_db)):
    """Single-day combined upload: one workbook with 'Essl In', 'Essl Out', and 'Tata'
    sheets (exactly the format HR pulls each day). ESSL sheets carry one time column per
    day (header like '22 W'), which doesn't carry an unambiguous date on its own, so the
    date is taken from the Tata sheet's own Date column unless target_date is passed
    explicitly (accepts 'DD/MM/YYYY' or 'YYYY-MM-DD')."""
    contents = file.file.read()
    file_hash = hashlib.sha256(contents).hexdigest()
    if not force:
        existing = db.query(UploadLog).filter(UploadLog.file_hash == file_hash).first()
        if existing:
            return {"status": "duplicate", "message": "Already uploaded", "previous_id": existing.id}

    try:
        wb = load_workbook(filename=bio.BytesIO(contents), read_only=True, data_only=True)
    except Exception as e:
        db.add(UploadLog(file_type="daywise", filename=file.filename, file_hash=file_hash, rows_processed=0, status="Failed", error_message=str(e)))
        db.commit()
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {str(e)}")

    essl_in_ws = _find_sheet_by_keywords(wb, ["essl", "in"], must_not_contain=["out"])
    essl_out_ws = _find_sheet_by_keywords(wb, ["essl", "out"])
    tata_ws = _find_sheet_by_keywords(wb, ["tata"])

    if essl_in_ws is None and essl_out_ws is None and tata_ws is None:
        wb.close()
        raise HTTPException(status_code=400, detail="Couldn't find 'Essl In', 'Essl Out', or 'Tata' sheets in this workbook. Expected exactly those (case/spacing doesn't matter).")

    # ---- Parse Tata sheet first (also used to auto-detect the date) ----
    tata_rows: List[Dict[str, Any]] = []
    if tata_ws is not None:
        raw = list(tata_ws.iter_rows(values_only=True))
        if raw:
            headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(raw[0])]
            for r in raw[1:]:
                if all(v is None for v in r):
                    continue
                tata_rows.append({headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))})

    resolved_date = parse_date_br(target_date) if target_date else None
    if resolved_date is None and tata_rows:
        date_counts: Dict[date, int] = {}
        for r in tata_rows:
            d = parse_date_br(r.get("Date", r.get("date")))
            if d:
                date_counts[d] = date_counts.get(d, 0) + 1
        if date_counts:
            resolved_date = max(date_counts, key=date_counts.get)

    if resolved_date is None:
        wb.close()
        raise HTTPException(status_code=400, detail="Couldn't determine the attendance date. Pass target_date (DD/MM/YYYY) explicitly, or make sure the Tata sheet's Date column is populated.")

    all_employees = db.query(Employee).all()
    emp_by_pr = {e.pr_number: e for e in all_employees if e.pr_number}
    emp_by_name = {e.name.strip().upper(): e for e in all_employees if e.name}

    def resolve_emp(code_val: Optional[str], name_val: Optional[str]):
        if code_val and code_val in emp_by_pr:
            return emp_by_pr[code_val]
        if name_val and name_val.strip().upper() in emp_by_name:
            return emp_by_name[name_val.strip().upper()]
        return None

    upload_log = UploadLog(file_type="daywise", filename=file.filename, file_hash=file_hash, rows_processed=0, status="Processing")
    db.add(upload_log)
    db.flush()
    upload_log_id = upload_log.id

    # ---- ESSL In / Out ----
    essl_processed = 0
    if essl_in_ws is not None or essl_out_ws is not None:
        in_data = _parse_essl_side_sheet(essl_in_ws) if essl_in_ws is not None else {}
        out_data = _parse_essl_side_sheet(essl_out_ws) if essl_out_ws is not None else {}
        all_prns = set(in_data.keys()) | set(out_data.keys())
        for prn in all_prns:
            in_rec = in_data.get(prn, {})
            out_rec = out_data.get(prn, {})
            name_val = in_rec.get("name") or out_rec.get("name") or ""
            in_time = in_rec.get("time")
            out_time = out_rec.get("time")
            essl_status = in_rec.get("status") or out_rec.get("status")
            emp = resolve_emp(prn, name_val)

            existing_essl = db.query(ESSLAttendance).filter(ESSLAttendance.pr_number == prn, ESSLAttendance.date == resolved_date).first()
            if existing_essl:
                existing_essl.in_time = safe_time(in_time) or existing_essl.in_time
                existing_essl.out_time = safe_time(out_time) or existing_essl.out_time
                existing_essl.status = safe_shift(essl_status) or existing_essl.status
                existing_essl.upload_log_id = upload_log_id
            else:
                db.add(ESSLAttendance(
                    employee_id=emp.id if emp else None, pr_number=prn, emp_code=prn, emp_name=name_val,
                    date=resolved_date, in_time=safe_time(in_time), out_time=safe_time(out_time), status=safe_shift(essl_status) if essl_status else None,
                    vendor=emp.vendor if emp else "", store=emp.store if emp else "", shift=safe_shift(emp.shift) if emp else "G",
                    upload_log_id=upload_log_id
                ))
            essl_processed += 1
            if essl_processed % 500 == 0:
                db.commit()
        db.commit()

    # ---- Tata (same upsert logic as /api/v1/upload/tata, scoped to this file) ----
    existing_tata_map = {
        rec.pr_number: rec for rec in db.query(TataAttendance).filter(TataAttendance.date == resolved_date).all()
    }
    tata_processed = 0
    for row in tata_rows:
        paycode = str(row.get("PayCode", row.get("Pay Code", ""))).strip()
        emp_name = str(row.get("Employee Name", row.get("EMP Name", ""))).strip()
        att_date = parse_date_br(row.get("Date", row.get("date", ""))) or resolved_date
        in_time = safe_time(parse_time_br(row.get("In Time", row.get("In_Time", row.get("IN", "")))))
        out_time = safe_time(parse_time_br(row.get("Out Time", row.get("Out_Time", row.get("OUT", "")))))
        man_hrs = float(row.get("Man Hrs", row.get("Man_Hrs", 0)) or 0)
        status_val = str(row.get("Status", row.get("status", "P"))).strip().upper()
        dept_from_row = str(row.get("Department", row.get("department", ""))).strip()
        division_from_row = str(row.get("Division", "")).strip()

        if not paycode or not att_date:
            continue

        emp = resolve_emp(paycode, emp_name)
        pr_number = emp.pr_number if emp else paycode

        row_shift_raw = row.get("Shift", row.get("Shift In Time", None))
        shift = safe_shift(str(row_shift_raw)) if row_shift_raw else (safe_shift(emp.shift) if emp else "G")
        category = normalize_category(str(row.get("WC/BC", ""))).strip() or (emp.wc if emp else "BC")
        if category == "":
            category = emp.wc if emp else "BC"

        if emp:
            dept_val = dept_from_row or division_from_row
            if dept_val and emp.department != dept_val:
                emp.department = dept_val
            safe_emp_shift = safe_shift(shift)
            if safe_emp_shift and emp.shift != safe_emp_shift:
                emp.shift = safe_emp_shift

        if status_val not in ["P", "A", "L", "HD", "WO", "SP"]:
            status_val = "P" if (in_time or out_time or man_hrs > 0) else "A"

        existing_tata = existing_tata_map.get(pr_number)
        ot_result = calculate_ot(shift, category, in_time, out_time, man_hrs)

        if existing_tata:
            existing_tata.in_time = safe_time(in_time) or existing_tata.in_time
            existing_tata.out_time = safe_time(out_time) or existing_tata.out_time
            existing_tata.man_hrs = man_hrs if man_hrs > 0 else existing_tata.man_hrs
            existing_tata.status = status_val
            existing_tata.ot_hours = ot_result["calculated_ot_hours"]
            existing_tata.shift = safe_shift(shift)
            existing_tata.department = dept_from_row or division_from_row or existing_tata.department
            existing_tata.upload_log_id = upload_log_id
        else:
            new_tata = TataAttendance(
                employee_id=emp.id if emp else None, pr_number=pr_number, emp_code=paycode, emp_name=emp_name,
                date=att_date, in_time=safe_time(in_time), out_time=safe_time(out_time), man_hrs=man_hrs, status=status_val,
                ot_hours=ot_result["calculated_ot_hours"],
                early_going=str(row.get("Early Going", "")).strip(), shift_late=str(row.get("Shift Late", "")).strip(),
                vendor=normalize_vendor(str(row.get("Contractor", emp.vendor if emp else ""))),
                store=str(row.get("Store", emp.store if emp else "")).strip(),
                department=dept_from_row or division_from_row or (emp.department if emp else ""),
                shift=safe_shift(shift),
                upload_log_id=upload_log_id
            )
            db.add(new_tata)
            existing_tata_map[pr_number] = new_tata  # guard against a dup paycode later in the same file
        tata_processed += 1

    try:
        db.commit()
    except Exception as commit_err:
        db.rollback()
        import traceback
        err_detail = traceback.format_exc()
        # Log what we were trying to insert for debugging
        print(f"[DAYWISE COMMIT ERROR] {commit_err}")
        print(f"[DAYWISE COMMIT ERROR] resolved_date={resolved_date}, essl_processed={essl_processed}, tata_processed={tata_processed}")
        # Check for long strings in pending objects
        for obj in db.new:
            if hasattr(obj, '__tablename__'):
                for col in obj.__table__.columns:
                    val = getattr(obj, col.name, None)
                    if val and isinstance(val, str) and len(val) > 50:
                        print(f"[DAYWISE LONG STRING] {obj.__tablename__}.{col.name} = {val[:100]}...")
        raise HTTPException(status_code=500, detail=f"Database commit failed: {str(commit_err)}")

    # Data is already committed above at this point. wb.close() failing here must
    # NOT be allowed to turn into an unhandled 500 that hides a successful upload -
    # so it's wrapped rather than left to raise directly like before.
    try:
        wb.close()
    except Exception as close_err:
        print(f"[DAYWISE] wb.close() failed after successful commit (non-fatal): {close_err}")

    upload_log.status = "Success"
    upload_log.rows_processed = essl_processed + tata_processed
    db.commit()

    return {
        "status": "success", "type": "daywise", "date": resolved_date.isoformat(),
        "sheets_found": {"essl_in": essl_in_ws is not None, "essl_out": essl_out_ws is not None, "tata": tata_ws is not None},
        "essl_rows_processed": essl_processed, "tata_rows_processed": tata_processed,
        "filename": file.filename,
        "message": f"Day-wise attendance stored for {resolved_date.isoformat()}. Run reconciliation for this date next.",
        "upload_log_id": upload_log_id
    }

# ============================================================================
# RECONCILIATION ENDPOINTS
# ============================================================================

def _get_essl_status(essl_record) -> Optional[str]:
    """FIX v3.2: Extract ESSL status from record (new status column or raw_punches fallback)."""
    if not essl_record:
        return None
    if essl_record.status and essl_record.status.upper() in ESSL_STATUS_MARKERS:
        return essl_record.status.upper()
    if essl_record.raw_punches:
        try:
            rp = json.loads(essl_record.raw_punches) if isinstance(essl_record.raw_punches, str) else essl_record.raw_punches
            status = rp.get("essl_status") if rp else None
            if status and status.upper() in ESSL_STATUS_MARKERS:
                return status.upper()
        except:
            pass
    return None

def _reconcile_single_date(d: date, db: Session) -> Dict[str, Any]:
    """
    Core reconciliation logic for one date. FIX v3.2: Uses batched queries for performance.
    """
    employees = db.query(Employee).filter(Employee.status == "ACTIVE").all()
    emp_ids = [e.id for e in employees]
    emp_by_id = {e.id: e for e in employees}

    # FIX v3.2: Pre-fetch ALL related records for this date in bulk queries
    # FIX v3.2.1: Use pr_number instead of employee_id for reliable matching
    all_pr_numbers = [e.pr_number for e in employees if e.pr_number]

    # TATA-ONLY DETECTION: employees with Tata data but NO ESSL data for this date
    tata_only_prs = set()
    tata_records_for_date = db.query(TataAttendance.pr_number).filter(
        TataAttendance.pr_number.in_(all_pr_numbers),
        TataAttendance.date == d
    ).distinct().all()
    tata_prs_today = {r[0] for r in tata_records_for_date if r[0]}
    essl_records_for_date = db.query(ESSLAttendance.pr_number).filter(
        ESSLAttendance.pr_number.in_(all_pr_numbers),
        ESSLAttendance.date == d
    ).distinct().all()
    essl_prs_today = {r[0] for r in essl_records_for_date if r[0]}
    tata_only_prs = tata_prs_today - essl_prs_today

    all_essl = db.query(ESSLAttendance).filter(
        ESSLAttendance.pr_number.in_(all_pr_numbers),
        ESSLAttendance.date == d
    ).all()
    essl_map = {e.pr_number: e for e in all_essl}

    all_tata = db.query(TataAttendance).filter(
        TataAttendance.pr_number.in_(all_pr_numbers),
        TataAttendance.date == d
    ).all()
    tata_map = {t.pr_number: t for t in all_tata}

    all_att = db.query(Attendance).filter(
        Attendance.pr_number.in_(all_pr_numbers),
        Attendance.date == d
    ).all()
    att_map = {a.pr_number: a for a in all_att}

    all_recon = db.query(AttendanceReconciliation).filter(
        AttendanceReconciliation.pr_number.in_([e.pr_number for e in employees]),
        AttendanceReconciliation.date == d
    ).all()
    recon_map = {(r.pr_number, r.date.isoformat()): r for r in all_recon}

    all_actions = db.query(HRAction).filter(
        HRAction.pr_number.in_([e.pr_number for e in employees]),
        HRAction.date == d
    ).all()
    action_map = {(a.pr_number, a.date.isoformat(), a.action_type): a for a in all_actions}

    all_ot = db.query(Overtime).filter(
        Overtime.pr_number.in_([e.pr_number for e in employees]),
        Overtime.date == d
    ).all()
    ot_map = {(o.pr_number, o.date.isoformat()): o for o in all_ot}

    attendance_created = 0
    reconciliation_issues = 0
    hr_actions_created = 0

    for emp in employees:
        essl = essl_map.get(emp.pr_number)
        tata = tata_map.get(emp.pr_number)

        essl_in = essl.in_time if essl else None
        essl_out = essl.out_time if essl else None
        tata_in = tata.in_time if tata else None
        tata_out = tata.out_time if tata else None
        tata_man_hrs = tata.man_hrs if tata else 0
        tata_status = tata.status if tata else None
        essl_status = _get_essl_status(essl)

        # Sunday handling: Sundays default to Week Off unless Tata shows the employee
        # actually worked (has a Tata record) - see Present override further below.
        is_sunday = d.weekday() == 6  # Monday=0 ... Sunday=6

        # FIX v3.2.8: If employee has NO ESSL and NO Tata data, create "Absent" record with issue "No Data"
        if not essl and not tata:
            shift = safe_shift(emp.shift) if emp.shift else "G"
            category = emp.wc or "BC"
            if is_sunday:
                display_status = "Week Off"
                issue = "Week Off"
                match_status = "Week Off"
                remark = "Sunday - Week Off (no attendance data)"
            else:
                display_status = "Absent"
                issue = "No Data"
                match_status = "No Data"
                remark = "No ESSL or Tata record found for this date"

            existing_att = att_map.get(emp.pr_number)
            if existing_att:
                existing_att.essl_in = None
                existing_att.essl_out = None
                existing_att.tata_in = None
                existing_att.tata_out = None
                existing_att.final_in = None
                existing_att.final_out = None
                existing_att.worked_hours = 0
                existing_att.man_hrs = 0
                existing_att.ot_hours = 0
                existing_att.ot_headcount = 0
                existing_att.attendance_status = display_status
                existing_att.late_minutes = 0
                existing_att.early_minutes = 0
                existing_att.single_punch = "No"
                existing_att.is_match = "No"
                existing_att.match_status = match_status
                existing_att.shift = shift
                existing_att.category = category
                existing_att.issue = issue
                existing_att.remark = remark
            else:
                db.add(Attendance(
                    employee_id=emp.id, pr_number=emp.pr_number, emp_code=emp.emp_code, date=d,
                    essl_in=None, essl_out=None, tata_in=None, tata_out=None,
                    final_in=None, final_out=None, worked_hours=0, man_hrs=0,
                    ot_hours=0, ot_headcount=0, attendance_status=display_status,
                    late_minutes=0, early_minutes=0, single_punch="No",
                    is_match="No", match_status=match_status,
                    vendor=emp.vendor, store=emp.store, department=emp.department,
                    shift=shift, category=category, remark=remark, issue=issue, source="reconciliation"
                ))
                attendance_created += 1
            continue

        # Continue with existing logic for employees who HAVE ESSL or Tata data

        # FIX v3.2.7: For WO employees, shift stays empty. For absent, use master shift.
        has_essl_punches = bool(essl_in or essl_out)
        has_tata_punches = bool(tata_in or tata_out)
        has_any_punches = has_essl_punches or has_tata_punches

        if essl_status == "WO" and not has_any_punches:
            # Week Off: no shift shown, status will be "Week Off"
            shift = None
        elif not has_any_punches:
            # Regular absent: use assigned shift from master
            shift = safe_shift(emp.shift) if emp.shift else "G"
        else:
            shift = safe_shift((tata.shift if tata and tata.shift else None) or (essl.shift if essl and essl.shift else None) or emp.shift) or "G"
        category = emp.wc or "BC"
        match_status = "Matched"
        match_delta_in = 0
        match_delta_out = 0
        reconciliation_severity = "Low"
        remark = ""

        has_essl = bool(essl_in or essl_out or essl_status)
        has_tata = bool(tata_in or tata_out)
        # FIX v3.2.5: Check actual punches separately from status
        has_essl_punches = bool(essl_in or essl_out)
        has_tata_punches = bool(tata_in or tata_out)
        has_any_punches = has_essl_punches or has_tata_punches

        # FIX v3.2.5: WO = Absent only if NO actual punches from either source
        if essl_status == "WO" and not has_any_punches:
            match_status = "Absent"
            reconciliation_severity = "Low"
            remark = "ESSL shows WO (treated as Absent - no punches)"
        elif essl_status == "WO" and has_any_punches:
            # ESSL says WO but punches exist → Present, flag mismatch
            match_status = "Mismatch"
            reconciliation_severity = "Medium"
            remark = "ESSL shows WO but punches exist"
        elif not has_essl and not has_tata:
            match_status = "No Data"
            reconciliation_severity = "Critical"
        elif emp.pr_number in tata_only_prs:
            # TATA-ONLY: Employee has no ESSL data but has Tata data → use Tata only, not an error
            match_status = "Tata Only"
            reconciliation_severity = "Low"
            remark = "Tata-only employee (no ESSL data)"
        elif not has_essl and has_tata:
            match_status = "Missing ESSL"
            reconciliation_severity = "High"
            remark = "No Punch ESSL"
        elif has_essl and not has_tata:
            match_status = "Missing Tata"
            reconciliation_severity = "High"
        elif has_essl and has_tata:
            if essl_in and tata_in:
                match_delta_in = abs(time_to_minutes(essl_in) - time_to_minutes(tata_in))
            if essl_out and tata_out:
                match_delta_out = abs(time_to_minutes(essl_out) - time_to_minutes(tata_out))
            if match_delta_in > 15 or match_delta_out > 15:
                match_status = "Mismatch"
                reconciliation_severity = "Medium"
            else:
                match_status = "Matched"
                reconciliation_severity = "Low"

        if match_status != "Matched":
            recon_key = (emp.pr_number, d.isoformat())
            if recon_key not in recon_map:
                db.add(AttendanceReconciliation(
                    pr_number=emp.pr_number, emp_name=emp.name, date=d,
                    essl_in=essl_in, essl_out=essl_out, tata_in=tata_in, tata_out=tata_out,
                    in_delta_minutes=match_delta_in, out_delta_minutes=match_delta_out,
                    match_status=match_status, severity=reconciliation_severity,
                    remarks=f"ESSL: {essl_in}-{essl_out}, Tata: {tata_in}-{tata_out}"
                ))
                reconciliation_issues += 1

        final_in = tata_in or essl_in
        final_out = tata_out or essl_out
        punch_count = sum([1 for p in [final_in, final_out] if p])
        single_punch = "Yes" if punch_count == 1 else "No"

        # TATA PRIORITY: Determine which source to use for OT/worked_hours calculation
        tata_has_punches = bool(tata_in or tata_out)
        tata_has_man_hrs = bool(tata_man_hrs and tata_man_hrs > 0)
        essl_has_punches = bool(essl_in or essl_out)

        # Priority: Tata punches > Tata man_hrs > ESSL punches > 0
        if tata_has_punches:
            # Tata has punches → calculate OT from Tata punches
            ot_in_time, ot_out_time = tata_in, tata_out
            force_man_hrs = False
        elif tata_has_man_hrs:
            # Tata has no punches but has man_hrs → use man_hrs for worked_hours
            ot_in_time, ot_out_time = None, None
            force_man_hrs = True
        elif essl_has_punches:
            # No Tata data → use ESSL punches
            ot_in_time, ot_out_time = essl_in, essl_out
            force_man_hrs = False
        else:
            ot_in_time, ot_out_time = None, None
            force_man_hrs = False

        summary = calculate_daily_summary(shift, category, tata_status, ot_in_time, ot_out_time, tata_man_hrs, force_man_hrs=force_man_hrs)
        worked_hours = summary["worked_hours"]
        attendance_status = summary["status"]
        late_minutes = summary["late_minutes"]
        ot_hours = summary["calculated_ot_hours"]
        ot_headcount = summary["ot_headcount"]

        status_map = {"P": "Present", "A": "Absent", "SP": "Single Punch", "L": "Leave", "HD": "Half Day", "WO": "Week Off"}
        display_status = status_map.get(attendance_status, attendance_status)

        # FIX v3.2.7: Override to "Week Off" if ESSL says WO and no punches
        if essl_status == "WO" and not has_any_punches:
            display_status = "Week Off"

        # Sunday + Tata data present -> employee worked, must show Present regardless of
        # what the generic shift/hours calculation produced.
        if is_sunday and tata:
            display_status = "Present"
            attendance_status = "P"

        # FIX v3.2.4: Status stays "Present" for Late Punch, Early Departure, etc.
        # The Issue column captures the specific problem. Only Present/Absent as main status.
        # FIX v3.2.10: Override to Half Day if worked hours <= threshold
        half_day_limit = get_half_day_threshold(shift)
        if display_status == "Present" and worked_hours > 0 and worked_hours <= half_day_limit:
            display_status = "Half Day"
            attendance_status = "HD"

        if summary["is_late_punch"] and display_status == "Present":
            pass  # Status stays "Present", issue = "Late Punch"
        elif summary["is_early_departure"] and display_status == "Present":
            pass  # Status stays "Present", issue = "Early Departure"
        elif summary["is_less_working_hours"] and display_status == "Present":
            pass  # Status stays "Present", issue = "Less Working Hours"
        # FIX v3.2.1: Removed "Missing ESSL" override. If Tata has punches, employee is Present.
        # match_status already captures "Missing ESSL" for reporting.
        # FIX v3.2.1: Compute Issue field for HR actionability
        issue = "-"
        if display_status == "Week Off":
            issue = "Week Off"
        elif display_status == "Absent":
            issue = "Absent"
        elif essl_status == "WO" and has_tata:
            issue = "ESSL WO vs Tata Punches"
        elif emp.pr_number in tata_only_prs:
            issue = "-"
        elif not has_essl and has_tata:
            issue = "Missing ESSL Punch"
        elif has_essl and not has_tata:
            issue = "Missing Tata Punch"
        elif match_status == "Mismatch":
            if abs(match_delta_in) > 30 or abs(match_delta_out) > 30:
                issue = "Time Difference (>30 min)"
            else:
                issue = "Time Difference"
        # FIX v3.2.4: Check summary flags instead of display_status for sub-issues
        elif summary["is_single_punch"]:
            issue = "Single Punch"
        elif summary["is_late_punch"]:
            issue = "Late Punch"
        elif summary["is_early_departure"]:
            issue = "Early Departure"
        elif summary["is_less_working_hours"]:
            issue = "Less Working Hours"

        # Check alternate shift
        # FIX v3.2.2: Only flag if both daily and master shifts are valid shift codes
        if emp.shift and shift:
            daily_shift = normalize_shift(shift)
            master_shift = normalize_shift(emp.shift)
            if daily_shift in VALID_SHIFT_CODES and master_shift in VALID_SHIFT_CODES and daily_shift != master_shift and display_status in ["Present", "Single Punch"]:
                issue = "Alternate Shift"


        early_minutes = 0
        if final_out and summary["is_early_departure"]:
            shift_end = get_shift_end_time(shift)
            out_mins = time_to_minutes(final_out)
            expected_mins = shift_end.hour * 60 + shift_end.minute
            if out_mins < expected_mins:
                early_minutes = expected_mins - out_mins

        # FIXED:
        existing_att = att_map.get(emp.pr_number)
        if existing_att:
            essl_has_real_punches = bool(essl_in) or bool(essl_out)
            tata_has_real_punches = bool(tata_in) or bool(tata_out)
            if not essl_has_real_punches and tata_has_real_punches:
                existing_att.essl_in = tata_in      # ← Now correct
                existing_att.essl_out = tata_out
            else:
                existing_att.essl_in = essl_in
                existing_att.essl_out = essl_out
            existing_att.tata_in = tata_in
            existing_att.tata_out = tata_out
            existing_att.final_in = final_in
            existing_att.final_out = final_out
            existing_att.worked_hours = worked_hours
            existing_att.man_hrs = tata_man_hrs
            existing_att.ot_hours = ot_hours
            existing_att.ot_headcount = ot_headcount
            existing_att.attendance_status = display_status
            existing_att.late_minutes = late_minutes
            existing_att.early_minutes = early_minutes
            existing_att.single_punch = single_punch
            existing_att.is_match = "Yes" if match_status == "Matched" else "No"
            existing_att.match_status = match_status
            existing_att.shift = shift
            existing_att.category = category
            existing_att.issue = issue
            existing_att.remark = remark
        else:
                        # UNIVERSAL MIRRORING: Fill ESSL with Tata when ESSL has no punch times
            essl_has_real_punches = bool(essl_in) or bool(essl_out)
            tata_has_real_punches = bool(tata_in) or bool(tata_out)
            if not essl_has_real_punches and tata_has_real_punches:
                essl_in_val = tata_in
                essl_out_val = tata_out
            else:
                essl_in_val = essl_in
                essl_out_val = essl_out
            db.add(Attendance(
                employee_id=emp.id, pr_number=emp.pr_number, emp_code=emp.emp_code, date=d,
                essl_in=essl_in_val, essl_out=essl_out_val, tata_in=tata_in, tata_out=tata_out,
                final_in=final_in, final_out=final_out, worked_hours=worked_hours, man_hrs=tata_man_hrs,
                ot_hours=ot_hours, ot_headcount=ot_headcount, attendance_status=display_status,
                late_minutes=late_minutes, early_minutes=early_minutes, single_punch=single_punch,
                is_match="Yes" if match_status == "Matched" else "No", match_status=match_status,
                vendor=emp.vendor, store=emp.store, department=emp.department,
                shift=shift, category=category, remark=remark, issue=issue, source="reconciliation"
            ))
            attendance_created += 1

        action_checks = [
            ("Single Punch", display_status == "Single Punch", "Medium", "HR Manager",
             f"{emp.name} ({emp.pr_number}) has single punch on {d.isoformat()}"),
            ("Late Punch", late_minutes > 0, "Low", "Supervisor",
             f"{emp.name} ({emp.pr_number}) punched {late_minutes} minutes late on {d.isoformat()}"),
            ("Early Departure", early_minutes > 0, "Low", "Supervisor",
             f"{emp.name} ({emp.pr_number}) left {early_minutes} minutes early on {d.isoformat()}"),
            ("Less Working Hours", display_status == "Less Working Hours", "Medium", "HR Manager",
             f"{emp.name} ({emp.pr_number}) worked {worked_hours}hrs instead of {get_standard_hours(shift)}hrs on {d.isoformat()}"),
            ("Reconciliation", match_status != "Matched",
             "High" if reconciliation_severity in ["Critical", "High"] else "Medium", "HR Manager",
             f"{emp.name} ({emp.pr_number}) has {match_status} on {d.isoformat()}. ESSL: {essl_in}-{essl_out}, Tata: {tata_in}-{tata_out}")
        ]

        for action_type, condition, priority, assigned_to, description in action_checks:
            if condition:
                action_key = (emp.pr_number, d.isoformat(), action_type)
                if action_key not in action_map:
                    db.add(HRAction(
                        pr_number=emp.pr_number, emp_name=emp.name, date=d,
                        action_type=action_type, description=description,
                        priority=priority, assigned_to=assigned_to
                    ))
                    hr_actions_created += 1

        if ot_headcount > 0 and summary["is_ot_eligible"]:
            ot_key = (emp.pr_number, d.isoformat())
            if ot_key not in ot_map:
                db.add(Overtime(
                    employee_id=emp.id, pr_number=emp.pr_number, name=emp.name,
                    store=emp.store, date=d, worked_hours=worked_hours,
                    ot_hours=ot_hours, calc_headcount=ot_headcount, status="Pending"
                ))
                hr_actions_created += 1
            # Check OT thresholds whenever an OT record exists for this day
            check_ot_thresholds(emp.pr_number, emp.name, d, db)

    db.commit()
    return {
        "status": "completed", "date": d.isoformat(),
        "attendance_records_created": attendance_created,
        "reconciliation_issues": reconciliation_issues,
        "hr_actions_created": hr_actions_created,
        "total_processed": len(employees)
    }

@app.post("/api/v1/reconciliation/run")
def run_reconciliation(target_date: str = Form(...), db: Session = Depends(get_db)):
    d = parse_date_br(target_date)
    if not d:
        raise HTTPException(status_code=400, detail="Invalid date format")
    return _reconcile_single_date(d, db)

@app.post("/api/v1/reconciliation/run-date")
def run_reconciliation_date(target_date: str = Form(...), clean: bool = Form(False), db: Session = Depends(get_db)):
    """Reconcile ONE day synchronously (no background thread). Returns in <10s.
    Use clean=true to wipe that day's derived data first."""
    d = parse_date_br(target_date)
    if not d:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")
    if clean:
        _clear_derived_data(d, d + timedelta(days=1), db)
    return _reconcile_single_date(d, db)

# ============================================================================
# BACKGROUND JOB TRACKING for month-long reconciliation.
#
# Railway's public-network edge proxy enforces a hard ~5 minute cap on any single
# HTTP request that neither side can raise - a full month's reconciliation can
# legitimately take longer than that. So run-month no longer does the work inline:
# it kicks off a background thread and returns a job_id immediately; the frontend
# polls /reconciliation/run-month/{job_id} for status/progress until it's done.
#
# In-memory dict is fine here since this runs as a single uvicorn process/worker
# (see the bottom of this file - no `workers=` argument). If that ever changes to
# multiple workers, this needs to move to a DB-backed jobs table instead, since
# each worker would have its own copy of this dict.
# ============================================================================
RECONCILIATION_JOBS: Dict[str, Dict[str, Any]] = {}

JOB_STALL_SECONDS = 150  # if a stage hasn't moved in this long, treat the job as dead

def _set_job_stage(job_id: Optional[str], stage: str):
    if job_id and job_id in RECONCILIATION_JOBS:
        RECONCILIATION_JOBS[job_id]["stage"] = stage
        RECONCILIATION_JOBS[job_id]["stage_updated_at"] = datetime.now()

# FIX v3.2: Completely rewritten with bulk pre-fetch to eliminate timeout
def _run_reconciliation_month_core(month: str, db: Session, job_id: Optional[str] = None) -> Dict[str, Any]:
    year, mon = parse_year_month(month)
    start = date(year, mon, 1)
    end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)
    total_days_in_month = (end - start).days

    _set_job_stage(job_id, "prefetching employee data")

    employees = db.query(Employee).all()
    emp_ids = [e.id for e in employees]
    emp_by_id = {e.id: e for e in employees}
    emp_by_pr = {e.pr_number: e for e in employees if e.pr_number}
    pr_numbers = [e.pr_number for e in employees]

    # MEMORY FIX: the previous version pre-fetched every ESSL/Tata/Attendance/
    # Reconciliation/HRAction/Overtime row for the ENTIRE month (~20-30k rows per
    # table) and held all six of those row sets, plus the insert/update lists
    # being built from them, in memory simultaneously for the whole run. With a
    # large workforce that peak footprint (a few hundred thousand mapped ORM
    # objects at once, on top of everything else the single process was already
    # serving) was large enough to risk exceeding the container's memory limit -
    # which reads externally as the process/container being killed and restarted
    # mid-request, i.e. an instant connection-reset ("502") on whatever request
    # happened to be in flight at that moment, exactly as observed in production.
    #
    # The whole-month "Tata-only" employee detection below only needs to know
    # WHICH pr_numbers have any ESSL/Tata row this month, not the row contents,
    # so it is computed with two cheap distinct-pr_number-only queries up front.
    # The actual per-day row data is then fetched CHUNK_DAYS days at a time in
    # the loop below, so peak memory is bounded by one chunk's worth of rows
    # instead of the whole month's, regardless of workforce size.
    CHUNK_DAYS = 3

    essl_prs_month = {
        pr for (pr,) in db.query(ESSLAttendance.pr_number).filter(
            ESSLAttendance.pr_number.in_(pr_numbers),
            ESSLAttendance.date >= start, ESSLAttendance.date < end
        ).distinct().all() if pr
    }
    tata_prs_month = {
        pr for (pr,) in db.query(TataAttendance.pr_number).filter(
            TataAttendance.pr_number.in_(pr_numbers),
            TataAttendance.date >= start, TataAttendance.date < end
        ).distinct().all() if pr
    }
    tata_only_prs = tata_prs_month - essl_prs_month

    days_processed = 0
    totals = {"attendance_records_created": 0, "reconciliation_issues": 0, "hr_actions_created": 0}
    month_ot_by_pr: Dict[str, List[tuple]] = {}  # accumulated across chunks for the OT-threshold post-check

    chunk_start = start
    while chunk_start < end:
        chunk_end = min(chunk_start + timedelta(days=CHUNK_DAYS), end)
        _set_job_stage(job_id, f"prefetching {chunk_start.isoformat()} to {(chunk_end - timedelta(days=1)).isoformat()}")

        all_essl = db.query(ESSLAttendance).filter(
            ESSLAttendance.pr_number.in_(pr_numbers),
            ESSLAttendance.date >= chunk_start,
            ESSLAttendance.date < chunk_end
        ).all()
        essl_map = {}
        for e in all_essl:
            key = (e.pr_number, e.date.isoformat())
            essl_map[key] = e

        all_tata = db.query(TataAttendance).filter(
            TataAttendance.pr_number.in_(pr_numbers),
            TataAttendance.date >= chunk_start,
            TataAttendance.date < chunk_end
        ).all()
        tata_map = {}
        for t in all_tata:
            key = (t.pr_number, t.date.isoformat())
            tata_map[key] = t

        all_att = db.query(Attendance).filter(
            Attendance.pr_number.in_(pr_numbers),
            Attendance.date >= chunk_start,
            Attendance.date < chunk_end
        ).all()
        att_map = {}
        for a in all_att:
            key = (a.pr_number, a.date.isoformat())
            att_map[key] = a

        all_recon = db.query(AttendanceReconciliation).filter(
            AttendanceReconciliation.pr_number.in_(pr_numbers),
            AttendanceReconciliation.date >= chunk_start,
            AttendanceReconciliation.date < chunk_end
        ).all()
        recon_map = {}
        for r in all_recon:
            key = (r.pr_number, r.date.isoformat())
            recon_map[key] = r

        all_actions = db.query(HRAction).filter(
            HRAction.pr_number.in_(pr_numbers),
            HRAction.date >= chunk_start,
            HRAction.date < chunk_end
        ).all()
        action_map = {}
        for a in all_actions:
            key = (a.pr_number, a.date.isoformat(), a.action_type)
            action_map[key] = a

        all_ot = db.query(Overtime).filter(
            Overtime.pr_number.in_(pr_numbers),
            Overtime.date >= chunk_start,
            Overtime.date < chunk_end
        ).all()
        ot_map = {}
        for o in all_ot:
            key = (o.pr_number, o.date.isoformat())
            ot_map[key] = o
            if o.ot_hours and o.ot_hours > 0:
                month_ot_by_pr.setdefault(o.pr_number, []).append((o.date, o.ot_hours))

        # Process this chunk entirely in memory
        attendance_to_insert = []
        attendance_updates = []  # dicts for bulk_update_mappings - see note below
        recon_to_insert = []
        actions_to_insert = []
        ot_to_insert = []


        d = chunk_start
        while d < chunk_end:
            d_iso = d.isoformat()
            for emp in employees:
                emp_id = emp.id
                key = (emp.pr_number, d_iso)

                essl = essl_map.get(key)
                tata = tata_map.get(key)

                essl_in = essl.in_time if essl else None
                essl_out = essl.out_time if essl else None
                tata_in = tata.in_time if tata else None
                tata_out = tata.out_time if tata else None
                tata_man_hrs = tata.man_hrs if tata else 0
                tata_status = tata.status if tata else None
                essl_status = _get_essl_status(essl)

                # Sunday handling: Sundays default to Week Off unless Tata shows the employee
                # actually worked (has a Tata record) - see Present override further below.
                is_sunday = d.weekday() == 6  # Monday=0 ... Sunday=6

                # FIX v3.2.8: If employee has NO ESSL and NO Tata data, create "Absent" record
                if not essl and not tata:
                    shift = safe_shift(emp.shift) if emp.shift else "G"
                    category = emp.wc or "BC"
                    if is_sunday:
                        display_status = "Week Off"
                        issue = "Week Off"
                        match_status = "Week Off"
                        remark = "Sunday - Week Off (no attendance data)"
                    else:
                        display_status = "Absent"
                        issue = "No Data"
                        match_status = "No Data"
                        remark = "No ESSL or Tata record found for this date"

                    att_key = (emp.pr_number, d_iso)
                    if att_key in att_map:
                        existing = att_map[att_key]
                        # NOTE: appending a plain dict for bulk_update_mappings() instead of
                        # mutating the ORM object directly. With ~30k rows/month, mutating
                        # tracked ORM objects makes SQLAlchemy issue one UPDATE per row at
                        # commit time (this was the actual cause of the 300s timeout) -
                        # bulk_update_mappings does it as a single fast batched statement.
                        attendance_updates.append({
                            "id": existing.id,
                            "essl_in": None, "essl_out": None, "tata_in": None, "tata_out": None,
                            "final_in": None, "final_out": None, "worked_hours": 0, "man_hrs": 0,
                            "ot_hours": 0, "ot_headcount": 0, "attendance_status": display_status,
                            "late_minutes": 0, "early_minutes": 0, "single_punch": "No",
                            "is_match": "No", "match_status": match_status,
                            "shift": shift, "category": category, "issue": issue, "remark": remark,
                        })
                    else:
                        attendance_to_insert.append({
                            "employee_id": emp.id, "pr_number": emp.pr_number, "emp_code": emp.emp_code, "date": d,
                            "essl_in": None, "essl_out": None, "tata_in": None, "tata_out": None,
                            "final_in": None, "final_out": None, "worked_hours": 0, "man_hrs": 0,
                            "ot_hours": 0, "ot_headcount": 0, "attendance_status": display_status,
                            "late_minutes": 0, "early_minutes": 0, "single_punch": "No",
                            "is_match": "No", "match_status": match_status,
                            "vendor": emp.vendor, "store": emp.store, "department": emp.department,
                            "shift": shift, "category": category, "remark": remark, "issue": issue, "source": "reconciliation"
                        })
                        totals["attendance_records_created"] += 1
                    continue

                # Continue with existing logic for employees who HAVE ESSL or Tata data

                # FIX v3.2.7: For WO employees, shift stays empty. For absent, use master shift.
                has_essl_punches = bool(essl_in or essl_out)
                has_tata_punches = bool(tata_in or tata_out)
                has_any_punches = has_essl_punches or has_tata_punches

                if essl_status == "WO" and not has_any_punches:
                    # Week Off: no shift shown, status will be "Week Off"
                    shift = None
                elif not has_any_punches:
                    # Regular absent: use assigned shift from master
                    shift = safe_shift(emp.shift) if emp.shift else "G"
                else:
                    shift = safe_shift((tata.shift if tata and tata.shift else None) or (essl.shift if essl and essl.shift else None) or emp.shift) or "G"
                category = emp.wc or "BC"

                has_essl = bool(essl_in or essl_out or essl_status)
                has_tata = bool(tata_in or tata_out)
                # FIX v3.2.5: Check actual punches separately from status
                has_essl_punches = bool(essl_in or essl_out)
                has_tata_punches = bool(tata_in or tata_out)
                has_any_punches = has_essl_punches or has_tata_punches

                match_delta_in = 0
                match_delta_out = 0

                # FIX v3.2.5: WO = Absent only if NO actual punches from either source
                if essl_status == "WO" and not has_any_punches:
                    match_status = "Absent"
                    reconciliation_severity = "Low"
                    remark = "ESSL shows WO (treated as Absent - no punches)"
                elif essl_status == "WO" and has_any_punches:
                    # ESSL says WO but punches exist → Present, flag mismatch
                    match_status = "Mismatch"
                    reconciliation_severity = "Medium"
                    remark = "ESSL shows WO but punches exist"
                elif not has_essl and not has_tata:
                    match_status = "No Data"
                    reconciliation_severity = "Critical"
                    remark = ""
                elif emp.pr_number in tata_only_prs:
                    # TATA-ONLY: Employee has no ESSL data but has Tata data → use Tata only, not an error
                    match_status = "Tata Only"
                    reconciliation_severity = "Low"
                    remark = "Tata-only employee (no ESSL data)"
                elif not has_essl and has_tata:
                    match_status = "Missing ESSL"
                    reconciliation_severity = "High"
                    remark = "No Punch ESSL"
                elif has_essl and not has_tata:
                    match_status = "Missing Tata"
                    reconciliation_severity = "High"
                    remark = ""
                else:
                    if essl_in and tata_in:
                        match_delta_in = abs(time_to_minutes(essl_in) - time_to_minutes(tata_in))
                    if essl_out and tata_out:
                        match_delta_out = abs(time_to_minutes(essl_out) - time_to_minutes(tata_out))
                    if match_delta_in > 15 or match_delta_out > 15:
                        match_status = "Mismatch"
                        reconciliation_severity = "Medium"
                    else:
                        match_status = "Matched"
                        reconciliation_severity = "Low"
                    remark = ""

                if match_status != "Matched":
                    recon_key = (emp.pr_number, d_iso)
                    if recon_key not in recon_map:
                        recon_to_insert.append({
                            "pr_number": emp.pr_number, "emp_name": emp.name, "date": d,
                            "essl_in": essl_in, "essl_out": essl_out, "tata_in": tata_in, "tata_out": tata_out,
                            "in_delta_minutes": match_delta_in, "out_delta_minutes": match_delta_out,
                            "match_status": match_status, "severity": reconciliation_severity,
                            "remarks": f"ESSL: {essl_in}-{essl_out}, Tata: {tata_in}-{tata_out}"
                        })
                        totals["reconciliation_issues"] += 1

                final_in = tata_in or essl_in
                final_out = tata_out or essl_out
                punch_count = sum([1 for p in [final_in, final_out] if p])
                single_punch = "Yes" if punch_count == 1 else "No"

                # TATA PRIORITY: Determine which source to use for OT/worked_hours calculation
                tata_has_punches = bool(tata_in or tata_out)
                tata_has_man_hrs = bool(tata_man_hrs and tata_man_hrs > 0)
                essl_has_punches = bool(essl_in or essl_out)

                # Priority: Tata punches > Tata man_hrs > ESSL punches > 0
                if tata_has_punches:
                    # Tata has punches → calculate OT from Tata punches
                    ot_in_time, ot_out_time = tata_in, tata_out
                    force_man_hrs = False
                elif tata_has_man_hrs:
                    # Tata has no punches but has man_hrs → use man_hrs for worked_hours
                    ot_in_time, ot_out_time = None, None
                    force_man_hrs = True
                elif essl_has_punches:
                    # No Tata data → use ESSL punches
                    ot_in_time, ot_out_time = essl_in, essl_out
                    force_man_hrs = False
                else:
                    ot_in_time, ot_out_time = None, None
                    force_man_hrs = False

                summary = calculate_daily_summary(shift, category, tata_status, ot_in_time, ot_out_time, tata_man_hrs, force_man_hrs=force_man_hrs)
                worked_hours = summary["worked_hours"]
                attendance_status = summary["status"]
                late_minutes = summary["late_minutes"]
                ot_hours = summary["calculated_ot_hours"]
                ot_headcount = summary["ot_headcount"]

                status_map = {"P": "Present", "A": "Absent", "SP": "Single Punch", "L": "Leave", "HD": "Half Day", "WO": "Week Off"}
                display_status = status_map.get(attendance_status, attendance_status)

                # FIX v3.2.7: Override to "Week Off" if ESSL says WO and no punches
                if essl_status == "WO" and not has_any_punches:
                    display_status = "Week Off"

                # Sunday + Tata data present -> employee worked, must show Present regardless of
                # what the generic shift/hours calculation produced.
                if is_sunday and tata:
                    display_status = "Present"
                    attendance_status = "P"

                # FIX v3.2.4: Status stays "Present" for Late Punch, Early Departure, etc.
                # The Issue column captures the specific problem. Only Present/Absent as main status.
                if summary["is_late_punch"] and display_status == "Present":
                    pass  # Status stays "Present", issue = "Late Punch"
                elif summary["is_early_departure"] and display_status == "Present":
                    pass  # Status stays "Present", issue = "Early Departure"
                elif summary["is_less_working_hours"] and display_status == "Present":
                    pass  # Status stays "Present", issue = "Less Working Hours"
                # FIX v3.2.1: Removed "Missing ESSL" override. Tata punches = Present.
                # match_status already reports "Missing ESSL" for data quality.

                # Compute Issue field for HR actionability
                issue = "-"
                if display_status == "Week Off":
                    issue = "Week Off"
                elif display_status == "Absent":
                    issue = "Absent"
                elif essl_status == "WO" and has_tata:
                    issue = "ESSL WO vs Tata Punches"
                elif emp.pr_number in tata_only_prs:
                    issue = "-"
                elif not has_essl and has_tata:
                    issue = "Missing ESSL Punch"
                elif has_essl and not has_tata:
                    issue = "Missing Tata Punch"
                elif match_status == "Mismatch":
                    if abs(match_delta_in) > 30 or abs(match_delta_out) > 30:
                        issue = "Time Difference (>30 min)"
                    else:
                        issue = "Time Difference"
                # FIX v3.2.4: Check summary flags instead of display_status for sub-issues
                elif summary["is_single_punch"]:
                    issue = "Single Punch"
                elif summary["is_late_punch"]:
                    issue = "Late Punch"
                elif summary["is_early_departure"]:
                    issue = "Early Departure"
                elif summary["is_less_working_hours"]:
                    issue = "Less Working Hours"

                # Check alternate shift
                # FIX v3.2.2: Only flag if both daily and master shifts are valid shift codes
                if emp.shift and shift:
                    daily_shift = normalize_shift(shift)
                    master_shift = normalize_shift(emp.shift)
                    if daily_shift in VALID_SHIFT_CODES and master_shift in VALID_SHIFT_CODES and daily_shift != master_shift and display_status in ["Present", "Single Punch"]:
                        issue = "Alternate Shift"

                early_minutes = 0
                if final_out and summary["is_early_departure"]:
                    shift_end = get_shift_end_time(shift)
                    out_mins = time_to_minutes(final_out)
                    expected_mins = shift_end.hour * 60 + shift_end.minute
                    if out_mins < expected_mins:
                        early_minutes = expected_mins - out_mins

                att_key = (emp.pr_number, d_iso)
                if att_key in att_map:
                    existing = att_map[att_key]
                    # UNIVERSAL MIRRORING: Fill ESSL with Tata when ESSL has no punch times
                    essl_has_real_punches = bool(essl_in) or bool(essl_out)
                    tata_has_real_punches = bool(tata_in) or bool(tata_out)
                    if not essl_has_real_punches and tata_has_real_punches:
                        essl_in_upd, essl_out_upd = tata_in, tata_out
                    else:
                        essl_in_upd, essl_out_upd = essl_in, essl_out
                    attendance_updates.append({
                        "id": existing.id,
                        "essl_in": essl_in_upd, "essl_out": essl_out_upd,
                        "tata_in": tata_in, "tata_out": tata_out,
                        "final_in": final_in, "final_out": final_out,
                        "worked_hours": worked_hours, "man_hrs": tata_man_hrs,
                        "ot_hours": ot_hours, "ot_headcount": ot_headcount,
                        "attendance_status": display_status,
                        "late_minutes": late_minutes, "early_minutes": early_minutes,
                        "single_punch": single_punch,
                        "is_match": "Yes" if match_status == "Matched" else "No",
                        "match_status": match_status, "issue": issue,
                        "shift": shift, "category": category, "remark": remark,
                    })
                else:
                # UNIVERSAL MIRRORING: Fill ESSL with Tata when ESSL has no punch times
                    essl_has_real_punches = bool(essl_in) or bool(essl_out)
                    tata_has_real_punches = bool(tata_in) or bool(tata_out)
                    if not essl_has_real_punches and tata_has_real_punches:
                        essl_in_val = tata_in
                        essl_out_val = tata_out
                    else:
                        essl_in_val = essl_in
                        essl_out_val = essl_out
                    attendance_to_insert.append({
                        "employee_id": emp.id, "pr_number": emp.pr_number, "emp_code": emp.emp_code, "date": d,
                        "essl_in": essl_in_val, "essl_out": essl_out_val, "tata_in": tata_in, "tata_out": tata_out,
                        "final_in": final_in, "final_out": final_out, "worked_hours": worked_hours, "man_hrs": tata_man_hrs,
                        "ot_hours": ot_hours, "ot_headcount": ot_headcount, "attendance_status": display_status,
                        "late_minutes": late_minutes, "early_minutes": early_minutes, "single_punch": single_punch,
                        "is_match": "Yes" if match_status == "Matched" else "No", "match_status": match_status,
                        "vendor": emp.vendor, "store": emp.store, "department": emp.department,
                        "shift": shift, "category": category, "remark": remark, "issue": issue, "source": "reconciliation"
                    })
                    totals["attendance_records_created"] += 1

                action_checks = [
                    ("Single Punch", display_status == "Single Punch", "Medium", "HR Manager",
                     f"{emp.name} ({emp.pr_number}) has single punch on {d_iso}"),
                    ("Late Punch", late_minutes > 0, "Low", "Supervisor",
                     f"{emp.name} ({emp.pr_number}) punched {late_minutes} minutes late on {d_iso}"),
                    ("Early Departure", early_minutes > 0, "Low", "Supervisor",
                     f"{emp.name} ({emp.pr_number}) left {early_minutes} minutes early on {d_iso}"),
                    ("Less Working Hours", display_status == "Less Working Hours", "Medium", "HR Manager",
                     f"{emp.name} ({emp.pr_number}) worked {worked_hours}hrs instead of {get_standard_hours(shift)}hrs on {d_iso}"),
                    ("Reconciliation", match_status != "Matched",
                     "High" if reconciliation_severity in ["Critical", "High"] else "Medium", "HR Manager",
                     f"{emp.name} ({emp.pr_number}) has {match_status} on {d_iso}. ESSL: {essl_in}-{essl_out}, Tata: {tata_in}-{tata_out}")
                ]

                for action_type, condition, priority, assigned_to, description in action_checks:
                    if condition:
                        action_key = (emp.pr_number, d_iso, action_type)
                        if action_key not in action_map:
                            actions_to_insert.append({
                                "pr_number": emp.pr_number, "emp_name": emp.name, "date": d,
                                "action_type": action_type, "description": description,
                                "priority": priority, "assigned_to": assigned_to
                            })
                            totals["hr_actions_created"] += 1

                if ot_headcount > 0 and summary["is_ot_eligible"]:
                    ot_key = (emp.pr_number, d_iso)
                    if ot_key not in ot_map:
                        ot_to_insert.append({
                            "employee_id": emp.id, "pr_number": emp.pr_number, "name": emp.name,
                            "store": emp.store, "date": d, "worked_hours": worked_hours,
                            "ot_hours": ot_hours, "calc_headcount": ot_headcount, "status": "Pending"
                        })
                        totals["hr_actions_created"] += 1

            days_processed += 1
            _set_job_stage(job_id, f"processed {days_processed}/{total_days_in_month} day(s) in memory")
            d += timedelta(days=1)

        _set_job_stage(job_id, "writing results to database")

        # FIX v3.2: Bulk insert everything at once
        # FIX v3.2.11: this was the actual cause of the 300s timeout. On a re-run (most
        # days already have an Attendance row from a prior reconciliation), the old code
        # mutated ~30,000 already-persistent ORM objects directly, which makes SQLAlchemy
        # issue one individual UPDATE per row at commit() - very slow, especially over the
        # network to a hosted Postgres instance. bulk_update_mappings() below does it as a
        # handful of fast batched statements instead.
        # FIX: bulk_insert_mappings works from plain dicts and skips ORM instrumentation/
        # identity-map bookkeeping entirely (unlike bulk_save_objects, which still builds
        # a full mapped object per row first). For 20-30k rows/month this is a meaningful
        # chunk of the "writing results to database" stage.
        if attendance_to_insert:
            db.bulk_insert_mappings(Attendance, attendance_to_insert)
        if attendance_updates:
            db.bulk_update_mappings(Attendance, attendance_updates)
        if recon_to_insert:
            db.bulk_insert_mappings(AttendanceReconciliation, recon_to_insert)
        if actions_to_insert:
            db.bulk_insert_mappings(HRAction, actions_to_insert)
        if ot_to_insert:
            db.bulk_insert_mappings(Overtime, ot_to_insert)
            # Accumulate newly-created OT rows into the whole-month tracker used by the
            # post-processing OT-threshold check below, before this chunk's local
            # ot_to_insert list goes out of scope. These are plain dicts (for
            # bulk_insert_mappings), so dict-key access is required here.
            for o in ot_to_insert:
                if o["ot_hours"] and o["ot_hours"] > 0:
                    month_ot_by_pr.setdefault(o["pr_number"], []).append((o["date"], o["ot_hours"]))

        db.commit()
        chunk_start = chunk_end


    _set_job_stage(job_id, "checking OT thresholds")

    # Post-process: Check OT thresholds for employees who accrued OT this month.
    # FIX: this used to call check_ot_thresholds() - 2 SUM queries + up to 2 existence
    # queries - once per employee PER DAY that had OT. For a full month with many
    # employees on OT, that was easily thousands of extra round-trips. Now computed
    # once from data already in memory (month_ot_by_pr, accumulated incrementally as
    # each chunk was processed above), plus 2 bulk existence queries total.
    #
    # NOTE: month_ot_by_pr is intentionally NOT rebuilt here from all_ot/ot_to_insert -
    # those are chunk-local variables that no longer exist once the chunk loop above
    # has finished, and the previous version of this block additionally had a latent
    # bug where it accessed ot_to_insert entries (plain dicts, for bulk_insert_mappings)
    # via attribute syntax (o.ot_hours) instead of dict-key syntax (o["ot_hours"]),
    # which would raise AttributeError and fail the job whenever new OT records were
    # created during the run.

    if month_ot_by_pr:
        prs_with_ot = list(month_ot_by_pr.keys())
        existing_weekly = db.query(HRAction.pr_number, HRAction.date).filter(
            HRAction.pr_number.in_(prs_with_ot),
            HRAction.action_type == "OT Weekly Threshold",
            HRAction.date >= start - timedelta(days=6), HRAction.date < end
        ).all()
        weekly_flagged_dates: Dict[str, List[date]] = {}
        for pr, dt in existing_weekly:
            weekly_flagged_dates.setdefault(pr, []).append(dt)

        existing_monthly_prs = {
            pr for (pr,) in db.query(HRAction.pr_number).filter(
                HRAction.pr_number.in_(prs_with_ot),
                HRAction.action_type == "OT Monthly Threshold",
                HRAction.date >= start, HRAction.date < end
            ).distinct().all()
        }

        threshold_actions_to_insert = []
        for pr, entries in month_ot_by_pr.items():
            emp = emp_by_pr.get(pr)
            if not emp:
                continue
            entries.sort(key=lambda x: x[0])
            already_flagged = weekly_flagged_dates.get(pr, [])
            weekly_alert_added = False

            for d_i, _ in entries:
                week_start_i = d_i - timedelta(days=6)
                weekly_sum = round(sum(h for dd, h in entries if week_start_i <= dd <= d_i), 2)
                if weekly_sum > OT_WEEKLY_THRESHOLD and not weekly_alert_added:
                    if not any(week_start_i <= fd <= d_i for fd in already_flagged):
                        threshold_actions_to_insert.append({
                            "pr_number": pr, "emp_name": emp.name, "date": d_i,
                            "action_type": "OT Weekly Threshold",
                            "description": (f"{emp.name} ({pr}) has exceeded the weekly OT limit: "
                                         f"{weekly_sum}h / {OT_WEEKLY_THRESHOLD}h (week ending {d_i.isoformat()})"),
                            "priority": "High", "assigned_to": "HR Manager"
                        })
                        weekly_alert_added = True
                        totals["hr_actions_created"] += 1

            monthly_sum = round(sum(h for _, h in entries), 2)
            if monthly_sum > OT_MONTHLY_THRESHOLD and pr not in existing_monthly_prs:
                last_date = entries[-1][0]
                threshold_actions_to_insert.append({
                    "pr_number": pr, "emp_name": emp.name, "date": last_date,
                    "action_type": "OT Monthly Threshold",
                    "description": (f"{emp.name} ({pr}) has exceeded the monthly OT limit: "
                                 f"{monthly_sum}h / {OT_MONTHLY_THRESHOLD}h"),
                    "priority": "High", "assigned_to": "HR Manager"
                })
                totals["hr_actions_created"] += 1

        if threshold_actions_to_insert:
            db.bulk_insert_mappings(HRAction, threshold_actions_to_insert)
            db.commit()

    return {
        "status": "completed",
        "month": month,
        "days_processed": days_processed,
        **totals
    }

def _run_reconciliation_job(job_id: str, month: str, clean: bool = False):
    """Runs on a background thread with its own DB session."""
    job_db = SessionLocal()
    try:
        if clean:
            year, mon = parse_year_month(month)
            start = date(year, mon, 1)
            end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)
            _clear_derived_data(start, end, job_db)

        result = _run_reconciliation_month_core(month, job_db, job_id=job_id)
        RECONCILIATION_JOBS[job_id]["status"] = "completed"
        RECONCILIATION_JOBS[job_id]["result"] = result
        RECONCILIATION_JOBS[job_id]["stage"] = "done"
        RECONCILIATION_JOBS[job_id]["finished_at"] = datetime.now()
    except Exception as e:
        job_db.rollback()
        import traceback
        tb = traceback.format_exc()
        RECONCILIATION_JOBS[job_id]["status"] = "failed"
        RECONCILIATION_JOBS[job_id]["error"] = f"{str(e)}\n\n{tb}"
        RECONCILIATION_JOBS[job_id]["finished_at"] = datetime.now()
    finally:
        job_db.close()

@app.post("/api/v1/reconciliation/run-month")
def run_reconciliation_month(month: str = Form(...), clean: bool = Form(False)):
    """Kicks off month reconciliation as a background job.

    Set clean=true to wipe all existing Attendance/Reconciliation/HRAction/OT
    for this month before re-running. Use this if your data looks corrupted.
    """
    # Prune old finished jobs
    cutoff = datetime.now() - timedelta(hours=2)
    for jid in [j for j, v in RECONCILIATION_JOBS.items()
                if v["status"] != "running" and v.get("finished_at", datetime.now()) < cutoff]:
        del RECONCILIATION_JOBS[jid]

    existing_running = next((jid for jid, v in RECONCILIATION_JOBS.items() if v["status"] == "running"), None)
    if existing_running:
        running_job = RECONCILIATION_JOBS[existing_running]
        return {
            "status": "already_running",
            "job_id": existing_running,
            "month": running_job["month"],
            "stage": running_job.get("stage"),
            "detail": f"A reconciliation job for {running_job['month']} is already in progress."
        }

    job_id = str(uuid.uuid4())
    RECONCILIATION_JOBS[job_id] = {
        "status": "running", "stage": "queued", "month": month,
        "result": None, "error": None, "started_at": datetime.now(), "finished_at": None,
        "stage_updated_at": datetime.now()
    }
    thread = threading.Thread(target=_run_reconciliation_job, args=(job_id, month, clean), daemon=True)
    thread.start()
    return {"status": "started", "job_id": job_id, "month": month, "clean": clean}

@app.get("/api/v1/reconciliation/run-month/{job_id}")
def get_reconciliation_month_job(job_id: str):
    job = RECONCILIATION_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found (it may have completed a while ago and been cleared, or the server restarted)")

    # FIX: if the background thread died/hung without hitting our except block
    # (e.g. a hard socket hang before the statement_timeout could even kick in,
    # or the process was killed), the job would otherwise sit at "running" with
    # a null finished_at forever, which is exactly what was showing on screen.
    # Treat "no stage progress for JOB_STALL_SECONDS" as a failure.
    if job["status"] == "running":
        last_progress = job.get("stage_updated_at") or job["started_at"]
        stalled_for = (datetime.now() - last_progress).total_seconds()
        if stalled_for > JOB_STALL_SECONDS:
            job["status"] = "failed"
            job["error"] = (f"Job stalled for over {JOB_STALL_SECONDS}s at stage "
                             f"'{job.get('stage')}' with no progress - likely a dropped DB "
                             f"connection. Please retry.")
            job["finished_at"] = datetime.now()

    return job

@app.get("/api/v1/reconciliation/summary")
def get_reconciliation_summary(target_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    d = parse_date_br(target_date) if target_date else date.today()
    if not d:
        d = date.today()
    # PERF FIX: was 8 separate COUNT queries (one per status/severity value).
    # A GROUP BY does the same job in 2 queries regardless of how many
    # match_status/severity values exist.
    total = db.query(AttendanceReconciliation).filter(AttendanceReconciliation.date == d).count()
    status_counts = dict(
        db.query(AttendanceReconciliation.match_status, func.count(AttendanceReconciliation.id))
        .filter(AttendanceReconciliation.date == d)
        .group_by(AttendanceReconciliation.match_status)
        .all()
    )
    critical = (
        db.query(func.count(AttendanceReconciliation.id))
        .filter(AttendanceReconciliation.date == d, AttendanceReconciliation.severity == "Critical")
        .scalar() or 0
    )
    return {
        "total": total,
        "matched": status_counts.get("Matched", 0),
        "mismatched": status_counts.get("Mismatch", 0),
        "missing_essl": status_counts.get("Missing ESSL", 0),
        "missing_tata": status_counts.get("Missing Tata", 0),
        "no_data": status_counts.get("No Data", 0),
        "week_off": status_counts.get("Week Off", 0),
        "critical": critical
    }

@app.get("/api/v1/reconciliation/records")
def get_reconciliation_records(target_date: Optional[str] = Query(None), page: int = Query(1, ge=1), page_size: int = Query(50, ge=1, le=200), db: Session = Depends(get_db)):
    d = parse_date_br(target_date) if target_date else date.today()
    if not d:
        d = date.today()
    total = db.query(AttendanceReconciliation).filter(AttendanceReconciliation.date == d).count()
    items = db.query(AttendanceReconciliation).filter(AttendanceReconciliation.date == d).order_by(desc(AttendanceReconciliation.created_at)).offset((page - 1) * page_size).limit(page_size).all()
    data = [{"id": r.id, "pr_number": r.pr_number, "date": r.date.isoformat(),
             "essl_in": r.essl_in, "essl_out": r.essl_out, "tata_in": r.tata_in, "tata_out": r.tata_out,
             "in_delta_minutes": r.in_delta_minutes, "out_delta_minutes": r.out_delta_minutes,
             "match_status": r.match_status, "severity": r.severity} for r in items]
    return {"data": data, "total": total, "page": page, "total_pages": max(1, (total + page_size - 1) // page_size)}

# ============================================================================
# LATE PUNCH ENDPOINTS
# ============================================================================

@app.post("/api/v1/late-punch/calculate")
def calculate_late_punch(pr_number: str = Form(...), month: str = Form(...), db: Session = Depends(get_db)):
    year, mon = parse_year_month(month)
    start = date(year, mon, 1)
    end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)
    late_count = db.query(Attendance).filter(Attendance.pr_number == pr_number, Attendance.date >= start, Attendance.date < end, Attendance.late_minutes > 0).count()
    penalty = calculate_late_punch_penalty(late_count)
    existing = db.query(LatePunchPenalty).filter(LatePunchPenalty.pr_number == pr_number, LatePunchPenalty.month == month).first()
    if existing:
        existing.late_count = late_count; existing.penalty_days = penalty["penalty_days"]; existing.calculated_at = datetime.utcnow()
    else:
        db.add(LatePunchPenalty(pr_number=pr_number, month=month, late_count=late_count, penalty_days=penalty["penalty_days"]))
    db.commit()
    return {"pr_number": pr_number, "month": month, "late_count": late_count, **penalty}

@app.post("/api/v1/late-punch/calculate-all")
def calculate_all_late_punches(month: str = Query(...), db: Session = Depends(get_db)):
    year, mon = parse_year_month(month)
    start = date(year, mon, 1)
    end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)
    employees = db.query(Employee).filter(Employee.status == "ACTIVE").all()
    processed = 0
    for emp in employees:
        late_count = db.query(Attendance).filter(Attendance.pr_number == emp.pr_number, Attendance.date >= start, Attendance.date < end, Attendance.late_minutes > 0).count()
        penalty = calculate_late_punch_penalty(late_count)
        existing = db.query(LatePunchPenalty).filter(LatePunchPenalty.pr_number == emp.pr_number, LatePunchPenalty.month == month).first()
        if existing:
            existing.late_count = late_count; existing.penalty_days = penalty["penalty_days"]; existing.calculated_at = datetime.utcnow()
        else:
            db.add(LatePunchPenalty(pr_number=emp.pr_number, month=month, late_count=late_count, penalty_days=penalty["penalty_days"]))
        processed += 1
    db.commit()
    return {"status": "completed", "month": month, "employees_processed": processed}

# ============================================================================
# HR BEHAVIORAL ALERTS
# (new endpoints only - existing /late-punch/calculate and /late-punch/calculate-all
#  above are untouched)
# ============================================================================


@app.get("/api/v1/employee/summary/{pr_number}")
def get_employee_summary(pr_number: str, month: Optional[str] = Query(None), db: Session = Depends(get_db)):
    """
    Complete employee attendance summary for a given month.
    Returns: attendance breakdown, late punch penalty, total salary penalty days.
    """
    emp = db.query(Employee).filter(Employee.pr_number == pr_number).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    target_month = month or datetime.now().strftime("%Y-%m")
    year, mon = parse_year_month(target_month)
    start = date(year, mon, 1)
    end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)

    # Get all attendance records for the month
    records = db.query(Attendance).filter(
        Attendance.pr_number == pr_number,
        Attendance.date >= start,
        Attendance.date < end
    ).order_by(Attendance.date).all()

    # Get late punch penalty
    late_penalty = db.query(LatePunchPenalty).filter(
        LatePunchPenalty.pr_number == pr_number,
        LatePunchPenalty.month == target_month
    ).first()

    if not late_penalty:
        # Auto-calculate if not exists
        late_count = db.query(Attendance).filter(
            Attendance.pr_number == pr_number,
            Attendance.date >= start,
            Attendance.date < end,
            Attendance.late_minutes > 0
        ).count()
        penalty = calculate_late_punch_penalty(late_count)
        late_penalty_data = {
            "late_count": late_count,
            "penalty_days": penalty["penalty_days"],
            "penalty_label": penalty["penalty"],
            "next_penalty_at": penalty["next_penalty_at"],
            "action_required": penalty["action_required"]
        }
    else:
        penalty = calculate_late_punch_penalty(late_penalty.late_count)
        late_penalty_data = {
            "late_count": late_penalty.late_count,
            "penalty_days": late_penalty.penalty_days,
            "penalty_label": penalty["penalty"],
            "next_penalty_at": penalty["next_penalty_at"],
            "action_required": penalty["action_required"]
        }

    # Build day-by-day breakdown
    days = []
    counts = {
        "present": 0, "absent": 0, "half_day": 0, "leave": 0,
        "weekoff": 0, "single_punch": 0, "late_punch": 0,
        "early_departure": 0, "less_working_hours": 0,
        "no_data": 0
    }

    for r in records:
        status = r.attendance_status or "No Data"
        issue = r.issue or "-"

        # Count categories
        if status == "Present":
            counts["present"] += 1
        elif status == "Absent":
            counts["absent"] += 1
        elif status == "Half Day":
            counts["half_day"] += 1
        elif status == "Leave":
            counts["leave"] += 1
        elif status == "Week Off":
            counts["weekoff"] += 1
        elif status == "Single Punch":
            counts["single_punch"] += 1
        else:
            counts["no_data"] += 1

        # Count issues
        if r.late_minutes and r.late_minutes > 0:
            counts["late_punch"] += 1
        if r.early_minutes and r.early_minutes > 0:
            counts["early_departure"] += 1
        if issue == "Less Working Hours":
            counts["less_working_hours"] += 1

        days.append({
            "date": r.date.isoformat(),
            "attendance_status": status,
            "issue": issue,
            "essl_in": r.essl_in,
            "essl_out": r.essl_out,
            "tata_in": r.tata_in,
            "tata_out": r.tata_out,
            "final_in": r.final_in,
            "final_out": r.final_out,
            "worked_hours": r.worked_hours,
            "ot_hours": r.ot_hours,
            "late_minutes": r.late_minutes,
            "early_minutes": r.early_minutes,
            "single_punch": r.single_punch,
            "shift": r.shift,
            "remark": r.remark
        })

    # Calculate total penalty days for salary deduction
    # Late punch penalty + half days count as 0.5 each
    total_penalty_days = late_penalty_data["penalty_days"] + (counts["half_day"] * 0.5)

    # Total OT hours for the month
    total_ot_hours = round(sum(r.ot_hours or 0 for r in records), 2)

    # Total working days in month
    total_days = counts["present"] + counts["absent"] + counts["half_day"] + counts["leave"] + counts["weekoff"] + counts["single_punch"]

    # Attendance percentage (present + half_day counted as present) / total working days
    effective_present = counts["present"] + counts["half_day"]
    attendance_percentage = round((effective_present / total_days * 100), 2) if total_days > 0 else 0

    return {
        "pr_number": pr_number,
        "name": emp.name,
        "month": target_month,
        "employee": {
            "name": emp.name,
            "vendor": emp.vendor,
            "store": emp.store,
            "department": emp.department,
            "designation": emp.designation,
            "category": emp.wc or emp.bc or "BC",
            "shift": emp.shift,
            "status": emp.status
        },
        "summary": {
            "total_days": total_days,
            "present": counts["present"],
            "absent": counts["absent"],
            "half_day": counts["half_day"],
            "leave": counts["leave"],
            "weekoff": counts["weekoff"],
            "single_punch": counts["single_punch"],
            "late_punch": counts["late_punch"],
            "early_departure": counts["early_departure"],
            "less_working_hours": counts["less_working_hours"],
            "no_data": counts["no_data"],
            "attendance_percentage": attendance_percentage,
            "effective_present_days": effective_present,
            "total_ot_hours": total_ot_hours
        },
        "penalty": {
            "late_punch_penalty_days": late_penalty_data["penalty_days"],
            "late_punch_count": late_penalty_data["late_count"],
            "late_punch_label": late_penalty_data["penalty_label"],
            "next_penalty_at": late_penalty_data["next_penalty_at"],
            "half_day_penalty_days": counts["half_day"] * 0.5,
            "total_penalty_days": total_penalty_days,
            "action_required": late_penalty_data["action_required"]
        },
        "days": days
    }

@app.get("/api/v1/alerts/late-punch-penalties")
def get_late_punch_penalty_alerts(month: str = Query(...), db: Session = Depends(get_db)):
    """List employees who currently have a non-zero late-punch penalty for the month.
    Run /api/v1/late-punch/calculate-all first (or after new attendance data lands) so
    this reflects the latest counts."""
    rows = db.query(LatePunchPenalty).filter(LatePunchPenalty.month == month, LatePunchPenalty.penalty_days > 0).order_by(desc(LatePunchPenalty.penalty_days)).all()
    emp_map = {e.pr_number: e for e in db.query(Employee).filter(Employee.pr_number.in_([r.pr_number for r in rows])).all()} if rows else {}
    data = []
    for r in rows:
        emp = emp_map.get(r.pr_number)
        next_penalty_at = ((r.late_count // LATE_PUNCH_RULES["half_day_after"]) + 1) * LATE_PUNCH_RULES["half_day_after"]
        data.append({
            "pr_number": r.pr_number, "name": emp.name if emp else None,
            "store": emp.store if emp else None, "vendor": emp.vendor if emp else None,
            "late_count": r.late_count, "penalty_days": r.penalty_days,
            "next_penalty_at_late_count": next_penalty_at
        })
    return {"month": month, "total": len(data), "data": data}

@app.post("/api/v1/alerts/behavioral/calculate")
def calculate_behavioral_alerts(month: str = Query(...), consecutive_days: int = Query(BEHAVIORAL_ALERT_CONFIG["default_consecutive_days"], ge=2, le=31), db: Session = Depends(get_db)):
    """Scan the month's attendance for each active employee and flag anyone with a
    run of `consecutive_days` (default 3) calendar-consecutive days of Early Departure,
    or the same for Half Day. No penalty is attached - this is purely a heads-up so HR
    can catch a pattern before it becomes one. Re-running replaces the prior alerts for
    this month."""
    year, mon = parse_year_month(month)
    start = date(year, mon, 1)
    end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)
    employees = db.query(Employee).filter(Employee.status == "ACTIVE").all()

    db.query(BehavioralAlert).filter(BehavioralAlert.month == month).delete()

    alerts_created = []
    for emp in employees:
        rows = db.query(Attendance).filter(
            Attendance.pr_number == emp.pr_number, Attendance.date >= start, Attendance.date < end
        ).order_by(Attendance.date.asc()).all()
        if not rows:
            continue

        early_streak = _find_max_consecutive_streak(rows, lambda r: bool(r.early_minutes) and r.early_minutes > 0)
        half_day_streak = _find_max_consecutive_streak(rows, lambda r: r.attendance_status == "Half Day")

        for alert_type, streak in (("Early Departure", early_streak), ("Half Day", half_day_streak)):
            if streak["length"] >= consecutive_days:
                db.add(BehavioralAlert(
                    pr_number=emp.pr_number, month=month, alert_type=alert_type,
                    streak_length=streak["length"], streak_start=streak["start"], streak_end=streak["end"],
                    threshold_used=consecutive_days
                ))
                alerts_created.append({
                    "pr_number": emp.pr_number, "name": emp.name, "alert_type": alert_type,
                    "streak_length": streak["length"],
                    "streak_start": streak["start"].isoformat() if streak["start"] else None,
                    "streak_end": streak["end"].isoformat() if streak["end"] else None
                })
    db.commit()
    return {"status": "completed", "month": month, "consecutive_days_threshold": consecutive_days,
            "employees_scanned": len(employees), "alerts_found": len(alerts_created), "alerts": alerts_created}

@app.get("/api/v1/alerts/behavioral")
def get_behavioral_alerts(month: str = Query(...), alert_type: Optional[str] = Query(None, description="Filter to 'Early Departure' or 'Half Day'"), db: Session = Depends(get_db)):
    """Read back the alerts produced by the most recent /api/v1/alerts/behavioral/calculate run."""
    q = db.query(BehavioralAlert).filter(BehavioralAlert.month == month)
    if alert_type:
        q = q.filter(BehavioralAlert.alert_type == alert_type)
    rows = q.order_by(desc(BehavioralAlert.streak_length)).all()
    emp_map = {e.pr_number: e for e in db.query(Employee).filter(Employee.pr_number.in_([r.pr_number for r in rows])).all()} if rows else {}
    data = []
    for r in rows:
        emp = emp_map.get(r.pr_number)
        data.append({
            "pr_number": r.pr_number, "name": emp.name if emp else None,
            "store": emp.store if emp else None, "vendor": emp.vendor if emp else None,
            "alert_type": r.alert_type, "streak_length": r.streak_length,
            "streak_start": r.streak_start.isoformat() if r.streak_start else None,
            "streak_end": r.streak_end.isoformat() if r.streak_end else None,
            "threshold_used": r.threshold_used
        })
    return {"month": month, "total": len(data), "data": data}

@app.get("/api/v1/alerts/summary")
def get_alerts_summary(month: str = Query(...), db: Session = Depends(get_db)):
    """Single counts endpoint for a dashboard tile - combines late-punch penalty
    alerts with the two behavioral streak alerts."""
    penalized = db.query(LatePunchPenalty).filter(LatePunchPenalty.month == month, LatePunchPenalty.penalty_days > 0).count()
    early_dep = db.query(BehavioralAlert).filter(BehavioralAlert.month == month, BehavioralAlert.alert_type == "Early Departure").count()
    half_day = db.query(BehavioralAlert).filter(BehavioralAlert.month == month, BehavioralAlert.alert_type == "Half Day").count()
    return {
        "month": month,
        "late_punch_penalty_alerts": penalized,
        "early_departure_streak_alerts": early_dep,
        "half_day_streak_alerts": half_day,
        "total_alerts": penalized + early_dep + half_day
    }

# ============================================================================
# OT THRESHOLD ALERTS
# ============================================================================

@app.post("/api/v1/alerts/ot-thresholds/calculate")
def calculate_ot_threshold_alerts(
    month: str = Query(..., description="YYYY-MM format"),
    db: Session = Depends(get_db)
):
    """Batch-calculate OT threshold alerts for a given month.
    Scans all OT records and creates HRAction alerts where weekly > 12h or monthly > 48h.

    PERF FIX: previously fired ~5 queries per unique (employee, date) pair with
    OT that month (1 employee lookup + 2 rolling-sum queries + up to 2 existing-
    alert checks via get_ot_totals/check_ot_thresholds) - for a few hundred
    employees x ~20+ OT days that was tens of thousands of sequential queries
    and could make "Recalculate" hang for minutes or time out. This now runs a
    fixed, small number of bulk queries regardless of month size, and does the
    rolling weekly/monthly sums in memory. Behavior (thresholds, dedup rules,
    alert descriptions) is unchanged from the original per-key logic.
    """
    year, mon = parse_year_month(month)
    start = date(year, mon, 1)
    end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)
    # Rolling 7-day weekly window can reach up to 6 days before month start.
    lookback_start = start - timedelta(days=6)

    ot_records = db.query(Overtime).filter(
        Overtime.date >= start,
        Overtime.date < end,
        Overtime.ot_hours > 0
    ).order_by(Overtime.pr_number, Overtime.date).all()

    if not ot_records:
        return {"status": "completed", "month": month, "alerts_created": 0, "details": []}

    pr_numbers = sorted({r.pr_number for r in ot_records})

    # One bulk fetch (instead of one query per key) of every OT row needed for
    # rolling-week math, including the 6-day lookback before the month starts.
    all_ot_rows = db.query(Overtime.pr_number, Overtime.date, Overtime.ot_hours).filter(
        Overtime.pr_number.in_(pr_numbers),
        Overtime.date >= lookback_start,
        Overtime.date < end,
        Overtime.ot_hours > 0
    ).all()
    ot_by_pr = defaultdict(list)  # pr_number -> [(date, hours), ...]
    for pr, d, hrs in all_ot_rows:
        ot_by_pr[pr].append((d, hrs or 0.0))

    # One bulk fetch of employee names instead of one lookup per key.
    employees = {
        e.pr_number: e.name
        for e in db.query(Employee.pr_number, Employee.name).filter(Employee.pr_number.in_(pr_numbers)).all()
    }

    # One bulk fetch of already-created alerts instead of an existence check per key.
    existing_weekly = db.query(HRAction.pr_number, HRAction.date).filter(
        HRAction.action_type == "OT Weekly Threshold",
        HRAction.date >= lookback_start, HRAction.date < end,
        HRAction.pr_number.in_(pr_numbers)
    ).all()
    existing_monthly_prs = {
        pr for pr, in db.query(HRAction.pr_number).filter(
            HRAction.action_type == "OT Monthly Threshold",
            HRAction.date >= start, HRAction.date < end,
            HRAction.pr_number.in_(pr_numbers)
        ).all()
    }
    weekly_alert_dates = defaultdict(list)
    for pr, d in existing_weekly:
        weekly_alert_dates[pr].append(d)

    alerts_created = []
    checked_keys = set()
    new_weekly_flags = defaultdict(list)  # dedup against alerts created earlier in this same run

    for ot in ot_records:
        key = (ot.pr_number, ot.date)
        if key in checked_keys:
            continue
        checked_keys.add(key)

        pr = ot.pr_number
        ref_date = ot.date
        emp_name = employees.get(pr, pr)
        week_start = ref_date - timedelta(days=6)
        month_start = date(ref_date.year, ref_date.month, 1)

        weekly_ot = round(sum(h for d, h in ot_by_pr.get(pr, []) if week_start <= d <= ref_date), 2)
        monthly_ot = round(sum(h for d, h in ot_by_pr.get(pr, []) if month_start <= d < end), 2)

        record_alerts = []

        if weekly_ot > OT_WEEKLY_THRESHOLD:
            already_flagged = any(
                week_start <= d <= ref_date
                for d in weekly_alert_dates.get(pr, []) + new_weekly_flags.get(pr, [])
            )
            if not already_flagged:
                db.add(HRAction(
                    pr_number=pr, emp_name=emp_name, date=ref_date,
                    action_type="OT Weekly Threshold",
                    description=(
                        f"{emp_name} ({pr}) has exceeded the weekly OT limit: "
                        f"{weekly_ot}h / {OT_WEEKLY_THRESHOLD}h (week ending {ref_date.isoformat()})"
                    ),
                    priority="High", assigned_to="HR Manager"
                ))
                new_weekly_flags[pr].append(ref_date)
                record_alerts.append({"type": "weekly", "hours": weekly_ot, "threshold": OT_WEEKLY_THRESHOLD})

        if monthly_ot > OT_MONTHLY_THRESHOLD and pr not in existing_monthly_prs:
            db.add(HRAction(
                pr_number=pr, emp_name=emp_name, date=ref_date,
                action_type="OT Monthly Threshold",
                description=(
                    f"{emp_name} ({pr}) has exceeded the monthly OT limit: "
                    f"{monthly_ot}h / {OT_MONTHLY_THRESHOLD}h for {ref_date.strftime('%Y-%m')}"
                ),
                priority="Critical", assigned_to="HR Manager"
            ))
            existing_monthly_prs.add(pr)
            record_alerts.append({"type": "monthly", "hours": monthly_ot, "threshold": OT_MONTHLY_THRESHOLD})

        if record_alerts:
            alerts_created.append({"pr_number": pr, "name": emp_name, "date": ref_date.isoformat(), "alerts": record_alerts})

    db.commit()
    return {
        "status": "completed",
        "month": month,
        "alerts_created": len(alerts_created),
        "details": alerts_created
    }


@app.get("/api/v1/alerts/ot-thresholds")
def get_ot_threshold_alerts(
    month: str = Query(..., description="YYYY-MM format"),
    db: Session = Depends(get_db)
):
    """List all open OT threshold alerts for the month."""
    year, mon = parse_year_month(month)
    start = date(year, mon, 1)
    end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)

    actions = db.query(HRAction).filter(
        HRAction.date >= start,
        HRAction.date < end,
        HRAction.action_type.in_(["OT Weekly Threshold", "OT Monthly Threshold"]),
        HRAction.status == "Open"
    ).order_by(desc(HRAction.created_at)).all()

    return {
        "month": month,
        "total": len(actions),
        "data": [
            {
                "id": a.id,
                "pr_number": a.pr_number,
                "name": a.emp_name,
                "date": a.date.isoformat(),
                "action_type": a.action_type,
                "description": a.description,
                "priority": a.priority,
                "status": a.status
            }
            for a in actions
        ]
    }


@app.get("/api/v1/alerts/ot-thresholds/summary")
def get_ot_threshold_summary(
    month: str = Query(..., description="YYYY-MM format"),
    db: Session = Depends(get_db)
):
    """Dashboard tile: count of weekly and monthly OT threshold breaches."""
    year, mon = parse_year_month(month)
    start = date(year, mon, 1)
    end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)

    weekly = db.query(HRAction).filter(
        HRAction.date >= start,
        HRAction.date < end,
        HRAction.action_type == "OT Weekly Threshold",
        HRAction.status == "Open"
    ).count()

    monthly = db.query(HRAction).filter(
        HRAction.date >= start,
        HRAction.date < end,
        HRAction.action_type == "OT Monthly Threshold",
        HRAction.status == "Open"
    ).count()

    return {
        "month": month,
        "weekly_breaches": weekly,
        "monthly_breaches": monthly,
        "total_breaches": weekly + monthly
    }


# ============================================================================
# DUMP REPORT
# ============================================================================




@app.get("/api/v1/reports/dump/download")
def download_dump_report(
    month: Optional[str] = Query(None),
    target_date: Optional[str] = Query(None),
    vendor: Optional[str] = Query(None),
    store: Optional[str] = Query(None),
    department: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Download Dump Report as Excel.

    - If `target_date` is provided: single sheet with that date's data.
    - If `month` is provided: one sheet per date in that month + Summary sheet.
    - If neither: uses current month.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import OrderedDict
    import time

    start_time = time.time()

    try:
        # ── Build query ──
        q = db.query(Attendance).join(Employee)

        if target_date:
            d = parse_date_br(target_date)
            if not d:
                raise HTTPException(status_code=400, detail="Invalid target_date. Use YYYY-MM-DD.")
            q = q.filter(Attendance.date == d)
            filename = f"seabird_dump_{d.isoformat()}.xlsx"
        elif month:
            year, mon = parse_year_month(month)
            start = date(year, mon, 1)
            end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)
            q = q.filter(Attendance.date >= start, Attendance.date < end)
            filename = f"seabird_dump_{month}.xlsx"
        else:
            today = date.today()
            start = date(today.year, today.month, 1)
            end = date(today.year, today.month + 1, 1) if today.month < 12 else date(today.year + 1, 1, 1)
            q = q.filter(Attendance.date >= start, Attendance.date < end)
            filename = f"seabird_dump_{today.strftime('%Y-%m')}.xlsx"

        if vendor: q = q.filter(Employee.vendor == vendor)
        if store: q = q.filter(Employee.store == store)
        if department: q = q.filter(Employee.department == department)

        # OPTIMIZATION: Count first to avoid empty result processing
        count = q.count()
        if count == 0:
            raise HTTPException(status_code=404, detail="No attendance records found for the given criteria.")

        print(f"[DUMP] Generating report for {count} records...")

        # OPTIMIZATION: Stream in chunks to avoid memory bloat
        records = q.order_by(Attendance.date.desc(), Attendance.pr_number).yield_per(500).all()

        # ── Group records by date ──
        records_by_date: OrderedDict[date, list] = OrderedDict()
        for r in records:
            d = r.date
            if d not in records_by_date:
                records_by_date[d] = []
            records_by_date[d].append(r)

        print(f"[DUMP] Grouped into {len(records_by_date)} date sheets")

        # ── Build Excel workbook ──
        wb = Workbook()
        wb.remove(wb.active)

        # Shared styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        alt_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color="D1D5DB"),
            right=Side(style='thin', color="D1D5DB"),
            top=Side(style='thin', color="D1D5DB"),
            bottom=Side(style='thin', color="D1D5DB")
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        headers = ["BioID", "PR", "WC/BC", "Vendor", "Name", "ESSL In", "ESSL Out", "Shift", "Total", "Remark", "Tata In", "Tata Out"]

        def style_header_row(ws, row_num=1):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

        def auto_width(ws, custom_headers=None):
            hdrs = custom_headers or headers
            for col_idx in range(1, len(hdrs) + 1):
                max_length = len(hdrs[col_idx - 1])
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 45)

        def write_data_rows(ws, recs, start_row=2):
            for i, r in enumerate(recs):
                emp = r.employee
                row_num = start_row + i
                # Compute remark from punch times for multi-shift detection
                remark = determine_worked_shift(r.essl_in, r.essl_out, r.tata_in, r.tata_out, r.shift, r.remark)
                ws.append([
                    emp.bio_id if emp else "",
                    r.pr_number,
                    r.category or (emp.wc if emp else ""),
                    r.vendor or (emp.vendor if emp else ""),
                    emp.name if emp else r.pr_number,
                    r.essl_in or "",
                    r.essl_out or "",
                    r.shift or "",       # Shift = assigned shift
                    r.worked_hours or 0,  # Total
                    remark,               # Remark = multi-shift info or DB value
                    r.tata_in or "",
                    r.tata_out or ""
                ])
                if i % 2 == 1:
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row_num, column=col).fill = alt_fill
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).border = thin_border
                    ws.cell(row=row_num, column=col).alignment = left_align

        # ── Single day mode ──
        if target_date:
            d = parse_date_br(target_date)
            sheet_name = d.strftime("%d-%b-%Y")
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)
            style_header_row(ws)
            write_data_rows(ws, records)
            auto_width(ws)
            ws.freeze_panes = "A2"
        else:
            # ── Monthly mode: Summary sheet + one sheet per date ──

            # 1. Summary sheet
            summary_ws = wb.create_sheet(title="Summary", index=0)
            summary_headers = ["Date", "Total", "Present", "Absent", "Late", "OT Hrs", "Single Punch", "Issues"]
            summary_ws.append(summary_headers)
            for col in range(1, len(summary_headers) + 1):
                cell = summary_ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            for d, recs in records_by_date.items():
                present = sum(1 for r in recs if r.attendance_status not in ["Absent", "Leave", "Week Off"])
                absent = sum(1 for r in recs if r.attendance_status in ["Absent", "Week Off"])
                late = sum(1 for r in recs if r.late_minutes > 0)
                ot = sum(r.ot_hours or 0 for r in recs)
                sp = sum(1 for r in recs if r.single_punch == "Yes")
                issues = sum(1 for r in recs if r.issue and r.issue != "-")
                summary_ws.append([
                    d.strftime("%d-%b-%Y"),
                    len(recs),
                    present,
                    absent,
                    late,
                    round(ot, 2),
                    sp,
                    issues
                ])

            for row in summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            auto_width(summary_ws, summary_headers)
            summary_ws.freeze_panes = "A2"

            # 2. One sheet per date (chronological)
            used_sheet_names = set()
            for d in sorted(records_by_date.keys()):
                recs = records_by_date[d]
                sheet_name = d.strftime("%d-%b-%Y")  # e.g. "01-Jun-2026"
                # Deduplicate sheet names (Excel requires unique names)
                original_name = sheet_name
                counter = 1
                while sheet_name in used_sheet_names:
                    suffix = f"_{counter}"
                    sheet_name = (original_name[:31 - len(suffix)]) + suffix
                    counter += 1
                used_sheet_names.add(sheet_name)
                sheet_name = sheet_name[:31]  # Excel max 31 chars

                ws = wb.create_sheet(title=sheet_name)
                ws.append(headers)
                style_header_row(ws)
                write_data_rows(ws, recs)
                auto_width(ws)
                ws.freeze_panes = "A2"

        # ── Save to buffer ──
        buf = bio.BytesIO()
        wb.save(buf)
        buf.seek(0)
        elapsed = round(time.time() - start_time, 2)
        print(f"[DUMP] Report generated in {elapsed}s, size={len(buf.getvalue())} bytes")

        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        elapsed = round(time.time() - start_time, 2)
        print(f"[DUMP ERROR] After {elapsed}s: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Report generation failed after {elapsed}s: {str(e)}")

@app.get("/api/v1/reports/dump")
def get_dump_report(month: Optional[str] = Query(None), vendor: Optional[str] = Query(None),
                     store: Optional[str] = Query(None), department: Optional[str] = Query(None),
                     page: int = Query(1, ge=1), page_size: int = Query(50, ge=1, le=500), db: Session = Depends(get_db)):
    q = db.query(Attendance).join(Employee)
    if month:
        year, mon = parse_year_month(month)
        start = date(year, mon, 1)
        end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)
        q = q.filter(Attendance.date >= start, Attendance.date < end)
    if vendor: q = q.filter(Employee.vendor == vendor)
    if store: q = q.filter(Employee.store == store)
    if department: q = q.filter(Employee.department == department)

    total = q.count()
    items = q.order_by(Attendance.date.desc(), Attendance.pr_number).offset((page - 1) * page_size).limit(page_size).all()
    data = []
    for r in items:
        emp = r.employee
        data.append({
            "id": r.id, "bio_id": emp.bio_id if emp else "", "pr_number": r.pr_number,
            "category": r.category or (emp.wc if emp else ""), "vendor": r.vendor,
            "name": emp.name if emp else r.pr_number, "essl_in": r.essl_in or "", "essl_out": r.essl_out or "",
            "shift": r.shift, "total_hours": r.worked_hours, "remark": r.remark or "",
            "tata_in": r.tata_in or "", "tata_out": r.tata_out or "", "attendance_status": r.attendance_status,
            "ot_hours": r.ot_hours, "ot_headcount": r.ot_headcount, "late_minutes": r.late_minutes,
            "early_minutes": r.early_minutes, "single_punch": r.single_punch, "is_match": r.is_match,
            "match_status": r.match_status or ""
        })
    return {"data": data, "total": total, "page": page, "total_pages": max(1, (total + page_size - 1) // page_size)}

@app.get("/api/v1/reports/ot/download")
def download_ot_report(month: Optional[str] = Query(None), vendor: Optional[str] = Query(None),
                        store: Optional[str] = Query(None), db: Session = Depends(get_db)):
    q = db.query(Overtime).join(Employee, Overtime.pr_number == Employee.pr_number)
    if month:
        year, mon = parse_year_month(month)
        start = date(year, mon, 1)
        end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)
        q = q.filter(Overtime.date >= start, Overtime.date < end)
    if vendor: q = q.filter(Employee.vendor == vendor)
    if store: q = q.filter(Employee.store == store)
    records = q.order_by(Overtime.date.desc(), Overtime.pr_number).limit(50000).all()

    output = bio.StringIO()
    writer = csv.writer(output)
    writer.writerow(["PR", "Name", "Store", "Date", "Worked Hours", "OT Hours", "OT Headcount", "Status", "Approved By"])
    for r in records:
        writer.writerow([r.pr_number, r.name, r.store, r.date.isoformat(), r.worked_hours, r.ot_hours, r.calc_headcount, r.status, r.approved_by or ""])
    output.seek(0)
    filename = f"seabird_ot_{month or date.today().strftime('%Y-%m')}.csv"
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={filename}"})

@app.get("/api/v1/reports/late/download")
def download_late_report(month: Optional[str] = Query(None), vendor: Optional[str] = Query(None),
                          store: Optional[str] = Query(None), db: Session = Depends(get_db)):
    q = db.query(Attendance).join(Employee).filter(Attendance.late_minutes > 0)
    if month:
        year, mon = parse_year_month(month)
        start = date(year, mon, 1)
        end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)
        q = q.filter(Attendance.date >= start, Attendance.date < end)
    if vendor: q = q.filter(Employee.vendor == vendor)
    if store: q = q.filter(Employee.store == store)
    records = q.order_by(Attendance.date.desc(), Attendance.pr_number).limit(50000).all()

    output = bio.StringIO()
    writer = csv.writer(output)
    writer.writerow(["PR", "Name", "Store", "Date", "Shift", "Final In", "Late Minutes"])
    for r in records:
        emp = r.employee
        writer.writerow([r.pr_number, emp.name if emp else "", r.store, r.date.isoformat(), r.shift, r.final_in or "", r.late_minutes])
    output.seek(0)
    filename = f"seabird_late_punch_{month or date.today().strftime('%Y-%m')}.csv"
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={filename}"})

@app.get("/api/v1/reports/monthly/download")
def download_monthly_report(month: Optional[str] = Query(None), db: Session = Depends(get_db)):
    ym = month or date.today().strftime("%Y-%m")
    # Validate the month format even though we only use it for string comparison
    parse_year_month(ym)
    records = db.query(MonthlyTataAttendance).filter(MonthlyTataAttendance.year_month == ym).order_by(MonthlyTataAttendance.pr_number).all()
    output = bio.StringIO()
    writer = csv.writer(output)
    writer.writerow(["PR", "Name", "Month", "Total Days", "Present", "Absent", "Leave", "Week Off", "Half Day", "Attendance %", "Total OT Hours"])
    for r in records:
        writer.writerow([r.pr_number, r.emp_name, r.year_month, r.total_days, r.present_days, r.absent_days, r.leave_days, r.weekoff_days, r.half_days, r.attendance_percentage, r.total_ot_hours])
    output.seek(0)
    filename = f"seabird_monthly_{ym}.csv"
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={filename}"})

# ============================================================================
# DATA QUALITY
# ============================================================================

@app.get("/api/v1/data-quality/unmatched")
def get_unmatched_records(source: str = Query(..., pattern="^(essl|tata)$"), db: Session = Depends(get_db)):
    if source == "essl":
        q = db.query(ESSLAttendance.pr_number, ESSLAttendance.emp_name, func.count(ESSLAttendance.id)).filter(
            ESSLAttendance.employee_id.is_(None)
        ).group_by(ESSLAttendance.pr_number, ESSLAttendance.emp_name).limit(200).all()
    else:
        q = db.query(TataAttendance.pr_number, TataAttendance.emp_name, func.count(TataAttendance.id)).filter(
            TataAttendance.employee_id.is_(None)
        ).group_by(TataAttendance.pr_number, TataAttendance.emp_name).limit(200).all()
    return [{"pr_number": r[0], "name": r[1], "record_count": r[2]} for r in q]

# ============================================================================
# TATA-ONLY FALLBACK ENDPOINTS
# ============================================================================

@app.get("/api/v1/tata-only/employees")
def get_tata_only_employees_endpoint(
    month: str = Query(..., description="YYYY-MM format"),
    db: Session = Depends(get_db)
):
    """List all employees who are Tata-only for the given month (have Tata data but NO ESSL data)."""
    year, mon = parse_year_month(month)
    start = date(year, mon, 1)
    end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)

    employees = db.query(Employee).filter(Employee.status == "ACTIVE").all()
    pr_numbers = [e.pr_number for e in employees if e.pr_number]

    tata_only_prs = get_tata_only_employees(pr_numbers, start, end, db)

    result = []
    for emp in employees:
        if emp.pr_number in tata_only_prs:
            # Get Tata data summary for this employee
            tata_count = db.query(TataAttendance).filter(
                TataAttendance.pr_number == emp.pr_number,
                TataAttendance.date >= start,
                TataAttendance.date < end
            ).count()
            result.append({
                "pr_number": emp.pr_number,
                "name": emp.name,
                "vendor": emp.vendor,
                "store": emp.store,
                "department": emp.department,
                "shift": emp.shift,
                "tata_records": tata_count,
                "essl_records": 0,
                "reason": "No ESSL data for this month - using Tata only"
            })

    return {
        "month": month,
        "tata_only_count": len(result),
        "total_active": len(employees),
        "employees": result
    }

@app.get("/api/v1/tata-only/summary")
def get_tata_only_summary(
    month: str = Query(..., description="YYYY-MM format"),
    db: Session = Depends(get_db)
):
    """Summary statistics for Tata-only employees in a month."""
    year, mon = parse_year_month(month)
    start = date(year, mon, 1)
    end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)

    employees = db.query(Employee).filter(Employee.status == "ACTIVE").all()
    pr_numbers = [e.pr_number for e in employees if e.pr_number]

    # Count employees with ESSL data
    essl_prs = db.query(ESSLAttendance.pr_number).filter(
        ESSLAttendance.pr_number.in_(pr_numbers),
        ESSLAttendance.date >= start,
        ESSLAttendance.date < end
    ).distinct().all()
    with_essl = len(set(r[0] for r in essl_prs if r[0]))

    # Count employees with Tata data
    tata_prs = db.query(TataAttendance.pr_number).filter(
        TataAttendance.pr_number.in_(pr_numbers),
        TataAttendance.date >= start,
        TataAttendance.date < end
    ).distinct().all()
    with_tata = len(set(r[0] for r in tata_prs if r[0]))

    # Tata-only = have Tata but no ESSL
    tata_only_prs = get_tata_only_employees(pr_numbers, start, end, db)

    return {
        "month": month,
        "total_active_employees": len(employees),
        "with_essl_data": with_essl,
        "with_tata_data": with_tata,
        "tata_only_employees": len(tata_only_prs),
        "both_essl_and_tata": with_essl,  # Since tata_only = tata - essl
        "percentage_tata_only": round((len(tata_only_prs) / len(employees) * 100), 2) if employees else 0
    }

@app.get("/api/v1/tata-only/check")
def check_employee_tata_only(
    pr_number: str = Query(...),
    month: str = Query(..., description="YYYY-MM format"),
    db: Session = Depends(get_db)
):
    """Check if a specific employee is Tata-only for a given month."""
    year, mon = parse_year_month(month)
    start = date(year, mon, 1)
    end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)

    emp = db.query(Employee).filter(Employee.pr_number == pr_number).first()
    if not emp:
        raise HTTPException(status_code=404, detail=f"Employee {pr_number} not found")

    essl_count = db.query(ESSLAttendance).filter(
        ESSLAttendance.pr_number == pr_number,
        ESSLAttendance.date >= start,
        ESSLAttendance.date < end
    ).count()

    tata_count = db.query(TataAttendance).filter(
        TataAttendance.pr_number == pr_number,
        TataAttendance.date >= start,
        TataAttendance.date < end
    ).count()

    is_tata_only = (tata_count > 0 and essl_count == 0)

    return {
        "pr_number": pr_number,
        "name": emp.name,
        "month": month,
        "essl_records": essl_count,
        "tata_records": tata_count,
        "is_tata_only": is_tata_only,
        "data_source": "Tata Only" if is_tata_only else ("ESSL + Tata" if essl_count > 0 else "No Data")
    }

# ============================================================================
# DEBUG / DIAGNOSTIC ENDPOINTS — See exactly what's in your database
# ============================================================================

@app.get("/debug/tables")
def debug_tables(db: Session = Depends(get_db)):
    """Show all tables, row counts, and last updated times."""
    tables = [
        ("vendors", Vendor),
        ("stores", Store),
        ("departments", Department),
        ("employees", Employee),
        ("essl_attendance", ESSLAttendance),
        ("tata_attendance", TataAttendance),
        ("attendance", Attendance),
        ("attendance_reconciliation", AttendanceReconciliation),
        ("hr_actions", HRAction),
        ("overtime", Overtime),
        ("upload_logs", UploadLog),
    ]
    result = []
    for name, model in tables:
        count = db.query(model).count()
        # Get last record timestamp if available
        last = db.query(model).order_by(desc(model.id)).first() if hasattr(model, 'id') else None
        last_time = None
        if last and hasattr(last, 'created_at'):
            last_time = last.created_at.isoformat() if last.created_at else None
        result.append({
            "table": name,
            "row_count": count,
            "last_record_at": last_time
        })
    return {"tables": result, "database": str(DATABASE_URL)}

@app.get("/debug/upload-log-test")
def debug_upload_log_test(db: Session = Depends(get_db)):
    """Try to insert+commit a throwaway UploadLog row and report exactly what happens.
    Also compares the live Postgres schema for upload_logs against the ORM model,
    so we can see if a column was added at the DB level that the model doesn't know about."""
    result = {}

    # 1. Compare live schema vs ORM model columns
    try:
        rows = db.execute(text(
            "SELECT column_name, is_nullable, column_default, data_type "
            "FROM information_schema.columns WHERE table_name = 'upload_logs' "
            "ORDER BY ordinal_position"
        )).fetchall()
        result["live_schema"] = [dict(r._mapping) for r in rows]
        model_cols = {c.name for c in UploadLog.__table__.columns}
        live_cols = {r[0] for r in rows}
        result["columns_in_db_not_in_model"] = list(live_cols - model_cols)
        result["columns_in_model_not_in_db"] = list(model_cols - live_cols)
    except Exception as schema_err:
        result["schema_check_error"] = str(schema_err)

    # 2. Try the actual insert+commit, exactly like the upload endpoints do
    try:
        test_log = UploadLog(file_type="debug_test", filename="test.xlsx",
                              file_hash="debug" + str(datetime.utcnow().timestamp()),
                              rows_processed=0, status="Success")
        db.add(test_log)
        db.commit()
        result["insert_test"] = "SUCCESS"
        result["inserted_id"] = test_log.id
        db.delete(test_log)
        db.commit()
    except Exception as insert_err:
        db.rollback()
        import traceback
        result["insert_test"] = "FAILED"
        result["insert_error"] = str(insert_err)
        result["insert_traceback"] = traceback.format_exc()

    return result

@app.get("/debug/employee/{pr_number}")
def debug_employee(pr_number: str, db: Session = Depends(get_db)):
    """Show FULL employee record including internal employee_id."""
    emp = db.query(Employee).filter(Employee.pr_number == pr_number).first()
    if not emp:
        raise HTTPException(status_code=404, detail=f"Employee PR{pr_number} not found")

    # Count related records
    essl_count = db.query(ESSLAttendance).filter(ESSLAttendance.pr_number == pr_number).count()
    tata_count = db.query(TataAttendance).filter(TataAttendance.pr_number == pr_number).count()
    attendance_count = db.query(Attendance).filter(Attendance.pr_number == pr_number).count()

    return {
        "employee_id": emp.id,  # THIS IS THE INTERNAL ID YOU WANTED
        "pr_number": emp.pr_number,
        "bio_id": emp.bio_id,
        "emp_code": emp.emp_code,
        "name": emp.name,
        "vendor": emp.vendor,
        "store": emp.store,
        "department": emp.department,
        "designation": emp.designation,
        "shift": emp.shift,
        "wc": emp.wc,
        "bc": emp.bc,
        "status": emp.status,
        "join_date": emp.join_date.isoformat() if emp.join_date else None,
        "phone": emp.phone,
        "email": emp.email,
        "created_at": emp.created_at.isoformat() if emp.created_at else None,
        "updated_at": emp.updated_at.isoformat() if emp.updated_at else None,
        "record_counts": {
            "essl_records": essl_count,
            "tata_records": tata_count,
            "attendance_records": attendance_count
        }
    }

@app.get("/debug/attendance/{target_date}")
def debug_attendance(target_date: str, db: Session = Depends(get_db)):
    """Show ALL attendance records for a date with FULL details."""
    today = parse_date_br(target_date)
    if not today:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD.")

    records = db.query(Attendance).filter(Attendance.date == today).all()

    # Summary stats
    total = len(records)
    present = sum(1 for r in records if r.attendance_status not in ["Absent", "Leave", "Week Off"])
    absent = sum(1 for r in records if r.attendance_status == "Absent")
    leave = sum(1 for r in records if r.attendance_status == "Leave")
    weekoff = sum(1 for r in records if r.issue == "Week Off")  # FIX v3.2.4: Check issue, not status
    late = sum(1 for r in records if r.late_minutes > 0)
    ot_positive = sum(1 for r in records if r.ot_hours > 0)
    total_ot = sum(r.ot_hours for r in records)
    single_punch = sum(1 for r in records if r.single_punch == "Yes")

    # Show first 20 records in detail
    detail = []
    for r in records[:20]:
        detail.append({
            "attendance_id": r.id,
            "employee_id": r.employee_id,
            "pr_number": r.pr_number,
            "emp_code": r.emp_code,
            "name": r.employee.name if r.employee else "NO EMPLOYEE LINK",
            "shift": r.shift,
            "master_shift": r.employee.shift if r.employee else None,
            "attendance_status": r.attendance_status,
            "essl_in": r.essl_in,
            "essl_out": r.essl_out,
            "tata_in": r.tata_in,
            "tata_out": r.tata_out,
            "final_in": r.final_in,
            "final_out": r.final_out,
            "worked_hours": r.worked_hours,
            "man_hrs": r.man_hrs,
            "ot_hours": r.ot_hours,
            "ot_headcount": r.ot_headcount,
            "late_minutes": r.late_minutes,
            "early_minutes": r.early_minutes,
            "single_punch": r.single_punch,
            "is_match": r.is_match,
            "match_status": r.match_status,
            "category": r.category,
            "remark": r.remark
        })

    return {
        "date": today.isoformat(),
        "summary": {
            "total_records": total,
            "present": present,
            "absent": absent,
            "leave": leave,
            "weekoff": weekoff,
            "late": late,
            "ot_positive_records": ot_positive,
            "total_ot_hours": round(total_ot, 2),
            "single_punch": single_punch
        },
        "first_20_records": detail,
        "note": f"Showing first 20 of {total} records. Use pagination for more."
    }

@app.get("/debug/reconcile-check")
def debug_reconcile_check(
    pr_number: str = Query(...),
    target_date: str = Query(...),
    db: Session = Depends(get_db)
):
    """Run reconciliation logic on ONE employee for ONE date and show EVERY step."""
    today = parse_date_br(target_date)
    if not today:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD.")

    emp = db.query(Employee).filter(Employee.pr_number == pr_number).first()
    if not emp:
        raise HTTPException(status_code=404, detail=f"Employee PR{pr_number} not found in Master")

    # Fetch raw data
    essl = db.query(ESSLAttendance).filter(
        ESSLAttendance.pr_number == pr_number,
        ESSLAttendance.date == today
    ).first()

    tata = db.query(TataAttendance).filter(
        TataAttendance.pr_number == pr_number,
        TataAttendance.date == today
    ).first()

    # Step-by-step reconciliation logic
    essl_in = parse_time_br(essl.in_time) if essl else None
    essl_out = parse_time_br(essl.out_time) if essl else None
    essl_status = str(essl.status).strip().upper() if essl and essl.status else None

    tata_in = parse_time_br(tata.in_time) if tata else None
    tata_out = parse_time_br(tata.out_time) if tata else None
    tata_status = str(tata.status).strip().upper() if tata and tata.status else None
    tata_shift = normalize_shift(tata.shift) if tata and tata.shift else None
    tata_man_hrs = float(tata.man_hrs) if tata and tata.man_hrs else 0.0

    # Final IN/OUT determination
    final_in = essl_in or tata_in
    final_out = essl_out or tata_out

    # Shift determination
    shift = tata_shift or (normalize_shift(essl.shift) if essl and essl.shift else None) or normalize_shift(emp.shift) or "G"

    # Category
    category = normalize_category(emp.wc)

    # Calculate
    worked_hours = calculate_work_duration(final_in, final_out)
    ot_result = calculate_ot(shift, category, final_in, final_out, tata_man_hrs)

    # Status determination
    status_determination = ""
    attendance_status = "Absent"
    if essl_status == "WO":
        attendance_status = "Week Off"; status_determination = "ESSL status = WO"
    elif essl_status == "A":
        attendance_status = "Absent"; status_determination = "ESSL status = A"
    elif essl_status == "L":
        attendance_status = "Leave"; status_determination = "ESSL status = L"
    elif essl_status == "P":
        attendance_status = "Present"; status_determination = "ESSL status = P"
    elif essl_status == "SP":
        attendance_status = "Present"; status_determination = "ESSL status = SP (single punch)"
    elif tata_status:
        attendance_status = tata_status; status_determination = f"No ESSL, using Tata status = {tata_status}"
    elif final_in and final_out:
        attendance_status = "Present"; status_determination = "No ESSL/Tata status, but has both punches"
    else:
        status_determination = "No ESSL, no Tata status, no punches → Absent"

    return {
        "employee": {
            "employee_id": emp.id,
            "pr_number": emp.pr_number,
            "name": emp.name,
            "master_shift": emp.shift,
            "master_category": emp.wc,
            "normalized_category": category,
            "vendor": emp.vendor,
            "store": emp.store
        },
        "date": today.isoformat(),
        "raw_data": {
            "essl": {
                "found": essl is not None,
                "in_time": essl_in,
                "out_time": essl_out,
                "status": essl_status,
                "shift": essl.shift if essl else None
            },
            "tata": {
                "found": tata is not None,
                "in_time": tata_in,
                "out_time": tata_out,
                "status": tata_status,
                "shift": tata_shift,
                "man_hrs": tata_man_hrs
            }
        },
        "reconciliation_logic": {
            "final_in": final_in,
            "final_out": final_out,
            "final_shift": shift,
            "worked_hours": worked_hours,
            "standard_hours": get_standard_hours(shift),
            "ot_calculation": ot_result,
            "status_determination": status_determination,
            "final_attendance_status": attendance_status,
            "is_late": is_late_punch(shift, final_in),
            "late_minutes": get_late_minutes(shift, final_in) if is_late_punch(shift, final_in) else 0,
            "is_early_departure": is_early_departure(shift, worked_hours),
            "has_single_punch": (final_in is not None) != (final_out is not None)
        }
    }

@app.get("/debug/employees-without-attendance")
def debug_employees_without_attendance(target_date: str = Query(...), db: Session = Depends(get_db)):
    """Show which ACTIVE employees have NO attendance record for a date."""
    today = parse_date_br(target_date)
    if not today:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD.")

    # Get all active employees
    active_employees = db.query(Employee).filter(Employee.status == "ACTIVE").all()

    # Get all PRs with attendance for this date
    attended_prs = set(
        r[0] for r in db.query(Attendance.pr_number).filter(Attendance.date == today).all() if r[0]
    )

    # Find missing
    missing = []
    for emp in active_employees:
        if emp.pr_number not in attended_prs:
            missing.append({
                "employee_id": emp.id,
                "pr_number": emp.pr_number,
                "name": emp.name,
                "vendor": emp.vendor,
                "store": emp.store,
                "shift": emp.shift,
                "reason": "No ESSL or Tata record for this date"
            })

    return {
        "date": today.isoformat(),
        "total_active": len(active_employees),
        "with_attendance": len(attended_prs),
        "without_attendance": len(missing),
        "missing_employees": missing[:50],  # First 50
        "note": f"Showing first 50 of {len(missing)} missing employees"
    }

# ============================================================================
# AUTH ENDPOINTS
# ============================================================================

class LoginRequest(BaseModel):
    username: str
    password: str

class LoginResponse(BaseModel):
    token: str
    user: Dict[str, Any]

@app.post("/api/v1/auth/login", response_model=LoginResponse)
def auth_login(req: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == req.username).first()
    if not user or not _verify_password(req.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    if not user.is_active:
        raise HTTPException(status_code=403, detail="Account disabled")
    # Rotate token on each login
    user.api_token = _generate_token()
    db.commit()
    return {
        "token": user.api_token,
        "user": {
            "id": user.id,
            "name": user.name,
            "username": user.username,
            "role": user.role,
        },
    }

@app.get("/api/v1/auth/me")
def auth_me(current_user: User = Depends(get_current_user)):
    return {
        "id": current_user.id,
        "name": current_user.name,
        "username": current_user.username,
        "role": current_user.role,
    }

@app.post("/api/v1/auth/logout")
def auth_logout(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    current_user.api_token = None
    db.commit()
    return {"status": "logged_out"}

# ============================================================================
# PREDICTIVE ANALYTICS ENDPOINTS
# Uses the SAME "present" definition as /api/v1/kpis:
#   attendance_status not in ["Absent", "Leave", "Week Off"]
# Stdlib only (statistics, collections) - no new dependencies to install.
# ============================================================================

from collections import defaultdict
import statistics


def _daily_present_totals(db: Session, months_back: int = 6, vendor: Optional[str] = None, store: Optional[str] = None):
    """
    Returns { date: {"present": int, "total": int} } for the lookback window.
    'total' = number of Attendance rows that day (i.e. employees reconciliation ran for).
    """
    cutoff = date.today() - timedelta(days=30 * months_back)
    q = db.query(Attendance.date, Attendance.attendance_status).filter(Attendance.date >= cutoff)
    if vendor:
        q = q.filter(Attendance.vendor == vendor)
    if store:
        q = q.filter(Attendance.store == store)

    daily = defaultdict(lambda: {"present": 0, "total": 0})
    for d, status in q.all():
        daily[d]["total"] += 1
        if status not in ("Absent", "Leave", "Week Off"):
            daily[d]["present"] += 1
    return daily


def _weekday_buckets(daily: Dict[date, Dict[str, int]]) -> Dict[str, list]:
    """Groups the per-day totals into weekday buckets: {"Monday": [(date, pct, present, total), ...]}"""
    buckets = defaultdict(list)
    for d, counts in daily.items():
        if counts["total"] == 0:
            continue
        pct = round(counts["present"] / counts["total"] * 100, 1)
        buckets[d.strftime("%A")].append((d, pct, counts["present"], counts["total"]))
    for wd in buckets:
        buckets[wd].sort(key=lambda x: x[0])  # chronological, oldest first
    return buckets


def _weighted_average(values: List[float]) -> float:
    """Exponential-decay weighted average: most recent value weighted highest."""
    n = len(values)
    if n == 0:
        return 0.0
    if n == 1:
        return round(values[0], 1)
    import math
    weights = [math.exp(-1 + (i / (n - 1))) for i in range(n)]
    total_weight = sum(weights)
    weighted_sum = sum(v * w for v, w in zip(values, weights))
    return round(weighted_sum / total_weight, 1)


def _linear_trend(y: List[float]):
    """Simple least-squares linear regression over evenly spaced x = 0..n-1. Returns (slope, intercept)."""
    n = len(y)
    if n < 2:
        return 0.0, (y[0] if y else 0.0)
    x = list(range(n))
    x_mean = sum(x) / n
    y_mean = sum(y) / n
    numerator = sum((x[i] - x_mean) * (y[i] - y_mean) for i in range(n))
    denominator = sum((x[i] - x_mean) ** 2 for i in range(n))
    slope = numerator / denominator if denominator != 0 else 0.0
    intercept = y_mean - slope * x_mean
    return slope, intercept


@app.get("/api/v1/analytics/day-of-week")
def get_day_of_week_analytics(
    months_back: int = Query(6, ge=1, le=24),
    vendor: Optional[str] = Query(None),
    store: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Historical attendance % by weekday - use this to plan important meetings
    on the days with the most reliable headcount.
    """
    daily = _daily_present_totals(db, months_back, vendor, store)
    buckets = _weekday_buckets(daily)

    weekday_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    result = []
    for wd in weekday_order:
        occurrences = buckets.get(wd, [])
        if not occurrences:
            continue
        pcts = [o[1] for o in occurrences]
        avg_pct = round(sum(pcts) / len(pcts), 1)
        std_dev = round(statistics.pstdev(pcts), 1) if len(pcts) > 1 else 0.0
        avg_present = round(sum(o[2] for o in occurrences) / len(occurrences), 1)
        result.append({
            "weekday": wd,
            "avg_attendance_pct": avg_pct,
            "avg_present_headcount": avg_present,
            "std_dev": std_dev,
            "occurrences_observed": len(occurrences),
            "confidence": "low" if len(occurrences) < 4 else "moderate" if len(occurrences) < 8 else "high"
        })

    result.sort(key=lambda x: x["avg_attendance_pct"], reverse=True)
    return {"months_back": months_back, "vendor": vendor, "store": store, "weekdays": result}


@app.get("/api/v1/analytics/best-meeting-days")
def get_best_meeting_days(
    top_n: int = Query(3, ge=1, le=7),
    months_back: int = Query(6, ge=1, le=24),
    vendor: Optional[str] = Query(None),
    store: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Ranked weekdays safest for scheduling important meetings (excludes near-zero attendance days like weekly-offs)."""
    full = get_day_of_week_analytics(months_back=months_back, vendor=vendor, store=store, db=db)
    candidates = [w for w in full["weekdays"] if w["avg_attendance_pct"] > 5]
    return {"top_n": top_n, "best_days": candidates[:top_n]}


@app.get("/api/v1/analytics/forecast")
def forecast_attendance(
    target_date: str = Query(..., description="YYYY-MM-DD - the date you're planning around"),
    months_back: int = Query(6, ge=1, le=24),
    vendor: Optional[str] = Query(None),
    store: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Weighted forecast for a SPECIFIC upcoming date, using history of that same
    weekday. Recent weeks weighted higher (exponential decay), so the forecast
    shifts automatically as new monthly sheets get ingested.
    """
    target = parse_date_br(target_date)
    if not target:
        raise HTTPException(status_code=400, detail="Invalid target_date. Use YYYY-MM-DD.")

    weekday_name = target.strftime("%A")
    daily = _daily_present_totals(db, months_back, vendor, store)
    buckets = _weekday_buckets(daily)
    occurrences = buckets.get(weekday_name, [])

    if not occurrences:
        return {"target_date": target_date, "weekday": weekday_name, "forecast_pct": None,
                "note": "No historical data for this weekday yet."}

    pcts = [o[1] for o in occurrences]
    n = len(pcts)
    forecast_pct = _weighted_average(pcts)
    historical_avg = round(sum(pcts) / n, 1)

    return {
        "target_date": target_date,
        "weekday": weekday_name,
        "forecast_attendance_pct": forecast_pct,
        "based_on_occurrences": n,
        "historical_avg_pct": historical_avg,
        "trend": "improving" if pcts[-1] > historical_avg else "declining",
        "confidence": "low (build up more weeks of data)" if n < 4 else "moderate" if n < 8 else "high"
    }


@app.get("/api/v1/analytics/ot-trend")
def get_ot_trend_forecast(
    weeks_back: int = Query(8, ge=2, le=52),
    vendor: Optional[str] = Query(None),
    store: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Weekly OT hours trend + next-week forecast (simple linear regression).
    Flags if OT is trending upward - useful to catch cost overruns BEFORE
    the monthly approval sign-off instead of after.
    """
    cutoff = date.today() - timedelta(weeks=weeks_back)
    q = db.query(Attendance.date, Attendance.ot_hours).filter(Attendance.date >= cutoff, Attendance.ot_hours > 0)
    if vendor:
        q = q.filter(Attendance.vendor == vendor)
    if store:
        q = q.filter(Attendance.store == store)

    weekly = defaultdict(float)
    for d, ot in q.all():
        iso_year, iso_week, _ = d.isocalendar()
        weekly[(iso_year, iso_week)] += (ot or 0)

    if len(weekly) < 2:
        return {"weeks_back": weeks_back, "note": "Not enough weeks of OT data yet to compute a trend.", "weekly_totals": []}

    sorted_weeks = sorted(weekly.keys())
    y = [round(weekly[w], 1) for w in sorted_weeks]

    slope, intercept = _linear_trend(y)
    next_week_forecast = round(slope * len(y) + intercept, 1)

    trend_label = "rising" if slope > 1 else "falling" if slope < -1 else "stable"

    return {
        "weeks_back": weeks_back,
        "vendor": vendor,
        "store": store,
        "weekly_totals": [{"iso_year": w[0], "iso_week": w[1], "ot_hours": round(weekly[w], 1)} for w in sorted_weeks],
        "trend": trend_label,
        "slope_hours_per_week": round(slope, 2),
        "next_week_forecast_hours": max(0, next_week_forecast),
        "flag": "OT rising - review staffing before next approval cycle" if slope > 5 else None
    }



# ============================================================================


# ============================================================================
# ADMIN / UTILITY ENDPOINTS
# Use these to fix corrupted states or clean up stale data.
# ============================================================================

@app.post("/api/v1/admin/purge-inactive-employees")
def purge_inactive_employees(db: Session = Depends(get_db)):
    """Permanently delete all employees with status INACTIVE.
    WARNING: This also removes their historical attendance/reconciliation records
    via cascade. Only use this if you are sure inactive employees are truly gone."""
    count = db.query(Employee).filter(Employee.status == "INACTIVE").delete(synchronize_session=False)
    db.commit()
    return {"status": "purged", "employees_deleted": count,
            "message": f"Deleted {count} inactive employees and all their related records."}

@app.post("/api/v1/admin/nuke-employees")
def nuke_all_employees(db: Session = Depends(get_db)):
    """NUCLEAR OPTION: Delete ALL employees from the database.

    Uses raw SQL TRUNCATE with CASCADE to bypass FK constraints.
    Uploads and vendors/stores are preserved.

    WARNING: This cannot be undone."""
    from sqlalchemy import text

    # Step 1: TRUNCATE all derived tables with CASCADE (PostgreSQL) 
    # This is faster than DELETE and handles FK constraints automatically
    tables = [
        "attendance", "attendance_reconciliation", "hr_actions", 
        "overtime", "late_punch_penalties", "behavioral_alerts", 
        "monthly_tata_attendance"
    ]
    for table in tables:
        try:
            db.execute(text(f'TRUNCATE TABLE {table} CASCADE'))
        except Exception:
            # If TRUNCATE fails (e.g. SQLite), fall back to DELETE
            pass
    db.commit()

    # Step 2: Also try DELETE as fallback for SQLite
    try:
        db.query(Attendance).delete(synchronize_session=False)
        db.query(AttendanceReconciliation).delete(synchronize_session=False)
        db.query(HRAction).delete(synchronize_session=False)
        db.query(Overtime).delete(synchronize_session=False)
        db.query(LatePunchPenalty).delete(synchronize_session=False)
        db.query(BehavioralAlert).delete(synchronize_session=False)
        db.query(MonthlyTataAttendance).delete(synchronize_session=False)
        db.commit()
    except Exception:
        db.rollback()

    # Step 3: Delete all employees
    count = db.query(Employee).delete(synchronize_session=False)
    db.commit()

    return {
        "status": "nuked",
        "employees_deleted": count,
        "message": f"Deleted ALL {count} employees and all their derived data. Upload a fresh Master to rebuild the roster."
    }

@app.post("/api/v1/admin/nuke-derived")
def nuke_all_derived(
    month: Optional[str] = Query(None, description="YYYY-MM — wipe only this month. Omit to wipe EVERYTHING derived."),
    db: Session = Depends(get_db)
):
    """NUCLEAR OPTION: Delete ALL derived data (Attendance, Reconciliation, HRAction,
    Overtime, LatePunchPenalty, BehavioralAlert, MonthlyTataAttendance).

    Use this when your dashboard shows phantom data from deleted uploads.
    Master employees and raw ESSL/Tata uploads are preserved.

    Pass month=YYYY-MM to limit the wipe to one month.
    """
    if month:
        year, mon = parse_year_month(month)
        start = date(year, mon, 1)
        end = date(year, mon + 1, 1) if mon < 12 else date(year + 1, 1, 1)
        _clear_derived_data(start, end, db)
        # Also clear monthly summaries
        monthly_del = db.query(MonthlyTataAttendance).filter(MonthlyTataAttendance.year_month == month).delete(synchronize_session=False)
        db.commit()
        return {
            "status": "nuked",
            "scope": month,
            "message": f"All derived data for {month} has been deleted. Re-upload and reconcile fresh."
        }
    else:
        # Wipe EVERYTHING derived
        counts = {
            "attendance": db.query(Attendance).delete(synchronize_session=False),
            "reconciliation": db.query(AttendanceReconciliation).delete(synchronize_session=False),
            "hr_actions": db.query(HRAction).delete(synchronize_session=False),
            "overtime": db.query(Overtime).delete(synchronize_session=False),
            "late_punch_penalties": db.query(LatePunchPenalty).delete(synchronize_session=False),
            "behavioral_alerts": db.query(BehavioralAlert).delete(synchronize_session=False),
            "monthly_tata": db.query(MonthlyTataAttendance).delete(synchronize_session=False),
        }
        db.commit()
        return {
            "status": "nuked",
            "scope": "all",
            "deleted_counts": counts,
            "message": "All derived data wiped. Uploads and Master employees are preserved. Reconcile fresh."
        }

@app.get("/api/v1/admin/upload-orphan-check")
def check_orphan_data(db: Session = Depends(get_db)):
    """Diagnostic: show how many Attendance/Reconciliation/HRAction/Overtime rows exist
    that have NO corresponding raw ESSL or Tata row anymore (orphaned by deletes)."""
    # Find dates where Attendance exists but no ESSL/Tata exists
    att_dates = db.query(Attendance.date).distinct().all()
    orphan_dates = []
    for (d,) in att_dates:
        has_essl = db.query(ESSLAttendance).filter(ESSLAttendance.date == d).first()
        has_tata = db.query(TataAttendance).filter(TataAttendance.date == d).first()
        if not has_essl and not has_tata:
            count = db.query(Attendance).filter(Attendance.date == d).count()
            orphan_dates.append({"date": d.isoformat(), "attendance_rows": count})

    return {
        "orphan_dates_found": len(orphan_dates),
        "total_orphan_attendance_rows": sum(d["attendance_rows"] for d in orphan_dates),
        "dates": orphan_dates[:20],  # first 20
        "recommendation": "Call POST /api/v1/admin/nuke-derived?month=YYYY-MM to clean these up."
    }


# ============================================================================
# RUN
# ============================================================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
    #finish